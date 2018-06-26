from flask_restful import Resource,abort
import os
from flask import Flask, request, redirect, url_for
from werkzeug.utils import secure_filename
import uuid
from Constants.Status import *
from Helpers.DatabaseHelper import DatabaseHelper
from Models.Document import Document
import datetime
import openpyxl as xls
import Helpers.utils as util
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Protection
from openpyxl.utils import get_column_letter
import Helpers.utils as util
import json
import ast
from operator import itemgetter
from datetime import datetime
import time
from Controllers.GenerateReportController import GenerateReportController as report
from Helpers.utils import autheticateTenant
from Helpers.authenticate import *
UPLOAD_FOLDER = './uploads/templates'
ALLOWED_EXTENSIONS = set(['txt', 'xls', 'xlsx'])
class DocumentController(Resource):
    def __init__(self):
        self.domain_info = autheticateTenant()
        if self.domain_info:
            tenant_info = json.loads(self.domain_info)
            self.tenant_info = json.loads(tenant_info['tenant_conn_details'])
            self.db=DatabaseHelper(self.tenant_info)

    @authenticate
    def allowed_file(self,filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
    def get(self, doc_id=None):
        if request.endpoint == 'report_list_ep':
            reporting_date = request.args.get('reporting_date')
            reporting_date_start = request.args.get('reporting_date_start')
            reporting_date_end = request.args.get('reporting_date_end')
            return self.render_report_list(reporting_date=reporting_date,reporting_date_start=reporting_date_start,reporting_date_end=reporting_date_end)
        if(request.endpoint == 'get_date_heads_for_report_ep'):
            startDate = request.args.get('start_date') if request.args.get('start_date') != None else '19000101'
            endDate = request.args.get('end_date') if request.args.get('end_date') != None else '39991231'
            return self.render_data_load_dates(startDate, endDate)
        if request.endpoint == 'drill_down_ep':
            report_id = request.args.get('report_id')
            sheet_id = request.args.get('sheet_id')
            cell_id = request.args.get('cell_id')
            report_snapshot = request.args.get('report_snapshot')
            return self.cell_drill_down_rules(report_id=report_id,sheet_id=sheet_id,cell_id=cell_id,report_snapshot=report_snapshot)
        if request.endpoint == 'drill_down_data_ep':
            drill_kwargs = eval(request.args.get('drill_kwargs'))
            print(drill_kwargs)
            return self.cell_drill_down_data(**drill_kwargs)
        if request.endpoint == 'get_report_template_suggestion_list_ep':
            reports = request.args.get('reports')
            country = request.args.get('country')
            return self.report_template_suggesstion_list(report_id=reports,country=country)
        if request.endpoint == 'get_report_export_to_excel_ep':
            self.report_id = request.args.get('report_id')
            reporting_date = request.args.get('reporting_date')
            cell_format_yn = request.args.get('cell_format_yn')
            if cell_format_yn == None or cell_format_yn == "":
                cell_format_yn = 'N'
            return self.export_to_excel(reporting_date=reporting_date,cell_format_yn=cell_format_yn)
        if request.endpoint == 'get_report_rule_export_to_excel_ep':
            self.report_id = request.args.get('report_id')
            return self.export_rules_to_excel()
        self.report_id = doc_id
        reporting_date = request.args.get('reporting_date')
        print(reporting_date)
        return self.render_report_json(reporting_date)
    def post(self):
        if 'file' not in request.files:
            return NO_FILE_SELECTED
        self.report_id = request.form.get('report_id')
        self.country = request.form.get('country').upper()
        self.report_description = request.form.get('report_description')
        if self.report_id == None or self.report_id == "":
            return REPORT_ID_EMPTY
        if self.country == None or self.country == "":
            return COUNTRY_EMPTY
        file = request.files['file']
        if file and not self.allowed_file(file.filename):
            return FILE_TYPE_IS_NOT_ALLOWED
        filename = str(uuid.uuid4()) + "_" + secure_filename(file.filename)
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        '''document = Document({
            'id': None,
            'file': filename,
            'uploaded_by': 1,
            'time_stamp': str (datetime.datetime.utcnow()),
            'ip': '1.1.1.1',
            'comment': "Sample comment by model"
        })'''
        if self.insert_report_def_catalog():
            self.load_report_template(filename)
            return self.render_report_json()
        else:
            return {"msg: Report capture failed. Please check."}, 400
    def insert_report_def_catalog(self):
        count = self.db.query("select count(*) as count from report_def_catalog where report_id=%s and country=%s",\
                            (self.report_id,self.country,)).fetchone()
        if not count['count']:
            res = self.db.transact("insert into report_def_catalog(report_id,country,report_description) values(%s,%s,%s)",\
                    (self.report_id,self.country,self.report_description,))
            self.db.commit()
            return res
        return 1

    def load_report_template(self,template_file_name):
        formula_dict = {'SUM': 'CALCULATE_FORMULA',
                        '=SUM': 'CALCULATE_FORMULA',
                        }
        cell_render_ref = None
        target_dir = UPLOAD_FOLDER + "/"
        wb = xls.load_workbook(target_dir + template_file_name)


        for sheet in wb.worksheets:

            self.db.transact('delete from report_def where report_id=%s and sheet_id=%s', (self.report_id, sheet.title,))

            # First capture the dimensions of the cells of the sheet
            rowHeights = [sheet.row_dimensions[r + 1].height for r in range(sheet.max_row)]
            colWidths = [sheet.column_dimensions[get_column_letter(c + 1)].width for c in range(sheet.max_column)]

            for row, height in enumerate(rowHeights):
                self.db.transact('insert into report_def(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref)\
                             values(%s,%s,%s,%s,%s)', (self.report_id, sheet.title, str(row + 1), 'ROW_HEIGHT', str(height)))

            for col, width in enumerate(colWidths):
                self.db.transact('insert into report_def(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref)\
                            values(%s,%s,%s,%s,%s)',
                            (self.report_id, sheet.title, get_column_letter(col + 1), 'COLUMN_WIDTH', str(width)))

            rng_startcell = []
            for rng in sheet.merged_cell_ranges:
                # print rng
                startcell, endcell = rng.split(':')
                # print sheet.cell(startcell).border
                rng_startcell.append(startcell)

                self.db.transact('insert into report_def(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref)\
                            values(%s,%s,%s,%s,%s)',
                            (self.report_id, sheet.title, rng, 'MERGED_CELL', sheet[startcell].value))

            for all_obj in sheet['A1':util.cell_index(sheet.max_column, sheet.max_row)]:
                for cell_obj in all_obj:
                    cell_ref = str(cell_obj.column) + str(cell_obj.row)
                    if (len(rng_startcell) > 0 and cell_ref not in rng_startcell) or (len(rng_startcell) == 0):
                        if cell_obj.value:
                            for key in formula_dict.keys():
                                cell_obj_value = str(cell_obj.value)
                                if key in cell_obj_value:
                                    cell_render_ref = formula_dict[key]
                                    break
                                else:
                                    cell_render_ref = 'STATIC_TEXT'

                            self.db.transact('insert into report_def(report_id,sheet_id ,cell_id ,cell_render_def ,cell_calc_ref)\
                                      values(%s,%s,%s,%s,%s)',
                                        (self.report_id, sheet.title, cell_ref, cell_render_ref, cell_obj_value.strip()))
        self.db.commit()
        print('====================================')
        print('End of load report template')
        print('====================================')
        return 0

    def render_report_json(self,reporting_date='19000101',cell_format_yn='Y'):


        cur = self.db.query("select distinct sheet_id from report_def where report_id=%s", (self.report_id,))
        sheets = cur.fetchall()
        print(sheets)

        agg_format_data ={}
        if cell_format_yn == 'Y':
            summary_set = report.create_report_summary_final(self,populate_summary = False,
                            cell_format_yn = cell_format_yn,
                            report_id = self.report_id,
                             business_date_from = str(reporting_date)[:8],
                             business_date_to = str(reporting_date)[8:])
            #print(summary_set)
            for e in summary_set:
                agg_format_data[e[0]+e[1]+e[2]] = e[3]
            #print(agg_format_data)

        sheet_d_list = []
        for sheet in sheets:
            matrix_list = []
            row_attr = {}
            col_attr = {}
            cur = self.db.query(
                "select cell_id,cell_render_def,cell_calc_ref from report_def where report_id=%s and sheet_id=%s",
                (self.report_id, sheet["sheet_id"]))
            report_template = cur.fetchall()

            data = self.db.query('select b.report_id,b.sheet_id,b.cell_id,a.cell_summary,\
                                b.reporting_scale,b.rounding_option \
                                from report_comp_agg_def b left join report_summary a\
                                on a.report_id=b.report_id and\
                                a.sheet_id=b.sheet_id and \
                                a.cell_id=b.cell_id and \
                                a.reporting_date=%s \
                                where b.report_id=%s \
                                and b.sheet_id=%s\
                                order by b.report_id,b.sheet_id,b.cell_id', (reporting_date, self.report_id,sheet["sheet_id"])).fetchall()

            for row in report_template:
                cell_d = {}
                if row["cell_render_def"] == 'STATIC_TEXT':
                    cell_d['cell'] = row['cell_id']
                    cell_d['value'] = row['cell_calc_ref']
                    cell_d['origin'] = "TEMPLATE"
                    matrix_list.append(cell_d)


                elif row['cell_render_def'] == 'MERGED_CELL':
                    start_cell, end_cell = row['cell_id'].split(':')
                    cell_d['cell'] = start_cell
                    cell_d['value'] = row['cell_calc_ref']
                    cell_d['merged'] = end_cell
                    cell_d['origin'] = "TEMPLATE"
                    matrix_list.append(cell_d)


                elif row['cell_render_def'] == 'ROW_HEIGHT':
                    if row['cell_calc_ref'] == 'None':
                        row_height = '12.5'
                    else:
                        row_height = row['cell_calc_ref']
                    row_attr[row['cell_id']] = {'height': row_height}


                elif row['cell_render_def'] == 'COLUMN_WIDTH':
                    if row['cell_calc_ref'] == 'None':
                        col_width = '13.88'
                    else:
                        col_width = row['cell_calc_ref']
                    col_attr[row['cell_id']] = {'width': col_width}

            for row in data:
                cell_d={}
                if cell_format_yn == 'Y':
                    # print(row["cell_id"],row["cell_summary"])
                    try:
                        cell_summary = agg_format_data[row['report_id']+row['sheet_id']+row['cell_id']]
                    except KeyError:
                        cell_summary = util.round_value(
                        float(util.if_null_zero(row["cell_summary"])) / float(row["reporting_scale"]),
                        row["rounding_option"])

                else:
                    cell_summary= float(util.if_null_zero(row["cell_summary"]))

                cell_d['cell']=row['cell_id']
                cell_d['value']=cell_summary
                cell_d['origin'] = "DATA"
                matrix_list.append(cell_d)


            sheet_d = {}
            sheet_d['sheet'] = sheet['sheet_id']
            # print(sheet_d['sheet'])
            sheet_d['matrix'] = matrix_list
            sheet_d['row_attr'] = row_attr
            sheet_d['col_attr'] = col_attr
            sheet_d_list.append(sheet_d)

        json_dump = (sheet_d_list)
        # print(json_dump)
        return json_dump



    def render_data_load_dates(self,start_reporting_date='19000101',end_reporting_date='39991231'):

        month_lookup={ '01': 'January',
                       '02':'February',
                       '03':'March',
                       '04':'April',
                       '05':'May',
                       '06':'June',
                       '07':'July',
                       '08':'August',
                       '09':'Sepember',
                       '10':'October',
                       '11':'November',
                       '12':'December'
                       }

        catalog=self.db.query("select distinct as_of_reporting_date from report_catalog where as_of_reporting_date between "+ start_reporting_date + " and " + end_reporting_date + " order by as_of_reporting_date").fetchall()

        catalog_list=[]

        for cat in catalog:
            year=cat['as_of_reporting_date'][:4]
            month_num=cat['as_of_reporting_date'][4:6]
            report_date=cat['as_of_reporting_date'][6:]
            month=month_lookup[month_num]

            #print(year,month,bus_date)
            #print(list(map(itemgetter('year'),catalog_list)))

            idx=list(map(itemgetter('year'),catalog_list)).index(year)\
                if year in map(itemgetter('year'),catalog_list) else None
            #print(list(map(itemgetter('year'), catalog_list)))
            if idx==None:
                d={'year':year,'month':{month:[report_date]}}
                catalog_list.append(d)
                #print(catalog_list)

            else:
                if month in catalog_list[idx]['month'].keys():
                    catalog_list[idx]['month'][month].append(report_date)
                else:
                    catalog_list[idx]['month'][month]=[report_date]


        return (catalog_list)

    def render_report_list(self,reporting_date=None, reporting_date_start=None, reporting_date_end=None):

        reporting_date = reporting_date if reporting_date and reporting_date != "undefined" else None
        reporting_date_start = reporting_date_start if reporting_date_start and reporting_date_start != "undefined" else None
        reporting_date_end = reporting_date_end if reporting_date_end and reporting_date_end != "undefined" else None

        if (not reporting_date) and ( not reporting_date_start or not reporting_date_end):
            print("Please supply parameters: reporting_date or (reporting_date_start and reporting_date_end)")

        if reporting_date:
            sql = "select * from report_catalog where as_of_reporting_date='"+reporting_date+"'"
        if reporting_date_start and reporting_date_end:
            data_sources={}
            data_sources['start_date']=reporting_date_start
            data_sources['end_date']=reporting_date_end
            sql = "select * from report_catalog where as_of_reporting_date between " \
                  + "'" + reporting_date_start + "' and '" + reporting_date_end + "'"
        reports = self.db.query(sql).fetchall()

        #print(data_sources)
        if reporting_date:
            return (reports)
        else:
            data_sources['data_sources']=reports
            return data_sources


    def export_to_excel(self, **kwargs):
        parameter_list = ['reporting_date','cell_format_yn']
        if set(parameter_list).issubset(set(kwargs.keys())):
            report_id = self.report_id
            reporting_date=kwargs['reporting_date']
            cell_format_yn=kwargs['cell_format_yn']
            target_file_name = report_id+ '_' + str(reporting_date) + '_' + str(time.time())+ '.xlsx'
        else:
            print("Please supply parameters: " + str(parameter_list))

        #Create report template
        wr = xls.Workbook()
        # target_dir='../output/'
        target_dir = './static/'

        sheets=self.db.query('select distinct sheet_id from report_def where report_id=%s', (report_id,)).fetchall()

        agg_format_data ={}
        if cell_format_yn == 'Y':
            summary_set = report.create_report_summary_final(self,populate_summary = False,
                            cell_format_yn = cell_format_yn,
                            report_id = self.report_id,
                             business_date_from = str(reporting_date)[:8],
                             business_date_to = str(reporting_date)[8:])
            #print(summary_set)
            for e in summary_set:
                agg_format_data[e[0]+e[1]+e[2]] = e[3]
            #print(agg_format_data)
        # print sheets
        # The default sheet of the workbook
        al = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
        ws = wr.worksheets[0]
        for sheet in sheets:
            # The first sheet title will be Sheet, so do not create any sheet, just rename the title
            if ws.title == 'Sheet':
                ws.title = sheet["sheet_id"]
            else:
                ws = wr.create_sheet(title=sheet["sheet_id"])

            report_template=self.db.query('select cell_id,cell_render_def,cell_calc_ref from report_def where report_id=%s and sheet_id=%s',
                        (report_id, sheet["sheet_id"],)).fetchall()

            for row in report_template:
                if row["cell_render_def"] == 'MERGED_CELL':
                    ws.merge_cells(row["cell_id"])
                    startcell, endcell = row["cell_id"].split(':')
                    ws[startcell].value = row["cell_calc_ref"]
                    ws[startcell].value = row["cell_calc_ref"]
                    ws[startcell].fill = PatternFill("solid", fgColor="DDDDDD")
                    ws[startcell].font = Font(size=9)
                    ws[startcell].alignment = al
                elif row["cell_render_def"] == 'STATIC_TEXT':
                    ws[row["cell_id"]].value = row["cell_calc_ref"]
                    ws[row["cell_id"]].fill = PatternFill("solid", fgColor="DDDDDD")
                    ws[row["cell_id"]].font = Font(size=9)
                    if row["cell_calc_ref"][:1] != '=' and row["cell_calc_ref"][
                                                           len(row["cell_calc_ref"]) - 1:1] != '%' and 'SUM' not in row[
                        "cell_calc_ref"]:
                        ws[row["cell_id"]].alignment = al
                elif row["cell_render_def"] == 'CALCULATE_FORMULA':
                    ws[row["cell_id"]].fill = PatternFill("solid", fgColor="DDDDDD")
                    ws[row["cell_id"]].font = Font(bold=True, size=9)
                    if '=' in row["cell_calc_ref"]:
                        ws[row["cell_id"]].value = row["cell_calc_ref"]
                    else:
                        ws[row["cell_id"]].value = '=' + row["cell_calc_ref"]
                if row["cell_render_def"] == 'ROW_HEIGHT' and row["cell_calc_ref"] != 'None':
                    # print('row ' + row["cell_id"])
                    ws.row_dimensions[int(row["cell_id"])].height = float(row["cell_calc_ref"])
                elif row["cell_render_def"] == 'COLUMN_WIDTH' and row["cell_calc_ref"] != 'None':
                    # print(row["cell_id"])
                    ws.column_dimensions[row["cell_id"]].width = float(row["cell_calc_ref"])
                else:
                    pass

        wr.save(target_dir + target_file_name)
        #End create report template

        #Create report body
        wb = xls.load_workbook(target_dir + target_file_name)
        data=self.db.query('select b.report_id,b.sheet_id,b.cell_id,a.cell_summary,\
                    b.reporting_scale,b.rounding_option \
                    from report_comp_agg_def b left join report_summary a\
                    on a.report_id=b.report_id and\
                    a.sheet_id=b.sheet_id and \
                    a.cell_id=b.cell_id and \
                    a.reporting_date=%s \
                    where b.report_id=%s \
                    order by b.report_id,b.sheet_id,b.cell_id',(reporting_date,report_id,)).fetchall()

        for row in data:
            ws = wb.get_sheet_by_name(row["sheet_id"])
            if cell_format_yn == 'Y':
                # print(row["cell_id"],row["cell_summary"])
                try:
                    cell_summary = agg_format_data[row['report_id']+row['sheet_id']+row['cell_id']]
                except KeyError:
                    cell_summary = util.round_value(
                    float(util.if_null_zero(row["cell_summary"])) / float(row["reporting_scale"]),
                    row["rounding_option"])
                ws[row["cell_id"]].value = cell_summary
            else:
                ws[row["cell_id"]].value = float(util.if_null_zero(row["cell_summary"]))

            ws[row["cell_id"]].font = Font(size=9)


        wb.save(target_dir + target_file_name)

        #End create report body

        return { "file_name": target_file_name }

    def export_rules_to_excel(self):
        if self.report_id != None and self.report_id != "":
            report_id = self.report_id
            target_file_name = report_id+ '_report_rules' + str(time.time())+ '.xlsx'
        else:
            print("Please supply parameters: report_id")

        #Create report template
        wr = xls.Workbook()
        # target_dir='../output/'
        target_dir = './static/'

        # print sheets
        # The default sheet of the workbook
        al = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
        ws = wr.worksheets[0]
        for sheet in [{'table_name':'report_calc_def'},{'table_name':'report_comp_agg_def'}]:
            # The first sheet title will be Sheet, so do not create any sheet, just rename the title
            if ws.title == 'Sheet':
                ws.title = sheet["table_name"]
            else:
                ws = wr.create_sheet(title=sheet["table_name"])

            cur = self.db.query('select * from ' + sheet["table_name"]+ ' where report_id=%s ',
                        (report_id,))
            report_rules=cur.fetchall()

            cols = []
            for i,c in enumerate(cur.description):
                cols.append({'cell_id': get_column_letter(i + 1) ,'cell_name':c[0]})
                print(cols,cols[i]['cell_id']+'1')
                ws[cols[i]['cell_id']+'1'].value = cols[i]['cell_name']
                ws[cols[i]['cell_id']+'1'].fill = PatternFill("solid", fgColor="DDDDDD")
                ws[cols[i]['cell_id']+'1'].font = Font(bold=True, size=9)
            for row,rule in enumerate(report_rules):
                for c in cols:
                    ws[c['cell_id']+str(row+2)].value = rule[c['cell_name']]
                    ws[cols[i]['cell_id']+str(row+2)].font = Font(bold=True, size=9)

        wr.save(target_dir + target_file_name)
        #End create report template

        return { "file_name": target_file_name }

    def cell_drill_down_rules(self,report_id,sheet_id,cell_id, report_snapshot):

        sql="select cell_calc_ref from report_def where report_id=%s and sheet_id=%s and (cell_id=%s or cell_id like %s) and cell_render_def='COMP_AGG_REF'"
        comp_agg_ref=self.db.query(sql,(report_id,sheet_id,cell_id,cell_id+":%")).fetchone()

        if report_snapshot and report_snapshot!='null' and report_snapshot!='undefined':
            report_snapshot = json.loads(report_snapshot)
        else:
            report_snapshot = None

        if report_snapshot:
            sql ="select cd.* from report_comp_agg_def cd, report_comp_agg_def_vers v " + \
                " where cd.report_id=%s and cd.sheet_id=%s and cd.cell_id=%s " + \
                " and v.version = {} and v.report_id=cd.report_id ".format(report_snapshot['report_comp_agg_def'],) + \
                " and instr(concat(',',v.id_list,','),concat(',',cd.id,','))"
        else:
            sql="select * from report_comp_agg_def where report_id=%s and sheet_id=%s and cell_id=%s and in_use='Y'"

        cell_calc_ref_list = ''
        comp_agg_rules=self.db.query(sql,(report_id,sheet_id,cell_id)).fetchall()
        if comp_agg_rules:
            formula = comp_agg_rules[0]['comp_agg_rule']
            variables = list(set([node.id for node in ast.walk(ast.parse(formula)) if isinstance(node, ast.Name)]))
            cell_calc_ref_list = ','.join(variables)

        agg_rules=[]

        if report_snapshot:
            src_list = '(-999,-999)'
            for src in report_snapshot['report_calc_def'].keys():
                src_list+=",({0},{1})".format(src,report_snapshot['report_calc_def'][src])

        if report_snapshot:
            sql = "select  a.* from report_calc_def a,data_source_information b, report_calc_def_vers v " + \
                " where a.source_id=b.source_id and a.report_id=%s and a.sheet_id=%s and a.cell_id=%s" + \
                " and a.report_id=v.report_id and a.source_id=v.source_id " + \
                " and (v.source_id,v.version) in ({})".format(src_list) + \
                " and instr(concat(',',v.id_list,','),concat(',',a.id,','))"
        else:
            sql = "select  a.* from report_calc_def a,data_source_information b where a.source_id=b.source_id and \
                report_id=%s and sheet_id=%s and cell_id=%s"
        if cell_calc_ref_list != '':
            if report_snapshot:
                sql += " union select  a.* from report_calc_def a,data_source_information b, report_calc_def_vers v " + \
                    " where a.source_id=b.source_id and a.report_id=%s and a.cell_calc_ref in (%s)" + \
                    " and a.report_id=v.report_id and a.source_id=v.source_id " + \
                    " and (v.source_id,v.version) in ({})".format(src_list) + \
                    " and instr(concat(',',v.id_list,','),concat(',',a.id,','))"
            else:
                sql += " union select  a.* from report_calc_def a,data_source_information b where a.source_id=b.source_id and \
                    report_id=%s and cell_calc_ref in (%s)"
            cell_rules = self.db.query(sql, (report_id, sheet_id, cell_id, report_id, cell_calc_ref_list)).fetchall()
        else:
            cell_rules = self.db.query(sql, (report_id, sheet_id, cell_id)).fetchall()


        #print(sql)


        for i,c in enumerate(cell_rules):
            print('Processing index ',i)
            for k,v in c.items():
                if isinstance(v,datetime):
                    c[k] = c[k].isoformat()
                    #print(c[k], type(c[k]))

        display_dict={}

        display_dict['comp_agg_ref']=comp_agg_ref['cell_calc_ref'] if comp_agg_ref else ''
        display_dict['comp_agg_rules']=comp_agg_rules
        display_dict['agg_rules']=agg_rules
        display_dict['cell_rules']=cell_rules
        display_dict['report_snapshot'] = report_snapshot if report_snapshot  else {}

        return display_dict

    def cell_drill_down_data(self,**kwargs):

        parameter_list = ['source_id', 'report_id', 'sheet_id', 'cell_id', 'cell_calc_ref','reporting_date','page']

        print(kwargs)
        if set(parameter_list).issubset(set(kwargs.keys())):
            source_id = kwargs['source_id']
            report_id = kwargs['report_id']
            sheet_id = kwargs['sheet_id']
            cell_id = kwargs['cell_id']
            cell_calc_ref=kwargs['cell_calc_ref']
            reporting_date = kwargs['reporting_date']
            page = kwargs['page']
            version=kwargs['version']
            filter=kwargs['filter']
            if 'export_to_csv' in kwargs.keys():
                export_to_csv = kwargs['export_to_csv']
            else:
                export_to_csv = 'N'
        else:
            print("Please supply parameters: " + str(parameter_list))
            print(kwargs.keys())


        filter_maps = {
            "starts":{"operator": "like", "start_wild_char":"", "end_wild_char":"%"},
            "notstarts":{"operator": "not like", "start_wild_char":"", "end_wild_char":"%"},
            "ends":{"operator": "like", "start_wild_char":"%", "end_wild_char":""},
            "notends":{"operator": "not like", "start_wild_char":"%", "end_wild_char":""},
            "includes":{"operator": "like", "start_wild_char":"%", "end_wild_char":"%"},
            "excludes":{"operator": "not like", "start_wild_char":"%", "end_wild_char":"%"},
            "equals":{"operator": "=", "start_wild_char":"", "end_wild_char":""},
            "notequals":{"operator": "!=", "start_wild_char":"", "end_wild_char":""},
        }

        src_inf=self.db.query("select * from data_source_information where source_id=" + str(source_id)).fetchone()

        key_column = 'id' #util.get_keycolumn(self.db._cursor(), src_inf['source_table_name'])

        self.db.transact("create temporary table tmp_rqd_id_list(idlist bigint)")
        self.db.transact("truncate table tmp_rqd_id_list")
        sql = ("select * from report_qualified_data_link " + \
            " where report_id='{0}' and sheet_id='{1}' and cell_id='{2}' and reporting_date='{3}'" + \
            " and cell_calc_ref='{4}' and version={5}") \
            .format(report_id,sheet_id,cell_id,reporting_date,cell_calc_ref,version)

        rqdldata = self.db.query(sql).fetchone()
        if rqdldata:
            qd_id_list = [(id,) for id in rqdldata['id_list'].split(',')]
            self.db.transactmany("insert into tmp_rqd_id_list(idlist) values (%s)",qd_id_list)

        startPage = int(page) * 100
        data_dict = {}
        filter_sql = ''
        if filter and filter != 'undefined':
            filter=json.loads(filter)
            for col in filter:
                col_filter_sql =''
                conditions = col['value'].split(",")
                for ss in conditions:
                    if ss != '':
                        ss = ss.lstrip().split(":")
                        hint_list = ss[0].split(" ") if len(ss) > 1 else ["and","includes"]
                        print("hint_list values {}".format(hint_list,))
                        hint_list = ["and"] + hint_list if len(hint_list)==1 else hint_list
                        print("hint_list values 2nd list {}".format(hint_list,))
                        hint_join = hint_list[0]
                        hint = hint_list[1] if hint_list[1] in filter_maps.keys() else "includes"
                        c = ss[1] if len(ss) > 1 else ss[0]
                        fm=filter_maps[hint]
                        print("fm values {}".format(fm,))
                        col_filter_sql += ' {0} '.format(hint_join,) if len(col_filter_sql) > 0 else ''
                        col_filter_sql += (col['id'] + ' {0} \'{1}' + c.replace("'","\'") + '{2}\'') \
                                      .format(fm["operator"],fm["start_wild_char"],fm["end_wild_char"])
                filter_sql +="and ({0}) ".format(col_filter_sql)

        if export_to_csv=='Y':
            limit_sql = ''
        else:
            limit_sql = ' limit {0},100'.format(startPage)

        sqlqry = "select {0} from  {1} a, tmp_rqd_id_list b where a.id=b.idlist {2} {3}"

        sql = sqlqry.format( 'a.*', src_inf['source_table_name'] ,filter_sql, limit_sql)
        print(sql)

        cur = self.db.query(sql)
        if export_to_csv=='Y':
            return cur

        data = cur.fetchall()

        cols = [i[0] for i in cur.description]
        print(cols)
        sql = sqlqry.format( 'count(*) as count', src_inf['source_table_name'] ,filter_sql, '')
        print(sql)
        count = self.db.query(sql).fetchone()
        # sql = "select a.* from " + src_inf['source_table_name'] + " a, tmp_rqd_id_list b" + \
        #      " where a." + key_column + "=b.idlist"
        data_dict['cols'] = cols
        data_dict['rows'] = data
        data_dict['count'] = count['count']
        data_dict['table_name'] = src_inf['source_table_name']
        data_dict['cell_details'] = {'report_id': report_id,
                            'sheet_id': sheet_id,
                            'cell_id': cell_id,
                            'reporting_date': reporting_date,
                            'cell_calc_ref': cell_calc_ref,
                            'version': version}

        return data_dict

    def suggesstion_list(self,search_string):

        catalog = self.db.query("select distinct report_id from report_catalog where report_id like '"+search_string+"%'" ).fetchall()

        if not catalog:
            return {"msg":"No report name found starting with '"+search_string+"'"}
        else:
            return catalog
    def report_template_suggesstion_list(self,report_id='ALL',country='ALL'):

        data_dict={}
        where_clause = ''

        sql = "select distinct country from report_def_catalog where 1 "
        country_suggestion = self.db.query(sql).fetchall()
        if country is not None and country !='ALL':
             where_clause =  " and instr('" + country.upper() + "', upper(country)) > 0"
        if report_id is not None and report_id !='ALL':
             where_clause +=  " and instr('" + report_id.upper() + "', upper(report_id)) > 0"

        country = self.db.query(sql + where_clause).fetchall()

        # sql = "select distinct report_id from report_def_catalog"
        # report_suggestion = self.db.query(sql).fetchall()
        data_dict['country'] = country
        for i,c in enumerate(data_dict['country']):
            sql = "select distinct report_id from report_def_catalog where country = '" + c['country'] + "'"
            if report_id is not None and report_id !='ALL':
                 where_clause =  " and instr(upper('" + report_id + "'), upper(report_id)) > 0"
            report = self.db.query(sql + where_clause).fetchall()
            print(data_dict['country'][i])
            data_dict['country'][i]['report'] = report
            where_report = ''
            for j,r in enumerate(data_dict['country'][i]['report']):
                sql = "select distinct report_id, valid_from, valid_to, last_updated_by from report_def where 1 "
                where_report =  " and report_id = '" + data_dict['country'][i]['report'][j]['report_id'] + "'"
                reportversions = self.db.query(sql + where_report).fetchone()
                print(data_dict['country'][i]['report'][j])
                data_dict['country'][i]['report'][j] = reportversions
            print(data_dict)
        #data_dict['report_suggestion'] = report_suggestion
        #data_dict['country_suggestion'] = country_suggestion

        if not data_dict:
            return {"msg":"No report matched found"},404
        else:
            return data_dict['country']
