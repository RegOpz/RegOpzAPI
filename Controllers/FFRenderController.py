from app import *
from flask_restful import Resource,abort
from Helpers.DatabaseHelper import DatabaseHelper
import openpyxl as xls
from openpyxl.utils import column_index_from_string,get_column_letter,coordinate_from_string,coordinate_to_tuple,range_boundaries
import Helpers.utils as util
import json
from Pipeline.PyDAG import DAG
from Helpers.utils import autheticateTenant
from Helpers.authenticate import *
import pandas as pd
import copy
from queue import Queue
from threading import Thread


TOP = 5
BOTTOM = -5
H_INTERSECT = 0
LEFT = 5
RIGHT = -5
V_INTERSECT = 0
INIT = -99
ERROR = -999
NONE=-999



class FFRenderController(Resource):

    def __init__(self):
        self.domain_info = autheticateTenant()
        self.db_master=DatabaseHelper()
        self.rectanguler_empty_spaces=Queue(maxsize=1000)
        self.running_threads=[]
        if self.domain_info:
            tenant_info = json.loads(self.domain_info)
            self.tenant_info = json.loads(tenant_info['tenant_conn_details'])
            self.db=DatabaseHelper(self.tenant_info)

    def get(self,report_id=None):
        if request.endpoint == 'view_free_formt_report_ep':
            self.report_id = report_id
            self.db_object_suffix= request.args.get('domain_type')
            self.reporting_date = request.args.get('reporting_date')
            self.version = request.args.get('version')
            report_parameters = request.args.get('report_parameters')
            report_snapshot = request.args.get('report_snapshot')
            # print(reporting_date)
            if not self.reporting_date:
                return self.render_template_json()

            return self.render_report_json()

    def get_position(self,this_section, other_section):
        this_start_row, this_start_col, this_end_row, this_end_col = this_section
        other_start_row, other_start_col, other_end_row, other_end_col = other_section
        v_pos = INIT
        h_pos = INIT

        if this_start_row >= other_start_row and this_start_row <= other_end_row:
            v_pos = V_INTERSECT
        if this_start_row > other_end_row:
            v_pos = TOP
        if this_end_row < other_start_row:
            v_pos = BOTTOM
        if this_start_col >= other_start_col and this_start_col <= other_end_col:
            h_pos = H_INTERSECT
        if other_end_col < this_start_col:
            h_pos = LEFT
        if this_end_col < other_start_col:
            h_pos = RIGHT

        if v_pos == V_INTERSECT and h_pos == H_INTERSECT:
            h_pos = ERROR
            v_pos = ERROR

        return (h_pos, v_pos)

    def render_template_json(self):

        app.logger.info("Rendering free format report template")
        try:
            def_object = "report_free_format_def"
            ref_object="report_free_format_ref"
            sec_object="report_free_format_section"

            if self.db_object_suffix:
                def_object += "_" + self.db_object_suffix
                ref_object+="_" + self.db_object_suffix
                sec_object+="_"+ self.db_object_suffix
                self.db = self.db_master

            template_df = pd.DataFrame(self.db.query("select a.report_id,a.sheet_id,a.cell_id,a.col_id,a.row_id,a.cell_type,a.cell_ref,\
                                    b.entity_type,b.entity_ref,b.content_type,b.content from {0} a,\
                                    {1} b where a.report_id=%s and a.report_id=b.report_id\
                                    and a.sheet_id=b.sheet_id and a.cell_ref=b.entity_ref".format(ref_object,def_object),(self.report_id,)).fetchall())
            hw_df=pd.DataFrame(self.db.query("select report_id,sheet_id,entity_type,entity_ref,content_type,content from {}\
                                     where report_id=%s and entity_type in ('ROW','COL') and content_type in ('ROW_HEIGHT','COLUMN_WIDTH')\
                                    ".format(def_object),(self.report_id,)).fetchall())
            sec_df=pd.DataFrame(self.db.query("select sheet_id,section_id,section_type,section_range,section_ht_range,section_dependency\
                                     from {} where report_id=%s".format(sec_object),(self.report_id,)).fetchall())

            sheet_content=[]
            for sheet in template_df['sheet_id'].unique():

                #Get row heights
                row_attr=hw_df.loc[(hw_df['sheet_id']==sheet) & (hw_df['entity_type']=='ROW') & (hw_df['content_type']=='ROW_HEIGHT')].to_dict('records')
                row_heights = [None] * len(row_attr)
                for row in row_attr:
                    if row['content'] == "None" or int(float(row['content'])) < 25:
                        row_heights[int(row['entity_ref']) - 1] = 25
                    else:
                        row_heights[int(row['entity_ref']) - 1] = int(float(row['content']))


                #Get column widths
                col_attr = hw_df.loc[(hw_df['sheet_id']==sheet) & (hw_df['entity_type']=='COL') & (hw_df['content_type']=='COLUMN_WIDTH')].to_dict('records')
                col_widths = [None] * len(col_attr)
                for col in col_attr:
                    # Note that column index from openpyxl utils starts at 1, but array starts at 0
                    # col width multiplier is 8 for rendering properly in UX
                    # app.logger.info("col {}".format(col))
                    if col['content'] == "None" or int(float(col['content']) * 8) < 90:
                        col_widths[column_index_from_string(col['entity_ref']) - 1] = 90
                    else:
                        col_widths[column_index_from_string(col['entity_ref']) - 1] = int(float(col['content']) * 8)

                if not sec_df.empty:
                    sec_attr=sec_df.loc[sec_df['sheet_id']==sheet].to_dict('records')
                else:
                    sec_attr={}

                section_details=[]
                # {"ht_range": [0, 0, 0, 1], "range": "A1:B1", "section_id": "s1", "section_position": [], "section_type": "FIXEDFORMAT"}
                for sec in sec_attr:
                    section_details.append({
                        'ht_range':json.loads("["+sec['section_ht_range']+"]"), # To convert string into array of numbers
                        'range':sec['section_range'],
                        'section_id':sec['section_id'],
                        'section_position':[] if sec['section_dependency']=='' else sec['section_dependency'].split(','),
                        'section_type':sec['section_type']
                        })




                data=[[None]*len(col_attr) for row in range(len(row_attr))]
                cell_refs=[[None]*len(col_attr) for row in range(len(row_attr))]
                merged_cells=[]
                sheet_styles = {'style_classes': {}, 'td_styles': []}

                sheet_template=template_df[template_df['sheet_id']==sheet].to_dict('records')

                for t in sheet_template:

                    if t['cell_type']=='MERGED':
                        cell = t['cell_id'].split(':')
                        start_xy = coordinate_from_string(cell[0])
                        # note that openpyxls util provides visual coordinates, but array elements starts with 0
                        start_row = start_xy[1] - 1
                        start_col = column_index_from_string(start_xy[0]) - 1
                        end_xy = coordinate_from_string(cell[1])
                        end_row = end_xy[1] - 1
                        end_col = column_index_from_string(end_xy[0]) - 1
                        merged_cells.append({'row': start_row, 'col': start_col, 'rowspan': end_row - start_row + 1,'colspan': end_col - start_col+1})

                    elif t['cell_type']=='SINGLE':
                        start_xy = coordinate_from_string(t['cell_id'])
                        start_row=start_xy[1]-1
                        start_col=column_index_from_string(start_xy[0]) - 1

                    # Get the reference for the cell
                    cell_refs[start_row][start_col] = t['cell_ref']
                    if t['content_type'] in ('BLANK','DATA',''):
                        data[start_row][start_col] = None
                    elif t['content_type']=='STATIC_TEXT':
                        data[start_row][start_col]=t['content']
                    elif t['content_type'] == 'CELL_STYLE':
                        td_style = json.loads(t['content'])
                        td_class_name = {'classes': ''}
                        util.process_td_class_names(td_style, td_class_name, sheet_styles)
                        sheet_styles['td_styles'].append({'row': start_row, 'col': start_col, 'class_name': td_class_name['classes']})

                sheet_d = {}
                sheet_d['sheet'] = sheet
                sheet_d['sheet_styles'] = sheet_styles
                sheet_d['row_heights'] = row_heights
                sheet_d['col_widths'] = col_widths
                sheet_d['data'] = data
                sheet_d['cell_refs'] = cell_refs
                sheet_d['merged_cells'] = merged_cells
                sheet_d['sections'] = section_details
                sheet_content.append(sheet_d)

            return sheet_content

        except Exception as e:
            app.logger.error(str(e))
            raise

    def render_fixed_format_section(self,wb, sheet, section, sec_df, template_df, hw_df,report_version_no):
        data_df = pd.DataFrame(self.db.query("select cell_id,cell_summary from report_summary where report_id=%s and \
                            reporting_date=%s and sheet_id=%s and section_id=%s and version=%s", (self.report_id, self.reporting_date, \
                                                                                   sheet, section,report_version_no)).fetchall())

        ws = wb.create_sheet(sheet + '-' + section)
        ws_template=wb.create_sheet(sheet+'-'+section+'-template')
        section_params =sec_df.loc[(sec_df['sheet_id'] == sheet) & (sec_df['section_id'] == section)].to_dict(orient='record')[0]
        # app.logger.info("Processing render_fixed_format_section section_params ...[{}]".format(section_params))
        section_start_row, section_start_col, section_end_row, section_end_col = map(int, section_params[
            'section_ht_range'].split(','))
        app.logger.info("Processing render_fixed_format_section boundaries ...{} {} [{}]".format(sheet,section,(section_start_row, section_start_col, section_end_row, section_end_col)))

        for row in range(section_start_row, section_end_row+1):
            # app.logger.info("Processing render_fixed_format_section ...row {}".format(row))
            for col in range(section_start_col, section_end_col+1):
                # app.logger.info("Processing render_fixed_format_section ...col {}".format(col))
                cell_template_lst = template_df.loc[
                    (template_df['report_id'] == self.report_id) & (template_df['sheet_id'] == sheet) & \
                    (template_df['section_id'] == section) & (template_df['row_id'] == (row + 1)) & \
                    (template_df['col_id'] == get_column_letter(col + 1))].to_dict(orient='record')
                if section=="S03":
                    app.logger.info("cross checking cell_template_lst... {}  {}   {}".format(section,row,cell_template_lst))
                if cell_template_lst:
                    cell_template = cell_template_lst[0]
                    if cell_template['cell_type'] == 'MERGED':
                        # app.logger.info("cell_template... {}".format(cell_template))
                        start_cell, end_cell = cell_template['cell_id'].split(':')
                        start_row, start_col = coordinate_to_tuple(start_cell)
                        end_row, end_col = coordinate_to_tuple(end_cell)
                        start_row -= section_start_row
                        end_row -= section_start_row
                        start_col -= section_start_col
                        end_col -= section_start_col
                        # app.logger.info("merged cell_template... {} {} {} {} {} {}".format(start_cell, end_cell,start_row, start_col,end_row, end_col))
                        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
                        ws_template.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
                    elif cell_template['cell_type'] == 'SINGLE':
                        start_row, start_col = coordinate_to_tuple(cell_template['cell_id'])
                        start_row -= section_start_row
                        start_col -= section_start_col

                    for template in cell_template_lst:
                        if row in [7] and col in [0]:
                            app.logger.info("Processing template ...{} {} {}".format(start_row,start_col,template))
                        if template['content_type'] == 'STATIC_TEXT':
                            ws.cell(row=start_row, column=start_col).value = template['content']
                        elif template['content_type'] in ('', 'BLANK', 'DATA'):
                            if not data_df.empty:
                                cell_data = data_df.loc[data_df['cell_id'] == template['cell_ref']].to_dict(orient='record')
                            else:
                                cell_data = None
                            if cell_data:
                                ws.cell(row=start_row, column=start_col).value = cell_data[0]['cell_summary']
                        elif template['content_type']=='CELL_STYLE':
                            td_style = json.loads(template['content'])
                            # app.logger.info("Processing td_style ...{} {}".format(td_style, hw_df))
                            row_height=hw_df.loc[(hw_df['sheet_id']==sheet)&(hw_df['entity_type']=='ROW')&(hw_df['entity_ref']==str(row+1)) &(hw_df['content_type']=='ROW_HEIGHT')].reset_index().at[0,'content']
                            # app.logger.info("Processing row_height ...{} ".format(row_height, ))
                            col_width=hw_df.loc[(hw_df['sheet_id']==sheet)&(hw_df['entity_type']=='COL')&(hw_df['entity_ref']==get_column_letter(col+1)) &(hw_df['content_type']=='COLUMN_WIDTH')].reset_index().at[0,'content']
                            t_content={'row_height':row_height,'col_width':col_width,
                                        'cell_ref':json.dumps({'cell_ref': template['cell_ref'],
                                                    'section':{'section_type': 'FIXEDFORMAT'}
                                                    }),
                                        'cell_style':td_style
                                        }
                            # app.logger.info("Processing t_content ...{}".format(t_content))
                            ws_template.cell(row=start_row,column=start_col).value=json.dumps(t_content)
        # ws.cell(row=start_row+1,column=1).value="Test Row"
        # ws_template.cell(row=start_row+1,column=1).value=json.dumps(t_content)
        # ws.cell(row=start_row+1,column=start_col+1).value="Test Col"
        # ws_template.cell(row=start_row+1,column=start_col+1).value=json.dumps(t_content)
        # ws.cell(row=start_row+2,column=start_col-1 if start_col > 1 else start_col).value="Test Col2"
        # ws_template.cell(row=start_row+2,column=start_col-1 if start_col > 1 else start_col).value=json.dumps(t_content)
        # ws.cell(row=start_row+3,column=1).value="END-OF-TECHNICAL-SECTION-DATA-FOR-INDIVIDUALSHEET"
        # ws_template.cell(row=start_row+3,column=1).value="END-OF-TECHNICAL-SECTION-DATA-FOR-INDIVIDUALSHEET"
        ws.cell(row=start_row+1,column=1).value="END-OF-TECHNICAL-SECTION-DATA-FOR-INDIVIDUALSHEET"
        ws_template.cell(row=start_row+1,column=1).value="END-OF-TECHNICAL-SECTION-DATA-FOR-INDIVIDUALSHEET"
        # ws.cell(row=start_row+1,column=start_col).value="END-OF-TECHNICAL-SECTION-DATA-FOR-INDIVIDUALSHEET"
        # ws_template.cell(row=start_row+1,column=start_col).value="END-OF-TECHNICAL-SECTION-DATA-FOR-INDIVIDUALSHEET"
        app.logger.info("Processing template final ws template...{} {}".format((sheet,section,ws.dimensions,ws['G7'].value,ws['A8'].value),(start_row+1,start_col,ws.max_column,ws.max_row)))

        # app.logger.info("Start and end position {}, {}".format((section_end_row - section_start_row) , (section_end_col - section_start_col)))
        return [(section_end_row - section_start_row) , (section_end_col - section_start_col)]

    def copy_section_to_final(self,ws_dest, ws_dest_template,sec_start_pos, ws_source,ws_source_template):
        sec_start_row, sec_start_col = sec_start_pos
        merged_cells = ws_source.merged_cell_ranges

        for cell_rng in merged_cells:
            # app.logger.info("Processing cell_rng [{}]".format(cell_rng))
            start_col, start_row, end_col, end_row = range_boundaries(cell_rng)
            ws_dest.merge_cells(start_row=start_row + sec_start_row, start_column=start_col + sec_start_col,
                                end_row=end_row + sec_start_row, end_column=end_col + sec_start_col)
            ws_dest_template.merge_cells(start_row=start_row + sec_start_row, start_column=start_col + sec_start_col,
                                end_row=end_row + sec_start_row, end_column=end_col + sec_start_col)

        # for row in ws_source.rows:
        # if ws_source_template.max_row > 6:
        app.logger.info("source template rows count {} row values {}".format(ws_source_template.dimensions, (sec_start_pos,ws_source['G7'].value,ws_source['A8'].value)))
        source_range_start, source_range_end = ws_source_template.dimensions.split(':')
        start_xy = coordinate_to_tuple(source_range_start)
        end_xy = coordinate_to_tuple(source_range_end)
        end_of_section = False
        for row in range(1,end_xy[0]+1):
            # app.logger.info("Processing row [{}]".format(row))
            if end_of_section:
                break
            for col in range(1,end_xy[1]+1):
                content = ws_source_template.cell(row=row,column=col).value
                if content and content == "END-OF-TECHNICAL-SECTION-DATA-FOR-INDIVIDUALSHEET":
                    end_of_section = True
                    break
                app.logger.info("Processing cell [{}] [{}]".format((sec_start_row,row,col,ws_source.max_column),ws_source.cell(row=row,column=col).value))
                cell_row = row
                cell_col = col
                # if cell_row in [7,8,9]:
                #     app.logger.info("Processing cell_col [{}] [{}]".format(cell_row,cell.value))
                ws_dest.cell(row=cell_row + sec_start_row, column=cell_col + sec_start_col).value = ws_source.cell(
                    row=cell_row, column=cell_col).value
                ws_dest_template.cell(row=cell_row + sec_start_row, column=cell_col + sec_start_col).value = ws_source_template.cell(
                    row=cell_row, column=cell_col).value


        app.logger.info("End of Processing cell 0 [{},{}] {}".format(ws_dest['G7'].value,ws_dest['A8'].value,(cell_col + sec_start_col -1, ws_dest.max_column)))
        # app.logger.info("End of Processing cell 1 [{}]".format((ws_source.max_row + sec_start_row, ws_source.max_column + sec_start_col)))
        # return (ws_source.max_row + sec_start_row, column_index_from_string(ws_source.max_column) + sec_start_col)
        # return [ws_source.max_row + sec_start_row -1, ws_source.max_column + sec_start_col -1]
        return [cell_row + sec_start_row -1, cell_col + sec_start_col -1]

    def get_start_pos(self,section_id, sec_df, section_positioning):
        # app.logger.info("get_start_pos 1 {} {}".format(section_positioning,sec_df.loc[sec_df['section_id'] == section_id]))
        section_pos = sec_df.loc[sec_df['section_id'] == section_id].reset_index().at[0, 'section_ht_range'].split(',')
        # app.logger.info("get_start_pos new {}".format(section_pos))
        h_predecessors = section_positioning.loc[section_positioning['section_id'] == section_id].reset_index().at[
            0, 'h_predecessors']
        v_predecessors = section_positioning.loc[section_positioning['section_id'] == section_id].reset_index().at[
            0, 'v_predecessors']
        h_distance = section_positioning.loc[section_positioning['section_id'] == section_id].reset_index().at[0, 'h_distance']
        v_distance = section_positioning.loc[section_positioning['section_id'] == section_id].reset_index().at[0, 'v_distance']
        start_pos = [None,None]

        # app.logger.info("h_predecessors {}".format(h_predecessors))
        if len(h_predecessors.keys())==0:
            start_pos[1] = int(section_pos[1])
        else:
            start_pos[1] = -99
            for prd in h_predecessors.keys():
                prd_end_pos = json.loads(section_positioning.loc[section_positioning['section_id'] == prd].reset_index().at[0, 'end_pos'])
                # app.logger.info("prd_end_pos {} {} {}".format(prd,h_distance[prd],prd_end_pos))
                if not prd_end_pos:
                    raise Exception
                if prd_end_pos[1] + h_distance[prd] > start_pos[1]:
                    start_pos[1] = prd_end_pos[1] + h_distance[prd]

        # app.logger.info("v_predecessors {}".format(v_predecessors))
        if len(v_predecessors.keys())==0:
            start_pos[0] = int(section_pos[0])
        else:
            start_pos[0] = -99
            for prd in v_predecessors.keys():
                prd_end_pos = json.loads(section_positioning.loc[section_positioning['section_id'] == prd].reset_index().at[0, 'end_pos'])
                # app.logger.info("prd_end_pos {}".format(prd_end_pos))
                if not prd_end_pos:
                    raise Exception

                if prd_end_pos[0] + v_distance[prd] > start_pos[0]:
                    start_pos[0] = prd_end_pos[0] + v_distance[prd]

        app.logger.info("start_pos {}".format(start_pos))
        return start_pos

    def render_report_intermediate(self,wb_data,report_version_no,**report_parameters):
        app.logger.info("Rendering free format report to intermediate")
        try:
            if report_parameters:
                self.report_id = report_parameters['report_id']
                self.reporting_date=report_parameters['business_date_from']+report_parameters['business_date_to']

            def_object = "report_free_format_def"
            ref_object = "report_free_format_ref"
            sec_object = "report_free_format_section"

            template_df = pd.DataFrame(self.db.query("select a.report_id,a.sheet_id,a.cell_id,a.col_id,a.row_id,a.cell_type,a.cell_ref,\
                                               a.section_id,b.entity_type,b.entity_ref,b.content_type,b.content from {0} a,\
                                               {1} b where a.report_id=%s and a.report_id=b.report_id\
                                               and a.sheet_id=b.sheet_id and a.cell_ref=b.entity_ref".format(ref_object,
                                                                                                             def_object),
                                                     (self.report_id,)).fetchall())
            hw_df = pd.DataFrame(self.db.query("select report_id,sheet_id,entity_type,entity_ref,content_type,content from {}\
                                                where report_id=%s and entity_type in ('ROW','COL') and content_type in ('ROW_HEIGHT','COLUMN_WIDTH')\
                                               ".format(def_object), (self.report_id,)).fetchall())

            # Move this portion to create report code -Code Start
            sec_df = pd.DataFrame(self.db.query("select sheet_id,section_id,section_type,section_range,section_ht_range,section_dependency\
                                                from {} where report_id=%s".format(sec_object),
                                                (self.report_id,)).fetchall())

            section_dependency = self.create_section_dependency(sec_df)

            for sec in section_dependency:
                app.logger.info("Processing section dependency {}".format((sec['section_dependency'], self.report_id, sec['sheet_id'], sec['section_id'])))
                self.db.transact(
                    "update {} set section_dependency=%s where report_id=%s and sheet_id=%s and section_id=%s".format(
                        sec_object),
                    (sec['section_dependency'], self.report_id, sec['sheet_id'], sec['section_id']))

            self.db.commit()

            # Move this portion to create report code - Code End

            sec_df = pd.DataFrame(self.db.query("select sheet_id,section_id,section_type,section_range,section_ht_range,section_dependency\
                                                            from {} where report_id=%s".format(sec_object),
                                                (self.report_id,)).fetchall())


            sheets=template_df['sheet_id'].unique()
            for sheet in sheets:
                sections=sec_df.loc[sec_df['sheet_id']==sheet]
                app.logger.info("sections for sheet {} are {}".format(sheet, sections))

                sheet_dep_graph=DAG()
                for sec_idx,sec in sections.iterrows():
                    sheet_dep_graph.add_node(sec['section_id'])

                for sec_idx, sec in sections.iterrows():
                    for dep in sec['section_dependency'].split(','):
                        if dep and dep != '':
                            sheet_dep_graph.add_edge(dep,sec['section_id'])

                nodes_in_order=sheet_dep_graph.topological_sort()

                for node in nodes_in_order:
                    section_type=sections.loc[sections['section_id']==node].reset_index().at[0,'section_type']
                    app.logger.info("Processing node ...[{}] [{}]".format(node,section_type))
                    if section_type=='FIXEDFORMAT':
                        self.render_fixed_format_section(wb_data,sheet,node,sec_df,template_df,hw_df,report_version_no)

                section_positioning=pd.DataFrame()
                for node in nodes_in_order:
                    # app.logger.info("Processing node [{}]".format(node))
                    predecessors=sheet_dep_graph.predecessors(node)
                    node_pos=[int(np) for np in sections.loc[sections['section_id']==node].reset_index().at[0,'section_ht_range'].split(',')]
                    h_predecessors={}
                    v_predecessors={}
                    h_distance={}
                    v_distance={}
                    for prd in predecessors:
                        app.logger.info("Processing {} predecessors [{}]".format(node,prd))
                        prd_pos=[int(p) for p in sections.loc[sections['section_id']==prd].reset_index().at[0,'section_ht_range'].split(',')]
                        h_pos,v_pos=self.get_position(node_pos,prd_pos)
                        if h_pos==LEFT:
                            h_predecessors[prd]=prd_pos
                            h_distance[prd]=node_pos[1]-prd_pos[3]
                        elif v_pos==TOP:
                            v_predecessors[prd]=prd_pos
                            v_distance[prd]=node_pos[0]-prd_pos[2]

                    section_positioning=section_positioning.append({'section_id':node,'predecessors':predecessors,\
                                       'h_predecessors':h_predecessors,'v_predecessors':v_predecessors,'h_distance':h_distance,\
                                        'v_distance':v_distance,'start_pos':[],'end_pos':[]},ignore_index=True)

                app.logger.info("{}".format(section_positioning))
                ws_dest=wb_data.create_sheet(sheet)
                ws_dest_template=wb_data.create_sheet(sheet+'-template')
                for node in nodes_in_order:
                    app.logger.info("Processing node in order [{}] {} {}".format(node,wb_data.sheetnames,(sheet+'-'+node,section_positioning)))
                    ws_source=wb_data[sheet+'-'+node]
                    ws_source_template=wb_data[sheet+'-'+node+'-template']
                    start_pos=self.get_start_pos(node,sections,section_positioning)
                    app.logger.info("Processing node in order start_pos [{}]".format(start_pos))
                    end_pos=self.copy_section_to_final(ws_dest,ws_dest_template,start_pos,ws_source,ws_source_template)
                    # app.logger.info("Processing node in order start_pos [{}] end_pos [{}]".format(start_pos,end_pos))
                    # app.logger.info(section_positioning.loc[section_positioning['section_id']==node])
                    section_positioning.loc[section_positioning['section_id']==node,['start_pos']]=json.dumps(start_pos)
                    section_positioning.loc[section_positioning['section_id'] == node,['end_pos']]= json.dumps(end_pos)
                    wb_data.remove(ws_source)
                    wb_data.remove(ws_source_template)

                ws_dest.cell(row=end_pos[0]+2,column=1).value="END-OF-TECHNICAL-SHEET-DATA-FOR-INDIVIDUALSHEET"
                # ws_dest.cell(row=end_pos[0]+2,column=end_pos[1]+1).value="END-OF-TECHNICAL-SHEET-DATA-FOR-INDIVIDUALSHEET"
                ws_dest_template.cell(row=end_pos[0]+2,column=1).value="END-OF-TECHNICAL-SHEET-DATA-FOR-INDIVIDUALSHEET"
                # ws_dest_template.cell(row=end_pos[0]+2,column=end_pos[1]+1).value="END-OF-TECHNICAL-SHEET-DATA-FOR-INDIVIDUALSHEET"
                app.logger.info("end of processing sheet wb {} {}".format(sheet,(end_pos,ws_dest['G7'].value,ws_dest['A8'].value)))
                #empty_zones=self.create_empty_zone_list(ws_dest,section_positioning)


            app.logger.info("end of processing now return wb_data end pos {} section positioning {}".format(end_pos,section_positioning))
            return wb_data

        except Exception as e:
            app.logger.error(str(e))
            raise

    def find_rectangular_empty_zone(self,start_pos,length,scan_start_pos,report_grid,max_dim):
        pass

    def create_empty_zone_list(self,ws,section_positioning):
        report_grid=[[0]*coordinate_from_string(ws.max_column) for x in range(ws.max_row)]
        empty_zone_start=[[(NONE,NONE) for x in range(coordinate_from_string(ws.max_column))] for y in range(ws.max_row)]

        for sec in section_positioning.itertuples():
            section_idx=int(sec.section_id[1:])
            section_start=sec.start_pos
            section_end=sec.end_pos
            for row in range(section_start[0]-1,section_end[0]):
                for col in range(section_start[1]-1,section_end[1]):
                    report_grid[row][col]=section_idx



        for row in range(ws.max_row):
            if row==0:
                previous_section = [(NONE, -1) for x in range(coordinate_from_string(ws.max_column))]  # (section_no,row_no)

            for col in range(coordinate_from_string(ws.max_column)):
                if col==0:
                    previous_col=NONE

                if report_grid[row][col]>0:
                    previous_section[col]=(report_grid[row][col],row)

                if (previous_col==NONE and report_grid[row][col]==0) or (previous_col >0 and report_grid[row][col]==0):
                    #Left side of an empty zone
                    if (previous_section[col][0]==NONE or previous_section[col][0] > 0) and previous_section[col][1]==row-1:
                        #start of an empty zone
                        empty_zone_start[row][col]=(row,col)
                    elif (previous_section[col][0]==NONE or previous_section[col][0] > 0) and previous_section[col][1]<row-1:
                        #continuation of an empty zone into next row
                        empty_zone_start[row][col]=empty_zone_start[row-1][col]
                elif (previous_col==0 and report_grid[row][col]>0) or (previous_col==0 and col==coordinate_from_string(ws.max_column)-1):
                    #right side of an emptyzone
                    if (previous_section[col][0] == NONE or previous_section[col][0] > 0) and previous_section[col][1] == row - 1:
                        #top row of an empty zone
                        if empty_zone_start[row][col-1][0]==row:
                            #first contiguous row of empty zone
                            #make call to the thread
                            thrd=Thread(target=self.find_rectangular_empty_zone,
                                        args=(empty_zone_start[row][col-1],col-empty_zone_start[row][col-1][1],
                                            (row+1,empty_zone_start[row][col-1][1],report_grid,
                                            (ws.max_row,column_index_from_string(ws.max_column)))))
                            thrd.start()
                            self.running_threads.append(thrd)
                elif previous_col >0 and report_grid[row][col] >0:
                    continue
                elif previous_col==0 and report_grid[row][col]==0:
                    empty_zone_start[row][col]=empty_zone_start[row][col-1]
                    continue

                previous_col=report_grid[row][col]


    def render_report_json(self):

        app.logger.info("Rendering free format report to web ")
        try:
            ff_summary_df=pd.DataFrame(self.db.query("select * from report_free_format_summary where report_id=%s and \
            reporting_date=%s and version=%s",(self.report_id,self.reporting_date,self.version)).fetchall())

            app.logger.info("Rendering free format report summary {} {} {}".format(self.report_id,self.reporting_date,self.version))
            sheet_content = []
            if not ff_summary_df.empty:
                for sheet in ff_summary_df['sheet_id'].unique():

                    # Get row heights
                    row_attr = ff_summary_df.loc[(ff_summary_df['sheet_id'] == sheet) & (ff_summary_df['entity_type'] == 'ROW') & (
                    ff_summary_df['cell_type'] == 'ROW_HEIGHT')].to_dict('records')
                    row_heights = [None] * len(row_attr)
                    for row in row_attr:
                        row_heights[int(row['entity_id']) - 1] = int(float(row['content']))

                    # Get column widths
                    col_attr = ff_summary_df.loc[(ff_summary_df['sheet_id'] == sheet) & (ff_summary_df['entity_type'] == 'COL') & (
                    ff_summary_df['cell_type'] == 'COLUMN_WIDTH')].to_dict('records')
                    col_widths = [None] * len(col_attr)
                    for col in col_attr:
                        col_widths[int(col['entity_id']) - 1] = int(float(col['content']))*8

                    data = [[None] * (len(col_attr)) for row in range(len(row_attr))]
                    cell_refs = [[None] * (len(col_attr)) for row in range(len(row_attr))]
                    merged_cells = []
                    sheet_styles = {'style_classes': {}, 'td_styles': []}

                    ff_summary = ff_summary_df[(ff_summary_df['sheet_id'] == sheet)&(ff_summary_df['entity_type'] == 'CELL')].to_dict('records')

                    for t in ff_summary:

                        if t['cell_type'] == 'MERGED':
                            cell = t['entity_id'].split(':')
                            start_xy = coordinate_from_string(cell[0])
                            # note that openpyxls util provides visual coordinates, but array elements starts with 0
                            start_row = start_xy[1] - 1
                            start_col = column_index_from_string(start_xy[0]) - 1
                            end_xy = coordinate_from_string(cell[1])
                            end_row = end_xy[1] - 1
                            end_col = column_index_from_string(end_xy[0]) - 1
                            merged_cells.append({'row': start_row, 'col': start_col, 'rowspan': end_row - start_row + 1,
                                                 'colspan': end_col - start_col + 1})

                        elif t['cell_type'] == 'SINGLE':
                            start_xy = coordinate_from_string(t['entity_id'])
                            start_row = start_xy[1] - 1
                            start_col = column_index_from_string(start_xy[0]) - 1

                        # Get the reference for the cell
                        cell_refs[start_row][start_col] = t['cell_ref']
                        data[start_row][start_col]=t['content']
                        td_style = json.loads(t['cell_style'])
                        td_class_name = {'classes': ''}
                        util.process_td_class_names(td_style, td_class_name, sheet_styles)
                        sheet_styles['td_styles'].append(
                            {'row': start_row, 'col': start_col, 'class_name': td_class_name['classes']})

                    sheet_d = {}
                    sheet_d['sheet'] = sheet
                    sheet_d['sheet_styles'] = sheet_styles
                    sheet_d['row_heights'] = row_heights
                    sheet_d['col_widths'] = col_widths
                    sheet_d['data'] = data
                    sheet_d['cell_refs'] = cell_refs
                    sheet_d['merged_cells'] = merged_cells
                    sheet_d['sections'] = []
                    sheet_content.append(sheet_d)

            return sheet_content

        except Exception as e:
            app.logger.error(str(e))
            raise(e)

    def create_section_dependency(self,sec_df):
        try:
            sec_dep=[]
            for sheet_name,sheet in sec_df.groupby('sheet_id'):
                dep_graph=DAG()
                dep_graph_hor=DAG()
                sec_list = {}
                for sec_name,sec in sheet.iterrows():
                    sec_list[sec['section_id']]=eval('['+sec['section_ht_range']+']')

                section_position={}
                for sec_row in sec_list.keys():
                    sec_row_position={}
                    for sec_col in sec_list.keys():
                        sec_row_position[sec_col]=[INIT,INIT]
                    section_position[sec_row]=sec_row_position

                for this_sec in section_position.keys():
                    this_sec_pos = sec_list[this_sec]
                    for other_sec in section_position[this_sec].keys():
                        if this_sec != other_sec:
                            other_sec_pos=sec_list[other_sec]
                            h_pos,v_pos=self.get_position(this_sec_pos,other_sec_pos)
                            if h_pos==ERROR and v_pos==ERROR:
                                raise Exception
                            section_position[this_sec][other_sec]=[h_pos,v_pos]

                for this_sec in section_position.keys():
                    dep_graph.add_node(this_sec)
                    dep_graph_hor.add_node(this_sec)

                for this_sec in section_position.keys():
                    for other_sec in section_position[this_sec].keys():
                        if this_sec!=other_sec:
                            h_pos,v_pos=section_position[this_sec][other_sec]
                            if v_pos == TOP : #and h_pos==H_INTERSECT:
                                dep_graph.add_edge(other_sec,this_sec)
                            elif v_pos==BOTTOM: # and h_pos==H_INTERSECT:
                                dep_graph.add_edge(this_sec,other_sec)
                            elif h_pos==LEFT: #v_pos==V_INTERSECT and
                                dep_graph_hor.add_edge(other_sec,this_sec)
                            elif h_pos==RIGHT: #v_pos==V_INTERSECT and
                                dep_graph_hor.add_edge(this_sec,other_sec)
                app.logger.info("graphs {} {}".format(dep_graph_hor.all_leaves(),dep_graph.all_leaves()))
                dep_graph=self.prune_graph(dep_graph,dep_graph_hor)

                for sec in sec_list.keys():
                    dep=dep_graph.predecessors(sec)
                    sec_dep.append({'sheet_id':sheet_name,'section_id':sec,'section_dependency':",".join(dep)})
            app.logger.info("Final section def ................ {}".format(sec_dep))
            return sec_dep


        except Exception as e:
            app.logger.error(str(e))
            raise

    def prune_graph(self,dep_graph, dep_graph_hor):
        dep_graph_copy = copy.deepcopy(dep_graph)
        dep_graph_hor_copy = copy.deepcopy(dep_graph_hor)

        dep_graph_levels = {}
        i = 0
        while dep_graph:
            app.logger.info("v graph {} {} {}".format(i,dep_graph.all_leaves(),dep_graph_levels))
            if len(dep_graph.all_leaves()) == 0:
                break
            dep_graph_levels[str(i)] = dep_graph.all_leaves()
            for node in dep_graph.all_leaves():
                dep_graph.delete_node(node)
            i += 1
        max_level = i - 1

        dep_graph_hor_levels = {}
        i = 0
        while dep_graph_hor:
            app.logger.info("h graph {}".format(i))
            if len(dep_graph_hor.all_leaves()) == 0:
                break
            dep_graph_hor_levels[str(i)] = dep_graph_hor.all_leaves()
            for node in dep_graph_hor.all_leaves():
                dep_graph_hor.delete_node(node)
            i += 1
        max_level_hor = i - 1

        for level in range(max_level + 1):
            nodes = dep_graph_levels[str(level)]
            for node in nodes:
                dep_graph.add_node(node)

        for level in range(max_level_hor + 1):
            nodes = dep_graph_hor_levels[str(level)]
            for node in nodes:
                dep_graph.add_node_if_not_exists(node)

        for level in range(max_level):
            app.logger.info("graph level  {}".format(level))
            for node_t in dep_graph_levels[str(level)]:
                app.logger.info("graph vertical level  {}".format(node_t))
                this_node_status = False
                for prev_l in range(level + 1, max_level + 1):
                    if this_node_status == True:
                        break
                    for node_p in dep_graph_levels[str(prev_l)]:
                        if node_p in dep_graph_copy.predecessors(node_t):
                            dep_graph.add_edge(node_p, node_t)
                            this_node_status = True

        for level in range(max_level_hor):
            app.logger.info("graph hor level  {}".format(level))
            for node_t in dep_graph_hor_levels[str(level)]:
                app.logger.info("graph hor level  {}".format(node_t))
                this_node_status = False
                for prev_l in range(level + 1, max_level_hor + 1):
                    if this_node_status == True:
                        break
                    for node_p in dep_graph_levels[str(prev_l)]:
                        if node_p in dep_graph_hor_copy.predecessors(node_t):
                            dep_graph.add_edge(node_p, node_t)
                            this_node_status = True

        return dep_graph
