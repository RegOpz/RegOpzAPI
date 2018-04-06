from app import *
from flask_restful import Resource,abort
from flask import Flask, request, redirect, url_for
from Helpers.DatabaseHelper import DatabaseHelper
import openpyxl as xls
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Protection
import Helpers.utils as util
import time
from Controllers.GenerateReportController import GenerateReportController as report
import json

class ViewReportController(Resource):
    def __init__(self):
        tenant_info = json.loads(request.headers.get('Tenant'))
        self.tenant_info = json.loads(tenant_info['tenant_conn_details'])
        self.db = DatabaseHelper(self.tenant_info)

    def get(self,report_id=None):
        if request.endpoint == 'view_report_ep':
            self.report_id = report_id
            reporting_date = request.args.get('reporting_date')
            # print(reporting_date)
            return self.render_report_json(reporting_date)

        if request.endpoint == 'report_list_ep':
            reporting_date = request.args.get('reporting_date')
            reporting_date_start = request.args.get('reporting_date_start')
            reporting_date_end = request.args.get('reporting_date_end')
            return self.render_report_list(reporting_date=reporting_date,reporting_date_start=reporting_date_start,
                                           reporting_date_end=reporting_date_end)

        if request.endpoint == 'get_report_export_to_excel_ep':
            self.report_id = request.args.get('report_id')
            reporting_date = request.args.get('reporting_date')
            cell_format_yn = request.args.get('cell_format_yn')
            if cell_format_yn == None or cell_format_yn == "":
                cell_format_yn = 'N'
            return self.export_to_excel(reporting_date=reporting_date,cell_format_yn=cell_format_yn)

    def render_report_json(self, reporting_date='19000101', cell_format_yn='Y'):

        app.logger.info("Rendering report")

        try:
            app.logger.info("Getting list of sheet for report {0}".format(self.report_id))
            sheets = self.db.query("select distinct sheet_id from report_def where report_id=%s",
                                   (self.report_id,)).fetchall()

            agg_format_data = {}
            if cell_format_yn == 'Y':
                app.logger.info("Creating formatted summary set")
                summary_set = report.create_report_summary_final(self, populate_summary=False,
                                                                 cell_format_yn=cell_format_yn,
                                                                 report_id=self.report_id,
                                                                 business_date_from=str(reporting_date)[:8],
                                                                 business_date_to=str(reporting_date)[8:])
                # print(summary_set)
                for e in summary_set:
                    agg_format_data[e[0] + e[1] + e[2]] = e[3]
                    # print(agg_format_data)

            sheet_d_list = []
            for sheet in sheets:
                matrix_list = []
                row_attr = {}
                col_attr = {}
                cell_style = {}
                app.logger.info("Getting report definition for report {0},sheet {1}".format(self.report_id,sheet["sheet_id"]))
                report_template = self.db.query(
                    "select cell_id,cell_render_def,cell_calc_ref from report_def where report_id=%s and sheet_id=%s",
                    (self.report_id, sheet["sheet_id"])).fetchall()

                app.logger.info("Getting data for report {0},sheet {1}".format(self.report_id, sheet["sheet_id"]))
                data = self.db.query('select b.report_id,b.sheet_id,b.cell_id,a.cell_summary,\
                                    b.reporting_scale,b.rounding_option \
                                    from report_comp_agg_def b left join report_summary a\
                                    on a.report_id=b.report_id and\
                                    a.sheet_id=b.sheet_id and \
                                    a.cell_id=b.cell_id and \
                                    a.reporting_date=%s \
                                    where b.report_id=%s \
                                    and b.sheet_id=%s\
                                    order by b.report_id,b.sheet_id,b.cell_id',
                                     (reporting_date, self.report_id, sheet["sheet_id"])).fetchall()

                app.logger.info("Writing report definition to dictionary")
                for row in report_template:
                    cell_d = {}
                    if row["cell_render_def"] == 'STATIC_TEXT':
                        cell_d['cell'] = row['cell_id']
                        cell_d['value'] = row['cell_calc_ref']
                        cell_d['origin'] = "TEMPLATE"
                        matrix_list.append(cell_d)


                    elif row['cell_render_def'] == 'MERGED_CELL':
                        start_cell, end_cell = row['cell_id'].split(':')
                        cell_d['cell'] = start_cell
                        cell_d['value'] = row['cell_calc_ref']
                        cell_d['merged'] = end_cell
                        cell_d['origin'] = "TEMPLATE"
                        matrix_list.append(cell_d)


                    elif row['cell_render_def'] == 'ROW_HEIGHT':
                        if row['cell_calc_ref'] == 'None':
                            row_height = '12.5'
                        else:
                            row_height = row['cell_calc_ref']
                        row_attr[row['cell_id']] = {'height': row_height}


                    elif row['cell_render_def'] == 'COLUMN_WIDTH':
                        if row['cell_calc_ref'] == 'None':
                            col_width = '13.88'
                        else:
                            col_width = row['cell_calc_ref']
                        col_attr[row['cell_id']] = {'width': col_width}

                    elif row['cell_render_def'] == 'CELL_STYLE':
                        if ':' in row['cell_id']:
                            start_cell, end_cell = row['cell_id'].split(':')
                        else:
                            start_cell=row['cell_id']

                        app.logger.info("Inside CELL_STYLE for cell {}".format(start_cell,))
                        cell_style[start_cell] = eval(row['cell_calc_ref'])

                if reporting_date == '19000101' or not reporting_date:
                    comp_agg_def = self.db.query("select cell_id,cell_render_def,cell_calc_ref from report_def where \
                    report_id=%s and sheet_id=%s and cell_render_def='COMP_AGG_REF'",
                                                 (self.report_id, sheet["sheet_id"])).fetchall()

                    for row in comp_agg_def:
                        cell_d = {}
                        if ':' in row['cell_id']:
                            start_cell, end_cell = row['cell_id'].split(':')
                        else:
                            start_cell=row['cell_id']

                        current_cell_exists=next((item for item in matrix_list if item['cell']==start_cell),False)
                        if not current_cell_exists:
                            cell_d['cell'] = start_cell
                            cell_d['value'] = '' #row['cell_calc_ref']
                            cell_d['origin'] = "TEMPLATE"
                            #print(cell_d)
                            matrix_list.append(cell_d)
                else:
                    app.logger.info("Writing report data to dictionary")
                    for row in data:
                        cell_d = {}
                        if cell_format_yn == 'Y':
                            # print(row["cell_id"],row["cell_summary"])
                            try:
                                cell_summary = agg_format_data[row['report_id'] + row['sheet_id'] + row['cell_id']]
                            except KeyError:
                                cell_summary = util.round_value(
                                    float(util.if_null_zero(row["cell_summary"])) / float(row["reporting_scale"]),
                                    row["rounding_option"])

                        else:
                            cell_summary = float(util.if_null_zero(row["cell_summary"]))

                        cell_d['cell'] = row['cell_id']
                        cell_d['value'] = cell_summary
                        cell_d['origin'] = "DATA"
                        matrix_list.append(cell_d)


                sheet_d = {}
                sheet_d['sheet'] = sheet['sheet_id']
                # print(sheet_d['sheet'])
                sheet_d['matrix'] = matrix_list
                sheet_d['row_attr'] = row_attr
                sheet_d['col_attr'] = col_attr
                sheet_d['cell_style'] = cell_style
                sheet_d_list.append(sheet_d)


            json_dump = (sheet_d_list)
            # print(json_dump)
            return json_dump
        except Exception as e:
            app.logger.error(e)
            return {"msg": e}, 500

    def render_report_list(self,reporting_date=None, reporting_date_start=None, reporting_date_end=None):
        #db=DatabaseHelper()
        try:
            app.logger.info("Getting list of report")
            if reporting_date:
                app.logger.info("Getting list of report for date {0}".format(reporting_date))
                sql = "select * from report_catalog where as_of_reporting_date='{0}'".format(reporting_date)

            if reporting_date_start and reporting_date_end:
                data_sources={}
                data_sources['start_date']=reporting_date_start
                data_sources['end_date']=reporting_date_end
                app.logger.info("Getting list of rpeort between dates {0} and {1}".format(reporting_date_start ,reporting_date_end))
                sql = "select rc.*,rdc.country,rdc.report_description,rdc.report_type from report_catalog rc, " + \
                    " report_def_catalog rdc where rc.report_id=rdc.report_id " +\
                    " and rc.as_of_reporting_date between '{0}' and '{1}'".format(reporting_date_start ,reporting_date_end)

            reports = self.db.query(sql).fetchall()

            #print(data_sources)
            if reporting_date:
                return (reports)
            else:
                data_sources['data_sources']=reports
                return data_sources
        except Exception as e:
            app.logger.error(e)
            return {"msg": e}, 500

    def export_to_excel(self, reporting_date,cell_format_yn):

        app.logger.info("Exporting report to excel")

        try:
            report_id = self.report_id
            target_file_name = report_id+ '_' + str(reporting_date) + '_' + str(time.time())+ '.xlsx'

            #Create report template
            wr = xls.Workbook()
            # target_dir='../output/'
            target_dir = './static/'
            #db=DatabaseHelper()

            app.logger.info("Getting sheets for thre report")
            sheets=self.db.query('select distinct sheet_id from report_def where report_id=%s', (report_id,)).fetchall()

            agg_format_data ={}
            if cell_format_yn == 'Y':
                app.logger.info("Creating formatted summary set")
                summary_set = report.create_report_summary_final(self,populate_summary = False,
                                cell_format_yn = cell_format_yn,
                                report_id = self.report_id,
                                 business_date_from = str(reporting_date)[:8],
                                 business_date_to = str(reporting_date)[8:])
                #print(summary_set)
                for e in summary_set:
                    agg_format_data[e[0]+e[1]+e[2]] = e[3]
                #print(agg_format_data)
            # print sheets
            # The default sheet of the workbook
            # al = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
            al = Alignment(horizontal="left", vertical="center", wrap_text=True, shrink_to_fit=True)
            ws = wr.worksheets[0]
            # img = xls.drawing.image.Image('/home/deb/Downloads/regopzdata/CloudMargin/SMTB.jpg')
            for sheet in sheets:
                # The first sheet title will be Sheet, so do not create any sheet, just rename the title
                if ws.title == 'Sheet':
                    ws.title = sheet["sheet_id"]
                else:
                    ws = wr.create_sheet(title=sheet["sheet_id"])

                app.logger.info("Getting report template information for sheet {}".format(sheet["sheet_id"]))
                report_template=self.db.query('select cell_id,cell_render_def,cell_calc_ref from report_def where report_id=%s and sheet_id=%s',
                            (report_id, sheet["sheet_id"],)).fetchall()

                app.logger.info("Writing report template to file")
                for row in report_template:
                    if row["cell_render_def"] == 'MERGED_CELL':
                        ws.merge_cells(row["cell_id"])
                        startcell, endcell = row["cell_id"].split(':')
                        ws[startcell].value = row["cell_calc_ref"]
                        ws[startcell].value = row["cell_calc_ref"]
                        # ws[startcell].fill = PatternFill("solid", fgColor="DDDDDD")
                        # ws[startcell].font = Font(size=9)
                        # ws[startcell].font = Font(bold=True, size=9)
                        # ws[startcell].alignment = al
                    elif row["cell_render_def"] == 'STATIC_TEXT':
                        ws[row["cell_id"]].value = row["cell_calc_ref"]
                        # ws[row["cell_id"]].fill = PatternFill("solid", fgColor="DDDDDD")
                        # ws[row["cell_id"]].font = Font(size=9)
                        # ws[row["cell_id"]].font = Font(bold=True, size=9)
                        # if row["cell_calc_ref"][:1] != '=' and row["cell_calc_ref"][
                                                            #    len(row["cell_calc_ref"]) - 1:1] != '%' and 'SUM' not in row[
                            # "cell_calc_ref"]:
                            # ws[row["cell_id"]].alignment = al
                    elif row["cell_render_def"] == 'CALCULATE_FORMULA':
                        # ws[row["cell_id"]].fill = PatternFill("solid", fgColor="DDDDDD")
                        # ws[row["cell_id"]].font = Font(bold=True, size=9)
                        if '=' in row["cell_calc_ref"]:
                            ws[row["cell_id"]].value = row["cell_calc_ref"]
                        else:
                            ws[row["cell_id"]].value = '=' + row["cell_calc_ref"]
                    if row["cell_render_def"] == 'ROW_HEIGHT' and row["cell_calc_ref"] != 'None':
                        # print('row ' + row["cell_id"])
                        ws.row_dimensions[int(row["cell_id"])].height = float(row["cell_calc_ref"])
                    elif row["cell_render_def"] == 'COLUMN_WIDTH' and row["cell_calc_ref"] != 'None':
                        # print(row["cell_id"])
                        ws.column_dimensions[row["cell_id"]].width = float(row["cell_calc_ref"])
                    elif row["cell_render_def"] == 'CELL_STYLE':
                        # print(row["cell_id"])
                        if ':' in row["cell_id"]:
                            startcell, endcell = row["cell_id"].split(':')
                        else:
                            startcell = row["cell_id"]
                        _cell_style=eval(row['cell_calc_ref'])
                        app.logger.info("{0} CELL_STYLE values are {1} {2}".format(row['cell_id'],_cell_style,row['cell_calc_ref']))
                        _al = Alignment(horizontal=_cell_style['alignment']['horizontal'], vertical=_cell_style['alignment']['vertical'])
                        _font = Font(name=_cell_style['font']['name'],
                                    bold= _cell_style['font']['bold'],
                                    italic= _cell_style['font']['italic'],
                                    color= _cell_style['font']['colour'] if _cell_style['font']['colour'] != 'None' else None,
                                    size= _cell_style['font']['size'])
                        _fill = PatternFill(fill_type=_cell_style['fill']['type'],
                                    fgColor=_cell_style['fill']['colour'] if _cell_style['fill']['colour'] != 'None' else None)
                        _border = Border(left=Side(style= _cell_style['border']['left']['style'],
                                                    color= _cell_style['border']['left']['colour'] if _cell_style['border']['left']['colour'] != 'None' else None),
                                        right=Side(style= _cell_style['border']['right']['style'],
                                                 color= _cell_style['border']['right']['colour'] if _cell_style['border']['right']['colour'] != 'None' else None),
                                        top=Side(style= _cell_style['border']['top']['style'],
                                                 color= _cell_style['border']['top']['colour'] if _cell_style['border']['top']['colour'] != 'None' else None),
                                        bottom=Side(style= _cell_style['border']['bottom']['style'],
                                                 color= _cell_style['border']['bottom']['colour'] if _cell_style['border']['bottom']['colour'] != 'None' else None))
                        ws[startcell].alignment = _al
                        ws[startcell].font = _font
                        ws[startcell].fill = _fill
                        ws[startcell].border = _border
                    else:
                        pass
            app.logger.info("Saving report template to file {}".format(target_dir + target_file_name))
            wr.save(target_dir + target_file_name)
            #End create report template

            #Create report body
            wb = xls.load_workbook(target_dir + target_file_name)
            app.logger.info("Getting report data")
            data=self.db.query('select b.report_id,b.sheet_id,b.cell_id,a.cell_summary,\
                        b.reporting_scale,b.rounding_option \
                        from report_comp_agg_def b left join report_summary a\
                        on a.report_id=b.report_id and\
                        a.sheet_id=b.sheet_id and \
                        a.cell_id=b.cell_id and \
                        a.reporting_date=%s \
                        where b.report_id=%s \
                        order by b.report_id,b.sheet_id,b.cell_id',(reporting_date,report_id,)).fetchall()

            app.logger.info("Writing report data")
            # for sheet in sheets:
                # ws = wb.get_sheet_by_name(sheet["sheet_id"])
                # img = xls.drawing.image.Image('/home/deb/Downloads/regopzdata/CloudMargin/SMTB.2.1.jpg')
                # # img.anchor(ws.cell('T1'))
                # ws.add_image(img,'N1')
            for row in data:
                ws = wb.get_sheet_by_name(row["sheet_id"])
                if cell_format_yn == 'Y':
                    # print(row["cell_id"],row["cell_summary"])
                    try:
                        cell_summary = agg_format_data[row['report_id']+row['sheet_id']+row['cell_id']]
                    except KeyError:
                        cell_summary = util.round_value(
                        float(util.if_null_zero(row["cell_summary"])) / float(row["reporting_scale"]),
                        row["rounding_option"])
                    ws[row["cell_id"]].value = cell_summary
                else:
                    ws[row["cell_id"]].value = float(util.if_null_zero(row["cell_summary"]))

                ws[row["cell_id"]].font = Font(size=9)

            app.logger.info("Saving report to file {}".format(target_dir + target_file_name))
            wb.save(target_dir + target_file_name)

            #End create report body

            return { "file_name": target_file_name }
        except Exception as e:
            app.logger.error(e)
            return {"msg": e}, 500
