from app import *
from flask_restful import Resource,abort
import os
from flask import Flask, request, redirect, url_for
from werkzeug.utils import secure_filename
import uuid
from Constants.Status import *
from Helpers.DatabaseHelper import DatabaseHelper
from Models.Document import Document
import datetime
import openpyxl as xls
import Helpers.utils as util
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Protection
from openpyxl.utils import column_index_from_string,get_column_letter,coordinate_from_string
import Helpers.utils as util
import json
import ast
from operator import itemgetter
from datetime import datetime
import time
import math
import re
from Pipeline.PyDAG import DAG
from Helpers.utils import autheticateTenant
from Helpers.authenticate import *

UPLOAD_FOLDER = './uploads/templates'
ALLOWED_EXTENSIONS = set(['txt', 'xls', 'xlsx'])

class FreeFormatReportController(Resource):
    def __init__(self):
        self.domain_info = autheticateTenant()
        self.db_master=DatabaseHelper()
        if self.domain_info:
            tenant_info = json.loads(self.domain_info)
            self.tenant_info = json.loads(tenant_info['tenant_conn_details'])
            self.db=DatabaseHelper(self.tenant_info)

    @authenticate
    def get(self,report_id=None):
        if request.endpoint == 'view_free_formt_report_ep':
            self.report_id = report_id
            self.db_object_suffix= request.args.get('domain_type')
            reporting_date = request.args.get('reporting_date')
            version = request.args.get('version')
            report_parameters = request.args.get('report_parameters')
            report_snapshot = request.args.get('report_snapshot')
            # print(reporting_date)
            return self.render_report_json(reporting_date=reporting_date,cell_format_yn='Y',version=version,report_snapshot=report_snapshot,report_parameters=report_parameters)

    def post(self):
        return self.capture_template()

    def put(self,report_id=None):
        if request.endpoint == 'validate_section_free_formt_report_ep':
            sections = request.get_json(force=True)
            self.DAG = DAG()
            return self.validate_sec_dependency(sections=sections)
        else:
            self.report_id = report_id
            self.db_object_suffix= request.args.get('domain_type')
            report_data = request.get_json(force=True)
            return self.save_hot_table_report_template(report_data)

    def capture_template(self):
        try:
            app.logger.info(request)
            if 'file' not in request.files:
                return NO_FILE_SELECTED

            self.report_id = request.form.get('report_id')
            self.report_type = request.form.get('report_type').upper()
            self.country = request.form.get('country').upper()
            self.report_description = request.form.get('report_description')
            self.db_object_suffix= request.form.get('domain_type')

            if self.report_id == None or self.report_id == "":
                return REPORT_ID_EMPTY

            if self.country == None or self.country == "":
                return COUNTRY_EMPTY

            file = request.files['file']
            self.selected_file=file.filename
            if file and not self.allowed_file(file.filename):
                return FILE_TYPE_IS_NOT_ALLOWED
            filename = str(uuid.uuid4()) + "_" + secure_filename(file.filename)
            file.save(os.path.join(UPLOAD_FOLDER, filename))
            res = self.insert_report_def_catalog()
            if res:
                return self.load_report_template(filename)

            else:
                return res
        except Exception as e:
            app.logger.error(str(e))
            return {"msg":str(e)},500

    def allowed_file(self,filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

    def insert_report_def_catalog(self):
        app.logger.info("Creating entry for report catalog")
        try:
            catalog_object = "report_def_catalog"
            if self.db_object_suffix:
                catalog_object += "_" + self.db_object_suffix
            app.logger.info("Checking if report {} for country {} already exists in catalog".format(self.report_id,self.country))
            count = self.db.query("select count(*) as count from {} where report_id=%s and country=%s".format(catalog_object,),\
                                (self.report_id,self.country,)).fetchone()
            app.logger.info("Creating catalog entry for country {} and report {}".format(self.country,self.report_id))
            if not count['count']:
                res = self.db.transact("insert into {}(report_id,report_type,country,report_description) values(%s,%s,%s,%s)".format(catalog_object,),\
                        (self.report_id,self.report_type,self.country,self.report_description,))
                # self.db.commit()
                return res
            return 1
        except Exception as e:
            app.logger.error(str(e))
            return {"msg":str(e)},500

    def load_report_template(self,template_file_name):
        app.logger.info("Loading report template")
        try:
            def_object = "report_free_format_def"
            if self.db_object_suffix:
                def_object += "_" + self.db_object_suffix

            formula_dict = {'SUM': 'CALCULATE_FORMULA',
                            '=SUM': 'CALCULATE_FORMULA',
                            }
            cell_render_ref = None
            target_dir = UPLOAD_FOLDER + "/"
            app.logger.info("Loading {} file from {} directory".format(template_file_name,target_dir))
            wb = xls.load_workbook(target_dir + template_file_name)

            #db = DatabaseHelper()
            app.logger.info("Deleting definition entries for report {}".format(self.report_id,))
            self.db.transact('delete from {} where report_id=%s'.format(def_object,), (self.report_id,))

            sheet_index=0
            for sheet in wb.worksheets:
                sheet_index+=1
                app.logger.info("Processing definition entries for sheet {} ,report {}".format(sheet.title,self.report_id))

                # First capture the dimensions of the cells of the sheet
                rowHeights = [sheet.row_dimensions[r + 1].height for r in range(sheet.max_row)]
                colWidths = [sheet.column_dimensions[get_column_letter(c + 1)].width for c in range(sheet.max_column)]

                app.logger.info("Creating entries for row height")
                for row, height in enumerate(rowHeights):
                    self.db.transact(('insert into {}(report_id,sheet_id,cell_id,row_id,cell_render_def,cell_calc_ref) '+\
                                 'values(%s,%s,%s,%s,%s,%s)').format(def_object,),
                                 (self.report_id, sheet.title, str(row + 1),(row + 1), 'ROW_HEIGHT', str(height)))

                app.logger.info("Creating entries for column width")
                for col, width in enumerate(colWidths):
                    self.db.transact(('insert into {}(report_id,sheet_id,cell_id,col_id,cell_render_def,cell_calc_ref) '+\
                                'values(%s,%s,%s,%s,%s,%s)').format(def_object,),
                                (self.report_id, sheet.title, get_column_letter(col + 1), get_column_letter(col + 1), 'COLUMN_WIDTH', str(width)))

                app.logger.info("Creating entries for merged cells")
                rng_startcell = []
                rng_boundary=[]
                for rng in sheet.merged_cell_ranges:
                    # print rng
                    startcell, endcell = rng.split(':')
                    # print sheet.cell(startcell).border
                    rng_startcell.append(startcell)
                    rng_boundary.append(rng)
                    start_xy = coordinate_from_string(startcell)
                    start_row = start_xy[1]
                    start_col = start_xy[0]
                    agg_ref='S'+str(sheet_index)+'AGG'+str(startcell)
                    _cell=sheet[startcell]
                    cell_style=util.get_css_style_from_openpyxl(_cell)

                    self.db.transact(('insert into {}(report_id,sheet_id,cell_id,col_id,row_id,cell_render_def,cell_calc_ref) '+\
                                'values(%s,%s,%s,%s,%s,%s,%s)').format(def_object,),
                                (self.report_id, sheet.title, rng, start_col, start_row,'MERGED_CELL', sheet[startcell].value))
                    self.db.transact(('insert into {}(report_id,sheet_id,cell_id,col_id,row_id,cell_render_def,cell_calc_ref) '+\
                                'values(%s,%s,%s,%s,%s,%s,%s)').format(def_object,),
                                (self.report_id, sheet.title, rng, start_col, start_row, 'COMP_AGG_REF', agg_ref))
                    self.db.transact(('insert into {}(report_id,sheet_id,cell_id,col_id,row_id,cell_render_def,cell_calc_ref) '+\
                                'values(%s,%s,%s,%s,%s,%s,%s)').format(def_object,),
                                (self.report_id, sheet.title, rng, start_col, start_row, 'CELL_STYLE', json.dumps(cell_style)))

                app.logger.info("Creating entries for static text and formulas")
                for all_obj in sheet['A1':util.cell_index(sheet.max_column, sheet.max_row)]:
                    for cell_obj in all_obj:
                        cell_ref = str(cell_obj.column) + str(cell_obj.row)
                        row = cell_obj.row
                        col = str(cell_obj.column)
                        agg_ref='S'+str(sheet_index)+'AGG'+str(cell_ref)
                        _cell=sheet[cell_ref]
                        cell_style=util.get_css_style_from_openpyxl(_cell)
                        if (len(rng_startcell) > 0 and cell_ref not in rng_startcell) or (len(rng_startcell) == 0):
                            if cell_obj.value and cell_obj.value.replace(' ','') !='':
                                for key in formula_dict.keys():
                                    cell_obj_value = str(cell_obj.value)
                                    if key in cell_obj_value:
                                        cell_render_ref = formula_dict[key]
                                        break
                                    else:
                                        cell_render_ref = 'STATIC_TEXT'

                                self.db.transact(('insert into {}'+\
                                        '(report_id,sheet_id ,cell_id ,col_id, row_id, cell_render_def ,cell_calc_ref) '+\
                                         'values(%s,%s,%s,%s,%s,%s,%s)').format(def_object,),
                                            (self.report_id, sheet.title, cell_ref, col, row, cell_render_ref, cell_obj_value.strip()))

                            if not self.check_if_cell_isin_range(cell_ref,rng_boundary):
                                self.db.transact(('insert into {}'+\
                                                '(report_id,sheet_id,cell_id,col_id,row_id,cell_render_def,cell_calc_ref) '+\
                                                 'values(%s,%s,%s,%s,%s,%s,%s)').format(def_object,),
                                                 (self.report_id, sheet.title, cell_ref, col, row, 'COMP_AGG_REF', agg_ref))
                                self.db.transact(('insert into {}'+\
                                                '(report_id,sheet_id,cell_id,col_id,row_id,cell_render_def,cell_calc_ref) '+\
                                                 'values(%s,%s,%s,%s,%s,%s,%s)').format(def_object,),
                                                 (self.report_id, sheet.title, cell_ref,col,row, 'CELL_STYLE', json.dumps(cell_style)))

            self.db.commit()
            return {"msg": "Report [" + self.report_id + "] template has been captured successfully using " + self.selected_file}, 200
        except Exception as e:
            app.logger.error(str(e))
            return {"msg": str(e)}, 500

    def check_if_cell_isin_range(self,cell_id,rng_boundary_list):

        def number_from_word(value):
            base='ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            result=0
            i=len(value)-1
            for j in range(0,len(value)):
              result+=math.pow(len(base),i)*(base.index(value[j])+1)
              i-=1
            return result

        def split_numbers_chars(string):
            sub=re.split('(\d+)', string)
            return sub

        def check_inclusion(cell,rng):
            start_cell, end_cell = rng.split(':')
            start_cell_sub=split_numbers_chars(start_cell)
            end_cell_sub=split_numbers_chars(end_cell)
            cell_sub=split_numbers_chars(cell)

            within_rng=False

            if number_from_word(cell_sub[0]) >=number_from_word(start_cell_sub[0]) \
            and number_from_word(cell_sub[0]) <= number_from_word(end_cell_sub[0]) \
            and  cell_sub[1] >= start_cell_sub[1] and cell_sub[1] <= end_cell_sub[1]:
                within_rng=True

            return within_rng

        incl=False
        for rng in rng_boundary_list:
            incl=check_inclusion(cell_id,rng)
            if incl:
                break
        return incl

    def report_template_check(self,report_id,country=None, domain_type=""):
        try:
            app.logger.info("Report ID inside the check function {}".format(report_id,))
            if domain_type and domain_type != 'undefined' and domain_type=="master":
                domain_type = "_" + domain_type
            else:
                domain_type = ""
            sql = "select report_id from report_def_catalog{} where report_id='{}'".format(domain_type,report_id)
            if country and country != 'undefined' and country != 'null':
                sql += " and country = '{}'".format(country)
            report = self.db.query(sql).fetchall()

            if report:
                return { "msg": "Report ID already exists.", "donotUseMiddleWare": True },200
            else:
                return {}, 200

        except Exception as e:
            app.logger.error(e)
            return {"msg": e}, 500

    def render_report_json(self, reporting_date='19000101', cell_format_yn='Y',version=1,report_snapshot="{}",report_parameters="{}"):

        app.logger.info("Rendering free format report")

        try:
            app.logger.info("Getting list of sheet for report {0} version {1}".format(self.report_id,version))
            def_object = "report_free_format_def"
            if self.db_object_suffix:
                def_object += "_" + self.db_object_suffix
            sheets = self.db.query("select distinct sheet_id from {} where report_id=%s".format(def_object,),
                                   (self.report_id,)).fetchall()

            sheet_d_list = []
            for sheet in sheets:
                # Get row heights
                row_attr = self.db.query(
                    ("select cell_id,cell_calc_ref from {} " + \
                    "where report_id=%s and sheet_id=%s and cell_render_def = 'ROW_HEIGHT' ").format(def_object,),
                    (self.report_id, sheet["sheet_id"])).fetchall()
                row_count = len(row_attr)
                row_heights = [None]*row_count
                for row in row_attr:
                    if row['cell_calc_ref']=="None" or int(float(row['cell_calc_ref'])) < 25:
                        row_heights[int(row['cell_id'])-1] = 25
                    else:
                        row_heights[int(row['cell_id'])-1] = int(float(row['cell_calc_ref']))

                # Get Column widths
                col_attr = self.db.query(
                    ("select cell_id,cell_calc_ref from {} " + \
                    "where report_id=%s and sheet_id=%s and cell_render_def = 'COLUMN_WIDTH' ").format(def_object,),
                    (self.report_id, sheet["sheet_id"])).fetchall()
                col_count = len(col_attr)
                col_widths = [None]*col_count
                for col in col_attr:
                    # Note that column index from openpyxl utils starts at 1, but array starts at 0
                    # col width multiplier is 8 for rendering properly in UX
                    # app.logger.info("col {}".format(col))
                    if col['cell_calc_ref']=="None" or int(float(col['cell_calc_ref'])*8) <90:
                        col_widths[column_index_from_string(col['cell_id'])-1] = 90
                    else:
                        col_widths[column_index_from_string(col['cell_id'])-1] =  int(float(col['cell_calc_ref'])*8)


                # Get the data for handsontable
                # Create placeholders for the data grid row_count x col_count
                data = [[None]*col_count for x in [None]*row_count]
                merged_cells=[]
                template = self.db.query(
                    ("select cell_id,cell_render_def,cell_calc_ref from {} " + \
                    "where report_id=%s and sheet_id=%s and cell_render_def in ('STATIC_TEXT','MERGED_CELL')").format(def_object,),
                    (self.report_id, sheet["sheet_id"])).fetchall()
                for t in template:
                    cell = t['cell_id'].split(':') # split it as we might have merged cells represented as A1:V10
                    start_xy = coordinate_from_string(cell[0])
                    # note that openpyxls util provides visual coordinates, but array elements starts with 0
                    start_row = start_xy[1]-1
                    start_col = column_index_from_string(start_xy[0])-1
                    if len(cell)==2:
                        end_xy = coordinate_from_string(cell[1])
                        end_row = end_xy[1]-1
                        end_col = column_index_from_string(end_xy[0])-1
                        merged_cells.append({'row': start_row, 'col': start_col,'rowspan': end_row - start_row+1, 'colspan': end_col - start_col+1})

                    data[start_row][start_col] = t['cell_calc_ref']

                # Create the stylessheet for the report
                sheet_styles={'style_classes':{}, 'td_styles':[]}
                styles = self.db.query(
                    ("select cell_id,cell_render_def,cell_calc_ref from {} " + \
                    "where report_id=%s and sheet_id=%s and cell_render_def in ('CELL_STYLE')").format(def_object,),
                    (self.report_id, sheet["sheet_id"])).fetchall()
                for s in styles:
                    td_style = json.loads(s['cell_calc_ref'])
                    cell = s['cell_id'].split(':') # split it as we might have merged cells represented as A1:V10
                    start_xy = coordinate_from_string(cell[0])
                    # note that openpyxls util provides visual coordinates, but array elements starts with 0
                    start_row = start_xy[1]-1
                    start_col = column_index_from_string(start_xy[0])-1
                    td_class_name = {'classes':''}
                    util.process_td_class_names(td_style,td_class_name, sheet_styles)
                    sheet_styles['td_styles'].append({'row': start_row, 'col': start_col, 'class_name': td_class_name['classes']})

                sections = self.db.query(
                    ("select cell_id,cell_render_def,cell_calc_ref from {} " + \
                    "where report_id=%s and sheet_id=%s and cell_render_def = 'SECTION_DEF'").format(def_object,),
                    (self.report_id, sheet["sheet_id"])).fetchall()

                section_details = []
                for sec in sections:
                    sec_def = json.loads(sec['cell_calc_ref'])
                    # cell = sec['cell_id'].split(':') # split it as we might have merged cells represented as A1:V10
                    # start_xy = coordinate_from_string(cell[0])
                    # # note that openpyxls util provides visual coordinates, but array elements starts with 0
                    # start_row = start_xy[1]-1
                    # start_col = column_index_from_string(start_xy[0])-1
                    #
                    # end_xy = coordinate_from_string(cell[1])
                    # end_row = end_xy[1]-1
                    # end_col = column_index_from_string(end_xy[1])-1
                    # # Add hot table cell range [startrow,startcol,endrow,endcol]
                    # sec_def['ht_cell_range'] = [start_row, start_col, end_row, end_col]
                    # sec_def['excel_cell_range'] = cell

                    section_details.append(sec_def)

                sheet_d = {}
                sheet_d['sheet'] = sheet['sheet_id']
                sheet_d['sheet_styles'] = sheet_styles
                sheet_d['row_heights'] = row_heights
                sheet_d['col_widths'] = col_widths
                sheet_d['data'] = data
                sheet_d['merged_cells'] = merged_cells
                sheet_d['sections'] = section_details
                sheet_d_list.append(sheet_d)

            if reporting_date != '19000101' and reporting_date and report_snapshot != 'null' and report_parameters!='null':
                app.logger.info("Creating formatted summary set")
                report_kwargs=json.loads(report_parameters)
                report_kwargs["populate_summary"]= False
                app.logger.info(report_snapshot)
                summary_set = {}
                summary_set = self.report.create_report_summary_final(report_version_no=version,
                                                                    report_snapshot=report_snapshot,
                                                                    cell_format_yn=cell_format_yn,
                                                                    **report_kwargs)

                print("After create report summary final...")
                for e in summary_set:
                    idx = [x['sheet_id'] for x in sheets].index(e['sheet_id'])
                    cell = e['cell_id']
                    start_xy = coordinate_from_string(cell)
                    # note that openpyxls util provides visual coordinates, but array elements starts with 0
                    start_row = start_xy[1]-1
                    start_col = column_index_from_string(start_xy[0])-1
                    # app.logger.info("idx {} {} {} {}".format(idx,start_row,start_col,cell))
                    sheet_d_list[idx]['data'][start_row][start_col] = e['cell_summary']


            json_dump = (sheet_d_list)
            # print(json_dump)
            return json_dump
        except Exception as e:
            app.logger.error(str(e))
            return {"msg": str(e)}, 500
            #raise e

    def save_hot_table_report_template(self,report_data):
        app.logger.info("Saving hot table report template")
        try:
            def_object = "report_free_format_def"
            if self.db_object_suffix:
                def_object += "_" + self.db_object_suffix

            cell_render_ref = None

            app.logger.info("Deleting definition entries for report {}".format(self.report_id,))
            self.db.transact('delete from {} where report_id=%s'.format(def_object,), (self.report_id,))
            sheet_index =0
            for sheet in report_data:
                sheet_index += 1
                app.logger.info("Processing definition entries for sheet {} ,report {}".format(sheet['sheet'],self.report_id))

                app.logger.info("Creating entries for row height")
                for row in sheet['rowHeights'].keys():
                    self.db.transact(('insert into {}(report_id,sheet_id,cell_id,row_id,cell_render_def,cell_calc_ref) '+\
                                 'values(%s,%s,%s,%s,%s,%s)').format(def_object,),
                                 (self.report_id, sheet['sheet'], str(row),row, 'ROW_HEIGHT', str(sheet['rowHeights'][str(row)])))

                app.logger.info("Creating entries for column width")
                for col in sheet['colWidths'].keys():
                    self.db.transact(('insert into {}(report_id,sheet_id,cell_id,col_id,cell_render_def,cell_calc_ref) '+\
                                'values(%s,%s,%s,%s,%s,%s)').format(def_object,),
                                (self.report_id, sheet['sheet'], col, col, 'COLUMN_WIDTH', str(sheet['colWidths'][col]/8)))

                app.logger.info("Creating entries for merged cells")
                # Note: Merged cell is brought as is, this row and col index starts at 0 (not 1 as it appears for visual co-ordinate
                merged_cell={}
                for rng in sheet['mergedCells']:
                    # print rng
                    startcell = get_column_letter(rng['col']+1) + str(rng['row']+1)
                    endcell = get_column_letter(rng['col']+rng['colspan']) + str(rng['row']+rng['rowspan'])
                    cell_rng = startcell + ':' + endcell
                    start_row = rng['row']+1
                    start_col = get_column_letter(rng['col']+1)
                    agg_ref='S'+str(sheet_index)+'AGG'+str(startcell)
                    merged_cell[startcell]={
                                            'cell_rng': cell_rng,
                                            'agg_ref': agg_ref,
                                            }

                app.logger.info("Creating entries for sections")
                for sec in sheet['sections']:
                    row = sec['ht_range'][0]+1
                    col = get_column_letter(sec['ht_range'][1]+1)
                    self.db.transact(('insert into {}(report_id,sheet_id,cell_id,col_id,row_id,cell_render_def,cell_calc_ref) '+\
                                'values(%s,%s,%s,%s,%s,%s,%s)').format(def_object,),
                                (self.report_id, sheet['sheet'], sec['range'], col, row, 'SECTION_DEF', json.dumps(sec)))
                # if sheet_index==1 :
                #     app.logger.info("Merged cell data {} {} {}".format(sheet_index, merged_cell,sheet))

                for r,cols in enumerate(sheet['sheetData']):
                    for c,val in enumerate(cols):
                        is_merged_cell = False
                        row = r+1
                        col = get_column_letter(c+1)
                        start_cell = col + str(row)
                        if start_cell in merged_cell.keys():
                            cell_id = merged_cell[start_cell]['cell_rng']
                            is_merged_cell = True
                        else:
                            cell_id = start_cell


                        if val or is_merged_cell:
                            agg_ref='S'+str(sheet_index)+'AGG'+str(start_cell)
                            cell_render_def = 'MERGED_CELL' if is_merged_cell else 'STATIC_TEXT'
                            # if sheet_index==1:
                            #     app.logger.info("Value of the cell .. {} {} {} {} {} {} {}".format(self.report_id, sheet['sheet'], cell_id, col, row,cell_render_def, val))
                            self.db.transact(('insert into {}(report_id,sheet_id,cell_id,col_id,row_id,cell_render_def,cell_calc_ref) '+\
                                    'values(%s,%s,%s,%s,%s,%s,%s)').format(def_object,),
                                    (self.report_id, sheet['sheet'], cell_id, col, row,cell_render_def, val))
                            self.db.transact(('insert into {}(report_id,sheet_id,cell_id,col_id,row_id,cell_render_def,cell_calc_ref) '+\
                                        'values(%s,%s,%s,%s,%s,%s,%s)').format(def_object,),
                                        (self.report_id, sheet['sheet'], cell_id, col, row, 'COMP_AGG_REF', agg_ref))

                        if start_cell in sheet['sheetStyles'].keys():
                            cell_style = sheet['sheetStyles'][start_cell]
                            if sheet_index==1:
                                app.logger.info("Value of the cell .. {} {} {} {} {} {} {}".format(self.report_id, sheet['sheet'], cell_id, col, row,'CELL_STYLE', json.dumps(cell_style)))
                            self.db.transact(('insert into {}(report_id,sheet_id,cell_id,col_id,row_id,cell_render_def,cell_calc_ref) '+\
                                    'values(%s,%s,%s,%s,%s,%s,%s)').format(def_object,),
                                    (self.report_id, sheet['sheet'], cell_id, col, row, 'CELL_STYLE', json.dumps(cell_style)))

            self.db.commit()
            return {"msg": "Report [" + self.report_id + "] template updates has been captured successfully "}, 200
        except Exception as e:
            app.logger.error(str(e))
            raise
            # return {"msg": str(e)}, 500

    def create_DAG(self,sections):
        try:
            self.DAG.reset_graph()
            # Sec dependency will be a list of sections like:
            # {section_id:'section id', section_position:[list of parent section_id], .....}
            # app.logger.info("sections: {}".format(sections,))
            for sec in sections:
                # Create the node for the section
                # app.logger.info("Creating node for {}, {}".format(sec['section_id'],sec))
                self.DAG.add_node(str(sec['section_id']))

            for sec in sections:
                # Now add dependencies
                # app.logger.info("Creating dependencies for {}".format(sec['section_position']))
                for dep in sec['section_position']:
                    self.DAG.add_edge(str(dep), str(sec['section_id']))
        except Exception as e:
            app.logger.error(str(e))
            raise(e)

    def validate_sec_dependency(self,sections):
        try:
            self.create_DAG(sections);

            while len(sections):
                self.DAG.reset_graph()
                self.create_DAG(sections)
                new_sections = []
                for node in self.DAG.ind_nodes():
                    app.logger.info("ind_node section [{}] {}".format(node,len(sections)))
                for sec in sections:
                    for node in self.DAG.ind_nodes():
                        if node in sec['section_position']:
                            sec['section_position'].remove(node)
                    if sec['section_id'] not in self.DAG.ind_nodes():
                        new_sections.append(sec)
                    # app.logger.info("New sections {}".format(new_sections))

                sections = new_sections

            return {"msg": "Sections validated successfully", "donotUseMiddleWare": True}, 200
        except Exception as e:
            app.logger.error(str(e))
            return {"msg": "Invalid sections definition detected (possible reasons: use of cyclic reference, invalid section refernce etc), please check!" + str(e) ,
                    "donotUseMiddleWare": True}, 400

    def traverse_DAG_nodes(self,sections):
        try:
            while len(sections):
                self.DAG.reset_graph()
                self.create_DAG(sections)
                new_sections = []
                for node in self.DAG.ind_nodes():
                    app.logger.info("ind_node section [{}] {}".format(node,len(sections)))
                    # Do sec data population for the report json
                # Once the independent node action complete, refresh the DAG with remaining nodes
                # so that appropriate actions can be taken for respective nodes
                for sec in sections:
                    for node in self.DAG.ind_nodes():
                        if node in sec['section_position']:
                            sec['section_position'].remove(node)
                    if sec['section_id'] not in self.DAG.ind_nodes():
                        new_sections.append(sec)
                    # app.logger.info("New sections {}".format(new_sections))

                sections = new_sections

            # return {"msg": "Sections validated successfully", "donotUseMiddleWare": True}, 200
        except Exception as e:
            app.logger.error(str(e))
            # return {"msg": "Invalid sections definition detected (possible reasons: use of cyclic reference, invalid section refernce etc), please check!" + str(e) ,
            #         "donotUseMiddleWare": True}, 400
