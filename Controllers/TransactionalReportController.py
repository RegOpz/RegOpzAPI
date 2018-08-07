from app import *
from flask_restful import Resource,abort
import os
from flask import Flask, request, redirect, url_for
from werkzeug.utils import secure_filename
import uuid
from Constants.Status import *
from Helpers.DatabaseHelper import DatabaseHelper
from Models.Document import Document
import datetime
from collections import defaultdict
import openpyxl as xls
import Helpers.utils as util
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Protection
from openpyxl.utils import get_column_letter, coordinate_from_string
import Helpers.utils as util
import json
import ast
from operator import itemgetter
from datetime import datetime
from Controllers.DefChangeController import DefChangeController
import time
import math
import re
import pandas as pd
from Models.Token import Token
from Helpers.utils import autheticateTenant
from Helpers.authenticate import *
from Controllers.OperationalLogController import OperationalLogController
import Parser.FormulaTranslator as fm_trns
from Parser.PandasLib import *

UPLOAD_FOLDER = './uploads/templates'
ALLOWED_EXTENSIONS = set(['xls', 'xlsx'])

class TransactionalReportController(Resource):
    def __init__(self):
        self.domain_info = autheticateTenant()
        self.master_db=DatabaseHelper()
        if self.domain_info:
            tenant_info = json.loads(self.domain_info)
            self.tenant_id = tenant_info['tenant_id']
            self.tenant_info = json.loads(tenant_info['tenant_conn_details'])
            self.db=DatabaseHelper(self.tenant_info)
            self.user_id = Token().authenticate()
            self.dcc_tenant = DefChangeController(tenant_info=self.tenant_info)

            self.opsLog = OperationalLogController()
            self.er = {}
            self.log_master_id = None
            self.user_id = Token().authenticate()

    @authenticate
    def get(self,report_id=None,cell_id=None, rule_cell_id=None,reporting_date=None):
        if request.endpoint == "trans_report_rule_audit_ep":
            report_id = request.args.get('report_id')
            sheet_id = request.args.get('sheet_id')
            section_id = request.args.get('section_id')
            return self.get_trans_report_audit_list(report_id=report_id, sheet_id=sheet_id, section_id=section_id)

        if report_id and reporting_date:
            self.report_id=report_id
            version= request.args.get('version') # needs to be amanded to check for request args
            return self.render_trans_view_report_json(reporting_date = reporting_date, version = version)

        if report_id and not reporting_date:
            self.report_id = report_id
            # print("Report id",self.report_id)
            self.db_object_suffix= request.args.get('domain_type')
            return self.render_trans_report_json()

        if cell_id:
            self.cell_id = cell_id
            self.report_id = request.args.get('report_id')
            self.sheet_id = request.args.get('sheet_id')
            self.db_object_suffix= request.args.get('domain_type')
            return self.get_trans_report_sec()

        if rule_cell_id:
            self.cell_id = rule_cell_id
            self.report_id = request.args.get('report_id')
            self.sheet_id = request.args.get('sheet_id')
            return self.get_trans_report_rules()

    def put(self,id=None,report=None):
        if id == None and report == None:
            return BUSINESS_RULE_EMPTY
        data = request.get_json(force=True)
        if request.endpoint == "trans_report_rule":
            res = self.dcc_tenant.update_or_delete_data(data, id)
            return res
        if request.endpoint == "report_parameter_ep":
            pass
            # return self.update_report_parameters(data, report)

    def post(self, calc_ref=None, report_id=None):

         if request.endpoint == "trans_bulk_process":
            data = request.get_json(force=True)
            res = self.delete_trans_report_rules(data)
            return res

         if request.endpoint == 'load_trans_report_template_ep':
            return self.capture_template()

         if request.endpoint=='update_trans_section_ep':
            params=request.get_json(force=True)
            print(params)
            self.report_id=params['report_id']
            self.sheet_id=params['sheet_id']
            cell_group=params['cell_group']
            section_id=params['section_id']
            section_type=params['section_type']
            self.db_object_suffix= request.args.get('domain_type')
            return self.update_section_ref(cell_group,section_id,section_type)

         if report_id:
            self.report_id = report_id
            report_info=request.get_json(force=True)
            return self.create_report(report_info)
         if request.endpoint == 'trans_report_rule':
             data = request.get_json(force=True)
             res = self.dcc_tenant.insert_data(data)
             return res

    def capture_template(self):
        try:
            if 'file' not in request.files:
                return NO_FILE_SELECTED

            self.report_id = request.form.get('report_id')
            self.country = request.form.get('country').upper()
            self.report_description = request.form.get('report_description')
            self.report_type=request.form.get('report_type')
            self.db_object_suffix= request.form.get('domain_type')

            if self.report_id == None or self.report_id == "":
                return REPORT_ID_EMPTY

            if self.country == None or self.country == "":
                return COUNTRY_EMPTY

            file = request.files['file']
            self.selected_file=file.filename
            if file and not self.allowed_file(file.filename):
                return FILE_TYPE_IS_NOT_ALLOWED
            filename = str(uuid.uuid4()) + "_" + secure_filename(file.filename)
            file.save(os.path.join(UPLOAD_FOLDER, filename))

            res = self.insert_report_def_catalog()
            if res:
                return self.load_report_template(filename)
            else:
                return res
        except Exception as e:
            app.logger.error(str(e))
            return {"msg":str(e)},500

    def create_report(self, report_info):
        try:
            print(report_info)
            if 'report_event'  in report_info.keys():
                report_id = report_info['report_id']
                report_create_date=report_info['report_create_date']
                #report_parameters = report_info['report_parameters']
                reporting_date = report_info['reporting_date']
                report_kwargs = eval("{'report_id':'" + report_id + "' ," + report_parameters.replace('"',"'") + "}")

                business_date_from=report_kwargs['business_date_from']
                business_date_to=report_kwargs['business_date_to']
                self.update_report_catalog(status='RUNNING',report_id=report_id,reporting_date=reporting_date,report_info=report_info,report_create_date=report_create_date)

            else:

                report_version_no = self.create_report_catalog(report_info)
                self.update_report_catalog(report_info = report_info
                                          , status='RUNNING'
                                          , version = report_version_no)

            report_snapshot = self.create_report_detail(report_version_no, report_info)

            self.update_report_catalog(report_info = report_info
                                       , status='SUCCESS'
                                       , report_snapshot=report_snapshot
                                       , version=report_version_no )

            report_id = report_info["report_id"]
            reporting_date = report_info["reporting_date"]
            if self.log_master_id:
                self.opsLog.write_log_detail(master_id=self.log_master_id
                    , operation_sub_type='End of Create Report'
                    , operation_status='Complete'
                    , operation_narration="Report generated SUCCESSFULLY for [{0}] Reporting date [{1}].".format(str(report_id),str(reporting_date))
                    )
                self.opsLog.update_master_status(id=self.log_master_id,operation_status="SUCCESS")

            return {"msg": "Report generated SUCCESSFULLY for ["+str(report_id)+"] Reporting date ["+str(reporting_date)+"]."}, 200
        except Exception as e:
            self.db.rollback()
            app.logger.error(str(e))
            if self.log_master_id:
                self.opsLog.write_log_detail(master_id=self.log_master_id
                    , operation_sub_type='Error occured while creating report'
                    , operation_status='Failed'
                    , operation_narration="Report not generated  for [{0}] Reporting date [{1}].".format(str(report_info['report_id']),str(report_info['reporting_date']))
                    )
                self.opsLog.update_master_status(id=self.log_master_id,operation_status="ERROR")
            #return {"msg":str(e)},500
            raise e

    def create_report_catalog(self,report_info):
        try:
            db=self.db
            report_id = report_info["report_id"]
            reporting_date = report_info["reporting_date"]
            report_create_date = report_create_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            report_create_status=report_info['report_create_status']
            as_of_reporting_date=report_info['as_of_reporting_date']

            report_version=self.db.query("select max(version) version from report_catalog where report_id=%s and reporting_date=%s",
                                         (report_id,reporting_date)).fetchone()
            report_version_no=1 if not report_version['version'] else  report_version['version']+1

            sql="insert into report_catalog(report_id,reporting_date,report_create_date,\
                report_parameters,report_create_status,as_of_reporting_date,version,report_created_by) values(%s,%s,%s,%s,%s,%s,%s,%s)"
            print(sql)
            catalog_id =         db.transact(sql,(report_id,reporting_date,report_create_date,report_info.__str__(),report_create_status,as_of_reporting_date,          report_version_no,self.user_id))
            self.log_master_id = self.opsLog.write_log_master(operation_type='Create Report'
                , operation_status = 'RUNNING'
                , operation_narration = 'Create report {0} for {1} as on {2}'.format(report_id,reporting_date,as_of_reporting_date)
                , entity_type = 'Report'
                , entity_name = report_id
                , entity_table_name = 'report_catalog'
                , entity_id = catalog_id
                )

            db.commit()
            return report_version_no
            if self.log_master_id:
                self.opsLog.write_log_detail(master_id=self.log_master_id
                    , operation_sub_type='Create catalog entry'
                    , operation_status='Complete'
                    , operation_narration='Report creation started with the parameters : {0}'.format(report_info,))
        except Exception as e:
            app.logger.error(e.__str__())
            if self.log_master_id:
                self.opsLog.write_log_detail(master_id=self.log_master_id
                    , operation_sub_type='Error occured while creating catalog'
                    , operation_status='Failed'
                    , operation_narration='Report creation Failed with error : {0}'.format(str(e),))
            raise (e)

    def update_report_catalog(self, report_info = None,  status = None, report_snapshot=None, version = 0):
        try:
            db=self.db
            report_id = report_info["report_id"]
            reporting_date = report_info["reporting_date"]
            report_create_date = report_create_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            update_clause = "report_create_status='{0}'".format(status,)
            if report_info != None:
                # Replace all singlequotes(') with double quote(") as update sql requires all enclosed in ''
                report_info_str = json.dumps(report_info)
                update_clause += ", report_parameters='{0}'".format(report_info_str,)
            if report_create_date != None:
                # Replace all singlequotes(') with double quote(") as update sql requires all enclosed in ''
                update_clause += ", report_create_date='{0}'".format(report_create_date.replace("'",'"'),)
            if report_snapshot !=None:
                update_clause +=", report_snapshot='{0}'".format(report_snapshot)
            sql = "update report_catalog set {0} where report_id='{1}' and reporting_date='{2}'and \
                    version={3}".format(update_clause,report_id,reporting_date, version)
            db.transact(sql)
            db.commit()
            if self.log_master_id:
                self.opsLog.write_log_detail(master_id=self.log_master_id
                    , operation_sub_type='Updated report catalog'
                    , operation_status='Complete'
                    , operation_narration='Report catalog updated with : {0}'.format(update_clause,))

        except Exception as e:
            app.logger.error(e.__str__())
            if self.log_master_id:
                self.opsLog.write_log_detail(master_id=self.log_master_id
                    , operation_sub_type='Error occured while updating report catalog'
                    , operation_status='Failed'
                    , operation_narration='Report creation Failed with error : {0}'.format(str(e),))
            raise(e)


    def allowed_file(self,filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

    def insert_report_def_catalog(self):
        app.logger.info("Creating entry for report catalog")
        try:
            catalog_object = "report_def_catalog"
            if self.db_object_suffix and self.db_object_suffix!='null' and self.db_object_suffix!='undefined':
                catalog_object += "_" + self.db_object_suffix
            app.logger.info("Checking if report {} for country {} already exists in catalog".format(self.report_id,self.country))
            count = self.db.query("select count(*) as count from {} where report_id=%s and country=%s".format(catalog_object,),\
                                (self.report_id,self.country,)).fetchone()
            if not count['count']:
                app.logger.info("Creating catalog entry for country {} and report {}".format(self.country,self.report_id))
                res = self.db.transact("insert into {}(report_id,country,report_description,report_type) values(%s,%s,%s,%s)".format(catalog_object,),\
                        (self.report_id,self.country,self.report_description,self.report_type))
                # self.db.commit()
                return True
            else:
                app.logger.info("Catalog entry exists for country {} and report {}".format(self.country,self.report_id))
                return True
        except Exception as e:
            app.logger.error(str(e))
            raise e
            #return {"msg":str(e)},500

    def load_report_template(self,template_file_name):
        app.logger.info("Loading report template")
        try:
            def_object = "report_dyn_trans_def"
            if self.db_object_suffix and self.db_object_suffix!='null' and self.db_object_suffix!='undefined':
                def_object += "_" + self.db_object_suffix
            formula_dict = {'SUM': 'CALCULATE_FORMULA',
                            '=SUM': 'CALCULATE_FORMULA',
                            }
            cell_render_ref = None
            target_dir = UPLOAD_FOLDER + "/"
            app.logger.info("Loading {} file from {} directory".format(template_file_name,target_dir))
            wb = xls.load_workbook(target_dir + template_file_name)

            #db = DatabaseHelper()

            sheet_index=0
            for sheet in wb.worksheets:
                sheet_index+=1
                app.logger.info("Deleting definition entries for sheet {} ,report {}".format(sheet.title,self.report_id))
                self.db.transact('delete from {} where report_id=%s and sheet_id=%s'.format(def_object,), (self.report_id, sheet.title,))

                # First capture the dimensions of the cells of the sheet
                rowHeights = [sheet.row_dimensions[r + 1].height for r in range(sheet.max_row)]
                colWidths = [sheet.column_dimensions[get_column_letter(c + 1)].width for c in range(sheet.max_column)]

                app.logger.info("Creating entries for row height")
                for row, height in enumerate(rowHeights):
                    self.db.transact('insert into {}(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref,row_id)\
                                 values(%s,%s,%s,%s,%s,%s)'.format(def_object,), (self.report_id, sheet.title, str(row + 1), 'ROW_HEIGHT', str(height),(row + 1)))

                app.logger.info("Creating entries for column width")
                for col, width in enumerate(colWidths):
                    self.db.transact('insert into {}(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref)\
                                values(%s,%s,%s,%s,%s)'.format(def_object,),
                                (self.report_id, sheet.title, get_column_letter(col + 1), 'COLUMN_WIDTH', str(width)))

                app.logger.info("Creating entries for merged cells")
                rng_startcell = []
                rng_boundary=[]
                for rng in sheet.merged_cell_ranges:
                    # print rng
                    startcell, endcell = rng.split(':')
                    colrow = coordinate_from_string(startcell)
                    # print sheet.cell(startcell).border
                    rng_startcell.append(startcell)
                    rng_boundary.append(rng)
                    #agg_ref='S'+str(sheet_index)+'AGG'+str(startcell)
                    _cell=sheet[startcell]
                    cell_style={"font":{"name":_cell.font.name,
                                        "bold": _cell.font.b,
                                        "italic": _cell.font.i,
                                        "colour": _cell.font.color.rgb if _cell.font.color else 'None',
                                        "size": _cell.font.sz},
                                "fill": {"type": _cell.fill.fill_type,
                                        "colour" : _cell.fill.fgColor.rgb if _cell.fill.fgColor else 'None'},
                                "alignment": {"horizontal": _cell.alignment.horizontal,
                                        "vertical": _cell.alignment.vertical},
                                "border":{"left":{"style": _cell.border.left.style,
                                                "colour":_cell.border.left.color.rgb if _cell.border.left.color else 'None'} ,
                                        "right":{"style": _cell.border.right.style,
                                                "colour":_cell.border.right.color.rgb if _cell.border.right.color else 'None'},
                                        "top":{"style": _cell.border.top.style,
                                                "colour":_cell.border.top.color.rgb if _cell.border.top.color else 'None'},
                                        "bottom":{"style": _cell.border.bottom.style,
                                                "colour":_cell.border.bottom.color.rgb if _cell.border.bottom.color else 'None'},}
                    }

                    self.db.transact('insert into {}(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref,col_id,row_id)\
                                values(%s,%s,%s,%s,%s,%s,%s)'.format(def_object,),
                                (self.report_id, sheet.title, rng, 'MERGED_CELL', sheet[startcell].value,colrow[0],str(colrow[1])))

                    self.db.transact('insert into {}(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref,col_id,row_id)\
                                values(%s,%s,%s,%s,%s,%s,%s)'.format(def_object,),
                                (self.report_id, sheet.title, rng, 'CELL_STYLE', str(cell_style),colrow[0],str(colrow[1])))

                app.logger.info("Creating entries for static text and formulas")
                for all_obj in sheet['A1':util.cell_index(sheet.max_column, sheet.max_row)]:
                    for cell_obj in all_obj:
                        cell_ref = str(cell_obj.column) + str(cell_obj.row)
                        #agg_ref='S'+str(sheet_index)+'AGG'+str(cell_ref)
                        _cell=sheet[cell_ref]
                        cell_style={"font":{"name":_cell.font.name,
                                            "bold": _cell.font.b,
                                            "italic": _cell.font.i,
                                            "colour": _cell.font.color.rgb if _cell.font.color else 'None',
                                            "size": _cell.font.sz},
                                    "fill": {"type": _cell.fill.fill_type,
                                            "colour" : _cell.fill.fgColor.rgb if _cell.fill.fgColor else 'None'},
                                    "alignment": {"horizontal": _cell.alignment.horizontal,
                                            "vertical": _cell.alignment.vertical},
                                    "border":{"left":{"style": _cell.border.left.style,
                                                    "colour":_cell.border.left.color.rgb if _cell.border.left.color else 'None'} ,
                                            "right":{"style": _cell.border.right.style,
                                                    "colour":_cell.border.right.color.rgb if _cell.border.right.color else 'None'},
                                            "top":{"style": _cell.border.top.style,
                                                    "colour":_cell.border.top.color.rgb if _cell.border.top.color else 'None'},
                                            "bottom":{"style": _cell.border.bottom.style,
                                                    "colour":_cell.border.bottom.color.rgb if _cell.border.bottom.color else 'None'},}
                        }
                        if (len(rng_startcell) > 0 and cell_ref not in rng_startcell) or (len(rng_startcell) == 0):
                            if not self.check_if_cell_isin_range(cell_ref,rng_boundary):
                                cell_obj_value = str(cell_obj.value) if cell_obj.value else ''
                                cell_render_ref = 'STATIC_TEXT'

                                self.db.transact('insert into {}(report_id,sheet_id ,cell_id ,cell_render_def ,cell_calc_ref,col_id,row_id)\
                                          values(%s,%s,%s,%s,%s,%s,%s)'.format(def_object,),
                                            (self.report_id, sheet.title, cell_ref, cell_render_ref, cell_obj_value.strip(),cell_obj.column,str(cell_obj.row)))

                                self.db.transact('insert into {}(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref,col_id,row_id)\
                                                 values(%s,%s,%s,%s,%s,%s,%s)'.format(def_object,),
                                                 (self.report_id, sheet.title, cell_ref, 'CELL_STYLE', str(cell_style),cell_obj.column,str(cell_obj.row)))

            self.db.commit()
            return {"msg": "Report [" + self.report_id + "] template has been captured successfully using " + self.selected_file}, 200
        except Exception as e:
            app.logger.error(e)
            return {"msg": str(e)}, 500

    def update_section_ref(self,cell_group,section_id,section_type):
        app.logger.info("Marking section {} for report {} and sheet".format(section_id,self.report_id,self.sheet_id))
        try:
            def_object = "report_dyn_trans_def"
            db = self.db
            if self.db_object_suffix and self.db_object_suffix!='null' and self.db_object_suffix!='undefined':
                def_object += "_" + self.db_object_suffix
                db = self.master_db
            rows=[]
            columns=[]
            for cell in cell_group:
                cell_index=re.split('(\d+)',cell)
                rows.append(cell_index[1])
                columns.append(cell_index[0])

            max_row=int(max(rows))
            min_row=int(min(rows))
            max_col=max(columns)
            min_col=min(columns)

            app.logger.info("Updating report_dyn_trans_def for making section {} as type {}".format(section_id,section_type))

            db.transact("update {0} set section_id=%s,section_type=%s where row_id between %s and %s and\
                          col_id between %s and %s and report_id=%s and sheet_id=%s".format(def_object,),(section_id,section_type,min_row,max_row,\
                          min_col,max_col,self.report_id,self.sheet_id))
            app.logger.info("update report_dyn_trans_def set section_id={0},section_type={1} where row_id between {2} and {3} and\
                          col_id between {4} and {5} and report_id={6} and sheet_id={7}".format(section_id,section_type,min_row,max_row,\
                          min_col,max_col,self.report_id,self.sheet_id))
            db.commit()
            return {"msg":"Cell {} marked successfully as section {} for report {},sheet {}"\
                    .format(cell_group,section_id,self.report_id,self.sheet_id)},200
        except Exception as e:
            app.logger.error(e)
            return {"msg":str(e)}, 500

    def check_if_cell_isin_range(self,cell_id,rng_boundary_list):

        def number_from_word(value):
            base='ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            result=0
            i=len(value)-1
            for j in range(0,len(value)):
              result+=math.pow(len(base),i)*(base.index(value[j])+1)
              i-=1
            return result

        def split_numbers_chars(string):
            sub=re.split('(\d+)', string)
            return sub

        def check_inclusion(cell,rng):
            start_cell, end_cell = rng.split(':')
            start_cell_sub=split_numbers_chars(start_cell)
            end_cell_sub=split_numbers_chars(end_cell)
            cell_sub=split_numbers_chars(cell)

            within_rng=False

            if number_from_word(cell_sub[0]) >=number_from_word(start_cell_sub[0]) \
            and number_from_word(cell_sub[0]) <= number_from_word(end_cell_sub[0]) \
            and  cell_sub[1] >= start_cell_sub[1] and cell_sub[1] <= end_cell_sub[1]:
                within_rng=True

            return within_rng

        for rng in rng_boundary_list:
            incl=check_inclusion(cell_id,rng)
            if incl:
                break
        return incl

    def render_trans_report_json(self):

        app.logger.info("Rendering Transactional report Template")

        try:
            app.logger.info("Getting list of sheet for report {0} {1}".format(self.report_id,self.tenant_id))
            def_object = "report_dyn_trans_def"
            db = self.db
            if self.db_object_suffix and self.db_object_suffix!='null' and self.db_object_suffix!='undefined':
                def_object += "_" + self.db_object_suffix
                db = self.master_db

            app.logger.info("Fetching report def for {0}".format(def_object,))
            sheets = db.query("select distinct sheet_id from {} where report_id=%s".format(def_object,),
                                   (self.report_id,)).fetchall()

            agg_format_data = {}

            sheet_d_list = []
            for sheet in sheets:
                matrix_list = []
                row_attr = {}
                col_attr = {}
                cell_style = {}
                app.logger.info("Getting report definition for report {0},sheet {1}".format(self.report_id,sheet["sheet_id"]))
                report_template = db.query(
                    "select cell_id,cell_render_def,cell_calc_ref,section_id,section_type,col_id,row_id " + \
                    " from {} where report_id=%s and sheet_id=%s".format(def_object,),
                    (self.report_id, sheet["sheet_id"])).fetchall()

                app.logger.info("Writing report definition to dictionary")
                for row in report_template:
                    cell_d = {}
                    if row["cell_render_def"] == 'STATIC_TEXT':
                        cell_d['cell'] = row['cell_id']
                        cell_d['value'] = row['cell_calc_ref'] + (" DYNDATA("+row['section_id']+")" if row['section_type']=="DYNDATA" else "")
                        cell_d['origin'] = "TEMPLATE"
                        cell_d['section'] = row['section_id']
                        cell_d['sectionType'] = row['section_type']
                        cell_d['col'] = row['col_id']
                        cell_d['row'] = row['row_id']
                        matrix_list.append(cell_d)


                    elif row['cell_render_def'] == 'MERGED_CELL':
                        start_cell, end_cell = row['cell_id'].split(':')
                        cell_d['cell'] = start_cell
                        cell_d['value'] = row['cell_calc_ref'] + (" DYNDATA("+row['section_id']+")" if row['section_type']=="DYNDATA" else "")
                        cell_d['merged'] = end_cell
                        cell_d['origin'] = "TEMPLATE"
                        cell_d['section'] = row['section_id']
                        cell_d['sectionType'] = row['section_type']
                        cell_d['col'] = row['col_id']
                        cell_d['row'] = row['row_id']
                        matrix_list.append(cell_d)


                    elif row['cell_render_def'] == 'ROW_HEIGHT':
                        if row['cell_calc_ref'] == 'None':
                            row_height = '12.5'
                        else:
                            row_height = row['cell_calc_ref']
                        row_attr[row['cell_id']] = {'height': row_height}


                    elif row['cell_render_def'] == 'COLUMN_WIDTH':
                        if row['cell_calc_ref'] == 'None':
                            col_width = '13.88'
                        else:
                            col_width = row['cell_calc_ref']
                        col_attr[row['cell_id']] = {'width': col_width}

                    elif row['cell_render_def'] == 'CELL_STYLE':
                        if ':' in row['cell_id']:
                            start_cell, end_cell = row['cell_id'].split(':')
                        else:
                            start_cell=row['cell_id']

                        app.logger.info("Inside CELL_STYLE for cell {}".format(start_cell,))
                        cell_style[start_cell] = eval(row['cell_calc_ref'])

                sheet_d = {}
                sheet_d['sheet'] = sheet['sheet_id']
                # print(sheet_d['sheet'])
                sheet_d['matrix'] = matrix_list
                sheet_d['row_attr'] = row_attr
                sheet_d['col_attr'] = col_attr
                sheet_d['cell_style'] = cell_style
                sheet_d_list.append(sheet_d)


            json_dump = (sheet_d_list)
            # print(json_dump)
            return json_dump
        except Exception as e:
            app.logger.error(str(e))
            return {"msg": str(e)}, 500

    def get_trans_report_sec(self):
        try:
            app.logger.info("Getting list of section for report {0} {1}".format(self.report_id,self.tenant_id))
            def_object = "report_dyn_trans_def"
            db = self.db
            if self.db_object_suffix and self.db_object_suffix!='null' and self.db_object_suffix!='undefined':
                def_object += "_" + self.db_object_suffix
                db = self.master_db
            section_id = db.query("select distinct section_id from {0} where report_id=%s and sheet_id=%s and cell_id=%s".format(def_object,), \
                    (self.report_id,self.sheet_id, self.cell_id)).fetchone()
            app.logger.info("Fetching section details {0} {1} {2} {3}".format(self.report_id,self.sheet_id, self.cell_id,section_id,))
            section_range = db.query("select section_id, min(col_id) min_col_id,min(row_id) min_row_id," + \
                    "max(col_id) max_col_id,max(row_id) max_row_id,max(section_type) section_type " + \
                    "from {0} \
                    where report_id=%s and sheet_id=%s and section_id=%s".format(def_object,), \
                    (self.report_id,self.sheet_id, str(section_id['section_id']))).fetchone()
            return section_range
        except Exception as e:
            app.logger.error(str(e))
            return {"msg": str(e)}, 500

    def get_trans_report_rules(self):
        try:
            section_id = self.db.query("select distinct section_id from report_dyn_trans_def where report_id=%s and sheet_id=%s and cell_id=%s", \
                    (self.report_id,self.sheet_id, self.cell_id)).fetchone()
            app.logger.info("Fetching section rule details {0} {1} {2} {3}".format(self.report_id,self.sheet_id, self.cell_id,section_id,))
            section_col = self.db.query("select distinct col_id from report_dyn_trans_def where report_id=%s and sheet_id=%s and section_id=%s", \
                    (self.report_id,self.sheet_id, section_id['section_id'])).fetchall()
            section_rules = self.db.query("select * " + \
                    "from report_dyn_trans_calc_def \
                    where report_id=%s and sheet_id=%s and section_id=%s", \
                    (self.report_id,self.sheet_id, str(section_id['section_id']))).fetchall()
            order_rules = self.db.query("select * " + \
                    "from report_dyn_trans_agg_def \
                    where report_id=%s and sheet_id=%s and section_id=%s", \
                    (self.report_id,self.sheet_id, str(section_id['section_id']))).fetchall()
            return {"section": section_id['section_id'],
                    "secRules": section_rules,
                    "secOrders": order_rules,
                    "secColumns": section_col}
        except Exception as e:
            app.logger.error(str(e))
            return {"msg": str(e)}, 500

    def insert_section_calc_rule(self,params):
        try:
            section_id=params['section_id']
            source_id=params['source_id']
            cell_calc_ref=params['cell_calc_ref']
            cell_calc_render_ref=params['cell_calc_render_ref']
            sql="insert into report_dyn_trans_calc_def (report_id,sheet_id,section_id,source_id,cell_calc_ref,cell_calc_render_ref,dml_allowed,in_use)"+\
            "values(%s,%s,%s,%s,%s,%s,'Y','Y')"
            self.db.transact(sql,(self.report_id,self.sheet_id,section_id,source_id,cell_calc_ref,cell_calc_render_ref))
            return {"msg": "New section calculation rule created for Report [{0}] sheet [{1}] section [{2}]".format(self.report_id, self.sheet_id, section_id)}, 200
        except Exception as e:
            app.logger.error(str(e))
            return {"msg": str(e)}, 500

    def update_section_calc_rule(self,params):
        try:
            section_id=params['section_id']
            source_id=params['source_id']
            cell_calc_ref=params['cell_calc_ref']
            cell_calc_render_ref=params['cell_calc_render_ref']
            id=params['id']
            sql="update report_dyn_trans_calc_def set cell_calc_render_ref=%s where id=%s"
            app.logger.info("update report_dyn_trans_calc_def set cell_calc_render_ref='{0}' where id={1}".format(cell_calc_render_ref,id,))
            self.db.transact(sql,(cell_calc_render_ref,id,))
            return {"msg": "Updated section calculation rule for Report [{0}] sheet [{1}] section [{2}] Ref [{3}]".format(self.report_id, self.sheet_id, section_id,cell_calc_ref)}, 200
        except Exception as e:
            app.logger.error(str(e))
            return {"msg": str(e)}, 500


    def create_report_detail(self, report_version_no, report_info):
        try:

            report_id = report_info['report_id']
            business_date_from=report_info['business_date_from']
            business_date_to=report_info['business_date_to']

            reporting_date=business_date_from+business_date_to
            app.logger.info("Getting list of sections for report {}".format(report_id))
            sheets = self.db.query("select distinct sheet_id,section_id from report_dyn_trans_def where report_id=%s and section_type='DYNDATA' ",
                               (report_id,)).fetchall()
            cardf = pd.DataFrame(self.db.query("select id, report_id, sheet_id , section_id from report_dyn_trans_agg_def where \
                                    report_id = %s and in_use = 'Y'",(report_id,)).fetchall())
            car_max_vers =  self.db.query("select report_id, version, id_list from report_dyn_trans_agg_def_vers where report_id = %s\
                                 and version = (select max(version) from report_dyn_trans_agg_def_vers where report_id = %s)",(report_id, report_id)).fetchone()
            if not cardf.empty:
                cardf['id']=cardf['id'].astype(dtype='int64',errors='ignore')
                car_id_list=list(map(int,cardf['id'].tolist()))
                car_id_list.sort()
                car_id_list_str=",".join(map(str,car_id_list))

                if not car_max_vers:
                    car_version_no = 1
                else:
                    old_id_list = list(map(int, car_max_vers['id_list'].split(',')))
                    car_version_no = car_max_vers['version'] + 1 if set(car_id_list) != set(old_id_list) else car_max_vers['version']

                if not car_max_vers or car_version_no != car_max_vers['version']:
                    self.db.transact("insert into report_dyn_trans_agg_def_vers (report_id, id_list, version) VALUES(%s, %s, %s)",
                                    (report_id, car_id_list_str, car_version_no))
                    self.db.commit()
                comp_agg_rule_version = car_version_no
            else:
                if not car_max_vers:
                    comp_agg_rule_version = 1
                else:
                    comp_agg_rule_version = car_max_vers['version']

            #print(sheets)
            all_calc_def = pd.DataFrame(self.db.query("select id, report_id, sheet_id, section_id, source_id from\
                             report_dyn_trans_calc_def where report_id = %s and in_use = 'Y'",(report_id,)).fetchall())
            calc_def_vers = self.db.query("select version, id_list from report_dyn_trans_calc_def_vers where report_id = %s\
                            and version = (select max(version) from report_dyn_trans_calc_def_vers where \
                            report_id = %s)",(report_id, report_id)).fetchone()

            if not all_calc_def.empty:
                all_calc_def['id'] = all_calc_def['id'].astype(dtype='int64', errors='ignore')
                rr_id_list = list(map(int, all_calc_def['id'].tolist()))
                rr_id_list.sort()
                rr_id_list_str=",".join(map(str,rr_id_list))
            else:
                rr_id_list = []
                rr_id_list_str = ''

            if not calc_def_vers:
                calc_def_vers_no = 1
            else:
                calc_def_old_list = list(map(int, calc_def_vers['id_list'].split(',')))
                calc_def_vers_no = calc_def_vers['version'] + 1 if set(rr_id_list) != set(calc_def_old_list) else                calc_def_vers['version']
            print("Calc_def: " , calc_def_vers_no,calc_def_vers)
            if not calc_def_vers or calc_def_vers_no != calc_def_vers['version']:
                self.db.transact("insert into report_dyn_trans_calc_def_vers(report_id,version,id_list) values(%s,%s,%s)",(report_id, calc_def_vers_no, rr_id_list_str))
                self.db.commit()
            report_rule_version = calc_def_vers_no


            qualified_data_version=defaultdict(dict)
            for sheet in sheets:
                sheet_id = sheet['sheet_id']
                section_id=sheet['section_id']
                trans_calc_def=self.get_dyn_trans_calc_def_details(report_id,sheet_id,section_id)
                if trans_calc_def:
                    trans_calc_def=pd.DataFrame(trans_calc_def)
                    app.logger.info("trans_calc_def post pd dataframe.. {}".format(trans_calc_def))

                    qualified_filtered_data=pd.DataFrame()
                    for source in trans_calc_def['source_id'].unique():
                        #source_id = source['source_id']


                        link_data_records = []
                        source_data, source_snapshot =self.get_qualified_source_data(source ,business_date_from,business_date_to)
                        qualified_data_version[str(source)] = source_snapshot

                        df_source_data=pd.DataFrame(source_data)
                        #print(df_source_data.columns)
                        if self.log_master_id:
                            self.opsLog.write_log_detail(master_id=self.log_master_id
                                , operation_sub_type='Processing Qualified data'
                                , operation_status='Started'
                                , operation_narration='Processing of qualified data started for source  : {0}'.format(source,))
                        if not df_source_data.empty:

                            # Create dummy columns (with truth values) on qualified business rules
                            qdr=df_source_data['business_rules'].str.get_dummies(sep=',')
                            # merge both the data frames into one to facilitate the filter
                            qd_source_data=pd.concat([df_source_data,qdr],axis=1)
                            app.logger.info("Before start of the qualified data loop...")
                            tcd_source=trans_calc_def.loc[trans_calc_def['source_id']==source]
                            for idx,rw in tcd_source.iterrows():
                                cell_calc_render_ref=eval(rw['cell_calc_render_ref'])
                                cell_calc_rule=cell_calc_render_ref['rule']
                                expr_str=""
                                for r in cell_calc_rule.split(','):
                                    if r is not None and r !='':
                                        expr_str= "(qd_source_data['{0}']==1)".format(r,) if expr_str=="" else expr_str + " & (qd_source_data['{0}']==1)".format(r,)
                                expr_str = "qd_source_data[" + expr_str + "]"
                                app.logger.info("Before dfr filter ...{0}".format(expr_str,))
                                dfr=eval(expr_str)
                                #print("evaluated string : " , dfr)
                                #qpdf_temp=dfr[['qualifying_key','business_date']]
                                dfr['source_id']=source
                                dfr['report_id'] = report_id
                                dfr['sheet_id'] = sheet_id
                                dfr['section_id'] = section_id
                                dfr['cell_calc_ref'] = rw['cell_calc_ref']
                                dfr['reporting_date'] = reporting_date
                                dfr['version'] = report_version_no
                                # app.logger.info("After dfr filter ...{0}".format(dfr,))
                                if not dfr.empty:
                                    data_dict={}
                                    cell_calc_columns=cell_calc_render_ref['calc']
                                    col_list=[]
                                    expr_str=""
                                    if cell_calc_columns:
                                        for col in cell_calc_columns.keys():
                                            # app.logger.info("Column value [{0}]".format(col))
                                            if cell_calc_columns[col]['column'] is not None and cell_calc_columns[col]['column'] !='':
                                                col_list.append(col)
                                                dfr=self.apply_formula_to_frame(dfr,cell_calc_columns[col]['column'],col)
                                                #expr_str = "\""+cell_calc_columns[col]['column']+"\":\""+col+"\"" if expr_str=="" else expr_str + ",\""+cell_calc_columns[col]['column']+"\":\""+col+"\""
                                    #expr_str="dfr.rename(columns={" + expr_str + "},inplace=True)"
                                    #app.logger.info("Before dfr column rename  ...{0}".format(expr_str,))
                                    #eval(expr_str)

                                    qualified_filtered_data=qualified_filtered_data.append(dfr,ignore_index=True)
                                    print(qualified_filtered_data)
                                    #qpdf=qpdf.append(qpdf_temp,ignore_index=True)
                                    # link_data_records.append((source,report_id,sheet_id,section_id,rw['cell_calc_ref'],row['business_date'],reporting_date))

                        app.logger.info("At the end of the qualified data loop...")
                        if self.log_master_id:
                            self.opsLog.write_log_detail(master_id=self.log_master_id
                                , operation_sub_type='Processing Qualified data'
                                , operation_status='Complete'
                                , operation_narration='Processing of qualified data completed for source  : {0}'.format(source,))


                    qualified_filtered_data.fillna('',inplace=True)
                    if not qualified_filtered_data.empty:
                        sql = "select cell_agg_render_ref from report_dyn_trans_agg_def where report_id = %s \
                                and sheet_id = %s and section_id = %s and in_use='Y'"
                        data = self.db.query(sql,(report_id, sheet_id, section_id)).fetchone()
                        print("Data",data)
                        if data:
                            data = eval(data['cell_agg_render_ref'])
                            if "sort_order" in data.keys():
                                cols = []
                                order = []
                                val = data["sort_order"]
                                for element in val:
                                    key = list(element.keys())[0]
                                    cols.append(key)
                                    if element[key] == 'ASC':
                                        order.append(1)
                                    else:
                                        order.append(0)
                                qualified_filtered_data.sort_values(cols, inplace = True , ascending = order)
                                if 'ranktype' in data.keys():
                                    num = int(data['rank_value'])
                                    if data['ranktype'] == 'TOP':
                                        #print("top")
                                        qualified_filtered_data = qualified_filtered_data.head(num)
                                    else :
                                        qualified_filtered_data = qualified_filtered_data.tail(num)
                        # print("Qualified Data", qualified_filtered_data)
                        qualified_filtered_data = qualified_filtered_data.reset_index()
                        row_id_list=[idx + 1 for idx,row in qualified_filtered_data.iterrows()]
                        qualified_filtered_data['row_id']=row_id_list
                        # app.logger.info("qualified_filtered_data ...{}".format(qualified_filtered_data))
                        qpdf=qualified_filtered_data[['source_id','report_id','sheet_id','section_id','cell_calc_ref','reporting_date','business_date','qualifying_key','row_id','version']]
                        columns = ",".join(qpdf.columns)
                        placeholders = ",".join(['%s'] * len(qpdf.columns))
                        data = list(qpdf.itertuples(index=False, name=None))
                        print(qpdf)
                        row_id=self.db.transactmany("insert into report_dyn_trans_qualified_data_link ({0}) \
                                                    values ({1})".format(columns, placeholders),data)

                        col_list.append('row_id')
                        sum_recs=qualified_filtered_data[col_list].to_dict(orient='records')
                        summary_records=[]

                        for rec in sum_recs:
                            #rec = dict(rec)
                            row_seq=rec['row_id']
                            rec.pop('row_id')
                            print(str(rec))
                            summary_records.append((report_id,sheet_id,section_id,row_seq,str(rec),reporting_date, report_version_no))


                        row_id=self.db.transactmany("insert into report_dyn_trans_summary(report_id,sheet_id,section_id,row_id,row_summary,reporting_date, version)\
                                                    values(%s,%s,%s,%s,%s,%s,%s)",summary_records)

                    self.db.commit()
                report_snapshot=json.dumps({"report_dyn_trans_calc_def":report_rule_version,"report_dyn_trans_agg_def":comp_agg_rule_version,
                                            "qualified_data":qualified_data_version})
                if self.log_master_id:
                    self.opsLog.write_log_detail(master_id=self.log_master_id
                        , operation_sub_type='Report Creation Completed'
                        , operation_status='Complete'
                        , operation_narration='Report creation completed for  report_id  : {0}'.format(report_info['report_id'],))
                return report_snapshot

        except Exception as e:
            self.db.rollback()
            app.logger.error(str(e))
            self.opsLog.write_log_detail(master_id=self.log_master_id
                , operation_sub_type='Error occured while creating report'
                , operation_status='Failed'
                , operation_narration='Report creation Failed with error : {0}'.format(str(e),))
            #return {"msg":str(e)},500
            raise e

    def apply_formula_to_frame(self, df, excel_formula,new_field_name):
        try:
            context=fm_trns.Context('df')
            rpn_expr=fm_trns.shunting_yard(excel_formula)
            G,root=fm_trns.build_ast(rpn_expr)
            df_expr=root.emit(G,context=context)
            for col in list(df.columns):
                if 'date' not in col:
                    df[col]=df[col].astype(dtype='float64',errors='ignore')
            df[new_field_name]=eval(df_expr)
            return df
        except Exception as e:
            app.loger.error(str(e))
            raise(e)

    def get_dyn_trans_calc_def_details(self, report_id, sheet_id,section_id):
        try:
            app.logger.info("Get transactional calc details for Report:{0} Sheet:{1} Section:{2}"\
                            .format(report_id,sheet_id,section_id))

            trans_calc_def=self.db.query("select * from report_dyn_trans_calc_def where report_id=%s and sheet_id=%s and section_id=%s and in_use='Y'",
                                         (report_id,sheet_id,section_id)).fetchall()
            return trans_calc_def

        except Exception as e:
            app.logger.error(e)
            raise e

    def get_qualified_source_data(self,source_id,business_date_from,business_date_to):
        try:
            app.logger.info("Fetching qualified source data for {}..".format(source_id))
            source_table_name = self.db.query("select source_table_name from data_source_information where source_id=%s",
                                          (source_id,)).fetchone()['source_table_name']
            key_column = util.get_keycolumn(self.db._cursor(), source_table_name)
            source_data = self.db.query("select src.*, qd.* from {0} src,qualified_data qd,(select a.business_date,a.source_id,a.version,a.id_list,\
                        a.br_version from qualified_data_vers a,(select business_date,source_id,max(version) version from qualified_data_vers\
                        where business_date between %s and %s and source_id=%s group by business_date,source_id) b\
                        where a.business_date=b.business_date and a.source_id=b.source_id and a.version=b.version) qdv\
                        where qd.business_date=qdv.business_date and instr(concat(',',qdv.id_list,','),concat(',',qd.id,',')) and qd.qualifying_key=src.{1}\
                        and src.business_date=qd.business_date".format(source_table_name,key_column), (business_date_from,business_date_to,source_id)).fetchall()

            max_vers=self.db.query("select a.business_date,a.source_id,a.version,a.id_list,\
                    a.br_version from qualified_data_vers a,(select business_date,source_id,max(version) version from qualified_data_vers\
                    where business_date between %s and %s and source_id=%s group by business_date,source_id) b\
                    where a.business_date=b.business_date and a.source_id=b.source_id and a.version=b.version",(business_date_from,business_date_to,source_id))

            source_snapshot={}
            for ver in max_vers:
                source_snapshot[ver['business_date']]=ver['version']
            return source_data,source_snapshot
        except Exception as e:
            app.logger.error(e)
            raise e

    def render_trans_view_report_json(self,reporting_date='2010010120100101', version = 1):

        app.logger.info("Rendering Transactional report Template")

        try:
            app.logger.info("Getting list of sheet for report {0}".format(self.report_id))
            sheets = self.db.query("select distinct sheet_id from report_dyn_trans_def where report_id=%s",
                                   (self.report_id,)).fetchall()

            agg_format_data = {}

            sheet_d_list = []
            for sheet in sheets:
                matrix_list = []
                row_attr = {}
                col_attr = {}
                cell_style = {}
                processing_row =1
                reference_row =1
                processing_cell_id=""
                processing_dyn_sec = ""
                app.logger.info("Getting report definition for report {0},sheet {1}".format(self.report_id,sheet["sheet_id"]))
                report_template = self.db.query(
                    "select cell_id,cell_render_def,cell_calc_ref,section_id,section_type,col_id,row_id " + \
                    " from report_dyn_trans_def where report_id=%s and sheet_id=%s order by row_id, col_id, cell_render_def asc",
                    (self.report_id, sheet["sheet_id"])).fetchall()

                app.logger.info("Writing report definition to dictionary")
                for row in report_template:
                    cell_d = {}
                    if row["cell_render_def"] == 'STATIC_TEXT':
                        if row['section_type'] != 'DYNDATA':
                            processing_row = (processing_row+1) if reference_row != row['row_id'] else processing_row
                            reference_row =row['row_id'] if reference_row != row['row_id'] else reference_row
                            cell_d['cell'] = row['col_id']+str(processing_row)
                            cell_d['value'] = row['cell_calc_ref'] + (" DYNDATA("+row['section_id']+")" if row['section_type']=="DYNDATA" else "")
                            cell_d['origin'] = "TEMPLATE"
                            cell_d['section'] = row['section_id']
                            cell_d['sectionType'] = row['section_type']
                            cell_d['col'] = row['col_id']
                            cell_d['row'] = processing_row
                            matrix_list.append(cell_d)
                        else:
                            # dyn_data_sec_columns=self.get_col_list_for_row(row['row_id'])
                            # Since one increment already have been given for ROW HEIGHT and STYLE
                            processing_row = processing_row - 1
                            if processing_dyn_sec != row['section_id']:
                                sql="select * from report_dyn_trans_summary where report_id=%s and sheet_id=%s and section_id=%s and reporting_date=%s and version = %s"
                                dyn_summary_data=self.db.query(sql,(self.report_id, sheet["sheet_id"],row['section_id'],reporting_date, version)).fetchall()
                                for data in dyn_summary_data:
                                    app.logger.info("Processing row...{0}".format(processing_row));
                                    processing_row = (processing_row+1)
                                    record=eval(data['row_summary'])
                                    for col in record.keys():
                                        cell_d = {}
                                        app.logger.info("Processing row...{0} for column {1}".format(processing_row,col));
                                        cell_d['cell'] = str(col) + str(processing_row)
                                        cell_d['value'] = record[str(col)]
                                        cell_d['origin'] = "SUMMARY"
                                        cell_d['section'] = row['section_id']
                                        cell_d['sectionType'] = row['section_type']
                                        cell_d['col'] = str(col)
                                        cell_d['row'] = processing_row
                                        matrix_list.append(cell_d)
                                        cell_style[col+str(processing_row)]=eval("{'font': {'name': 'Arial', 'colour': 'None', 'size': 10.0, 'bold': False, 'italic': False}, 'border': {'bottom': {'style': None, 'colour': 'None'}, 'top': {'style': None, 'colour': 'None'}, 'left': {'style': None, 'colour': 'None'}, 'right': {'style': None, 'colour': 'None'}}, 'alignment': {'vertical': None, 'horizontal': None}, 'fill': {'colour': '00000000', 'type': None}}")
                                    row_attr[str(processing_row)] = {'height': '12.5'}
                                processing_dyn_sec = row['section_id']

                    elif row['cell_render_def'] == 'MERGED_CELL':
                        start_cell, end_cell = row['cell_id'].split(':')
                        if row['section_type'] != 'DYNDATA':
                            processing_row = (processing_row+1) if reference_row != row['row_id'] else processing_row
                            reference_row =row['row_id'] if reference_row != row['row_id'] else reference_row
                            cell_d['cell'] = start_cell
                            cell_d['value'] = row['cell_calc_ref'] + (" DYNDATA("+row['section_id']+")" if row['section_type']=="DYNDATA" else "")
                            cell_d['merged'] = end_cell
                            cell_d['origin'] = "TEMPLATE"
                            cell_d['section'] = row['section_id']
                            cell_d['sectionType'] = row['section_type']
                            cell_d['col'] = row['col_id']
                            cell_d['row'] = row['row_id']
                            matrix_list.append(cell_d)
                        else:
                            pass


                    elif row['cell_render_def'] == 'ROW_HEIGHT':
                        if row['section_type'] != 'DYNDATA':
                            processing_row = (processing_row+1) if reference_row != row['row_id'] else processing_row
                            reference_row =row['row_id'] if reference_row != row['row_id'] else reference_row
                            if row['cell_calc_ref'] == 'None':
                                row_height = '12.5'
                            else:
                                row_height = row['cell_calc_ref']
                            row_attr[row['cell_id']] = {'height': row_height}
                        else:
                            pass



                    elif row['cell_render_def'] == 'COLUMN_WIDTH':
                        if row['cell_calc_ref'] == 'None':
                            col_width = '13.88'
                        else:
                            col_width = row['cell_calc_ref']
                        col_attr[row['cell_id']] = {'width': col_width}

                    elif row['cell_render_def'] == 'CELL_STYLE':
                        if row['section_type'] != 'DYNDATA':
                            processing_row = (processing_row+1) if reference_row != row['row_id'] else processing_row
                            reference_row =row['row_id'] if reference_row != row['row_id'] else reference_row
                            if ':' in row['cell_id']:
                                start_cell, end_cell = row['cell_id'].split(':')
                            else:
                                start_cell=row['cell_id']

                            #app.logger.info("Inside CELL_STYLE for cell {}".format(start_cell,))
                            cell_style[start_cell] = eval(row['cell_calc_ref'])
                        else:
                            pass


                sheet_d = {}
                sheet_d['sheet'] = sheet['sheet_id']
                # print(sheet_d['sheet'])
                sheet_d['matrix'] = matrix_list
                sheet_d['row_attr'] = row_attr
                sheet_d['col_attr'] = col_attr
                sheet_d['cell_style'] = cell_style
                sheet_d_list.append(sheet_d)


            json_dump = (sheet_d_list)
            # print(json_dump)
            return json_dump
        except Exception as e:
            app.logger.error(str(e))
            return {"msg": str(e)}, 500

    def delete_trans_report_rules(self, data):
        app.logger.info("Deleting trans calc and agg rules")
        try:
            data_list=data['data']
            sec_rule_count = 0
            sec_order_count = 0
            for i in data_list:
                id=i['id']
                if "calc_def" in i['table_name']:
                    sec_rule_count += 1
                if "agg_def" in i['table_name']:
                    sec_order_count += 1
                res=self.dcc_tenant.update_or_delete_data(i,id)
            return {"msg": "Marked {0} sec column rules, {1} sec order rules successfully for deletion review.".format(sec_rule_count,sec_order_count)},200
        except Exception as e:
            app.logger.info(str(e))
            return {"msg":"ERROR"+str(e)},500
        return None

    def get_trans_report_audit_list(self, report_id=None, sheet_id=None, section_id=None):
        app.logger.info("Getting report audit list")
        try:
            audit_list=[]
            sql = "SELECT id FROM {0} WHERE 1 "
            if report_id:
                sql += " AND report_id='{}' ".format(report_id,)
                if sheet_id:
                    sql += " AND sheet_id='{}'".format(sheet_id,)
                if section_id:
                    sql += " AND section_id='{}'".format(section_id)
                calc_id_list = self.db.query(sql.format('report_dyn_trans_calc_def',)).fetchall()
                if calc_id_list:
                    calc_id_list = ",".join(map(str,[id['id'] for id in calc_id_list]))
                else:
                    calc_id_list = "-99999999"
                audit_list+=self.dcc_tenant.get_audit_history(id_list=calc_id_list,table_name='report_dyn_trans_calc_def')

                agg_id_list = self.db.query(sql.format('report_dyn_trans_agg_def',)).fetchall()
                if agg_id_list:
                    agg_id_list = ",".join(map(str,[id['id'] for id in agg_id_list]))
                else:
                    agg_id_list = "-99999999"
                audit_list+=self.dcc_tenant.get_audit_history(id_list=agg_id_list,table_name='report_dyn_trans_agg_def')
                return audit_list
        except Exception as e:
            app.logger.error(e)
            return {"msg": e}, 500
