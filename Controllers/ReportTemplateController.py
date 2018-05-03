from app import *
from flask_restful import Resource,abort
import os
from flask import Flask, request, redirect, url_for
from werkzeug.utils import secure_filename
import uuid
from Constants.Status import *
from Helpers.DatabaseHelper import DatabaseHelper
from Models.Document import Document
import datetime
import openpyxl as xls
import Helpers.utils as util
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Protection
from openpyxl.utils import get_column_letter
import Helpers.utils as util
import json
import ast
from operator import itemgetter
from datetime import datetime
import time
import math
import re
from Helpers.utils import autheticateTenant
from Helpers.authenticate import *

UPLOAD_FOLDER = './uploads/templates'
ALLOWED_EXTENSIONS = set(['txt', 'xls', 'xlsx'])

class ReportTemplateController(Resource):
    def __init__(self):
        self.domain_info = autheticateTenant()
        if self.domain_info:
            tenant_info = json.loads(self.domain_info)
            self.tenant_info = json.loads(tenant_info['tenant_conn_details'])
            self.db=DatabaseHelper(self.tenant_info)

    @authenticate
    def get(self):

        if request.endpoint == 'get_report_template_suggestion_list_ep':
            reports = request.args.get('reports')
            country = request.args.get('country')
            report_type = request.args.get('reportType')
            return self.report_template_suggesstion_list(report_id=reports,country=country,report_type=report_type)

    def post(self):
        try:
            if 'file' not in request.files:
                return NO_FILE_SELECTED

            self.report_id = request.form.get('report_id')
            self.report_type = request.form.get('report_type').upper()
            self.country = request.form.get('country').upper()
            self.report_description = request.form.get('report_description')
            self.db_object_suffix= request.form.get('domain_type')

            if self.report_id == None or self.report_id == "":
                return REPORT_ID_EMPTY

            if self.country == None or self.country == "":
                return COUNTRY_EMPTY

            file = request.files['file']
            self.selected_file=file.filename
            if file and not self.allowed_file(file.filename):
                return FILE_TYPE_IS_NOT_ALLOWED
            filename = str(uuid.uuid4()) + "_" + secure_filename(file.filename)
            file.save(os.path.join(UPLOAD_FOLDER, filename))
            res = self.insert_report_def_catalog()
            if res ==1:
                return self.load_report_template(filename)

            else:
                return res
        except Exception as e:
            app.logger.error(str(e))
            return {"msg":str(e)},500

    def allowed_file(self,filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

    def insert_report_def_catalog(self):
        app.logger.info("Creating entry for report catalog")
        try:
            catalog_object = "report_def_catalog"
            if self.db_object_suffix:
                catalog_object += "_" + self.db_object_suffix
            app.logger.info("Checking if report {} for country {} already exists in catalog".format(self.report_id,self.country))
            count = self.db.query("select count(*) as count from {} where report_id=%s and country=%s".format(catalog_object,),\
                                (self.report_id,self.country,)).fetchone()
            app.logger.info("Creating catalog entry for country {} and report {}".format(self.country,self.report_id))
            if not count['count']:
                res = self.db.transact("insert into {}(report_id,report_type,country,report_description) values(%s,%s,%s,%s)".format(catalog_object,),\
                        (self.report_id,self.report_type,self.country,self.report_description,))
                self.db.commit()
                return res
            return 1
        except Exception as e:
            app.logger.error(str(e))
            return {"msg":str(e)},500

    def load_report_template(self,template_file_name):
        app.logger.info("Loading report template")
        try:
            def_object = "report_def"
            if self.db_object_suffix:
                def_object += "_" + self.db_object_suffix

            formula_dict = {'SUM': 'CALCULATE_FORMULA',
                            '=SUM': 'CALCULATE_FORMULA',
                            }
            cell_render_ref = None
            target_dir = UPLOAD_FOLDER + "/"
            app.logger.info("Loading {} file from {} directory".format(template_file_name,target_dir))
            wb = xls.load_workbook(target_dir + template_file_name)

            #db = DatabaseHelper()

            sheet_index=0
            for sheet in wb.worksheets:
                sheet_index+=1
                app.logger.info("Deleting definition entries for sheet {} ,report {}".format(sheet.title,self.report_id))
                self.db.transact('delete from {} where report_id=%s and sheet_id=%s'.format(def_object,), (self.report_id, sheet.title,))

                # First capture the dimensions of the cells of the sheet
                rowHeights = [sheet.row_dimensions[r + 1].height for r in range(sheet.max_row)]
                colWidths = [sheet.column_dimensions[get_column_letter(c + 1)].width for c in range(sheet.max_column)]

                app.logger.info("Creating entries for row height")
                for row, height in enumerate(rowHeights):
                    self.db.transact('insert into {}(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref)\
                                 values(%s,%s,%s,%s,%s)'.format(def_object,), (self.report_id, sheet.title, str(row + 1), 'ROW_HEIGHT', str(height)))

                app.logger.info("Creating entries for column width")
                for col, width in enumerate(colWidths):
                    self.db.transact('insert into {}(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref)\
                                values(%s,%s,%s,%s,%s)'.format(def_object,),
                                (self.report_id, sheet.title, get_column_letter(col + 1), 'COLUMN_WIDTH', str(width)))

                app.logger.info("Creating entries for merged cells")
                rng_startcell = []
                rng_boundary=[]
                for rng in sheet.merged_cell_ranges:
                    # print rng
                    startcell, endcell = rng.split(':')
                    # print sheet.cell(startcell).border
                    rng_startcell.append(startcell)
                    rng_boundary.append(rng)
                    agg_ref='S'+str(sheet_index)+'AGG'+str(startcell)
                    _cell=sheet[startcell]
                    cell_style={"font":{"name":_cell.font.name,
                                        "bold": _cell.font.b,
                                        "italic": _cell.font.i,
                                        "colour": _cell.font.color.rgb if _cell.font.color else 'None',
                                        "size": _cell.font.sz},
                                "fill": {"type": _cell.fill.fill_type,
                                        "colour" : _cell.fill.fgColor.rgb if _cell.fill.fgColor else 'None'},
                                "alignment": {"horizontal": _cell.alignment.horizontal,
                                        "vertical": _cell.alignment.vertical},
                                "border":{"left":{"style": _cell.border.left.style,
                                                "colour":_cell.border.left.color.rgb if _cell.border.left.color else 'None'} ,
                                        "right":{"style": _cell.border.right.style,
                                                "colour":_cell.border.right.color.rgb if _cell.border.right.color else 'None'},
                                        "top":{"style": _cell.border.top.style,
                                                "colour":_cell.border.top.color.rgb if _cell.border.top.color else 'None'},
                                        "bottom":{"style": _cell.border.bottom.style,
                                                "colour":_cell.border.bottom.color.rgb if _cell.border.bottom.color else 'None'},}
                    }

                    self.db.transact('insert into {}(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref)\
                                values(%s,%s,%s,%s,%s)'.format(def_object,),
                                (self.report_id, sheet.title, rng, 'MERGED_CELL', sheet[startcell].value))
                    self.db.transact('insert into {}(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref)\
                                values(%s,%s,%s,%s,%s)'.format(def_object,),
                                (self.report_id, sheet.title, rng, 'COMP_AGG_REF', agg_ref))
                    self.db.transact('insert into {}(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref)\
                                values(%s,%s,%s,%s,%s)'.format(def_object,),
                                (self.report_id, sheet.title, rng, 'CELL_STYLE', str(cell_style)))

                app.logger.info("Creating entries for static text and formulas")
                for all_obj in sheet['A1':util.cell_index(sheet.max_column, sheet.max_row)]:
                    for cell_obj in all_obj:
                        cell_ref = str(cell_obj.column) + str(cell_obj.row)
                        agg_ref='S'+str(sheet_index)+'AGG'+str(cell_ref)
                        _cell=sheet[cell_ref]
                        cell_style={"font":{"name":_cell.font.name,
                                            "bold": _cell.font.b,
                                            "italic": _cell.font.i,
                                            "colour": _cell.font.color.rgb if _cell.font.color else 'None',
                                            "size": _cell.font.sz},
                                    "fill": {"type": _cell.fill.fill_type,
                                            "colour" : _cell.fill.fgColor.rgb if _cell.fill.fgColor else 'None'},
                                    "alignment": {"horizontal": _cell.alignment.horizontal,
                                            "vertical": _cell.alignment.vertical},
                                    "border":{"left":{"style": _cell.border.left.style,
                                                    "colour":_cell.border.left.color.rgb if _cell.border.left.color else 'None'} ,
                                            "right":{"style": _cell.border.right.style,
                                                    "colour":_cell.border.right.color.rgb if _cell.border.right.color else 'None'},
                                            "top":{"style": _cell.border.top.style,
                                                    "colour":_cell.border.top.color.rgb if _cell.border.top.color else 'None'},
                                            "bottom":{"style": _cell.border.bottom.style,
                                                    "colour":_cell.border.bottom.color.rgb if _cell.border.bottom.color else 'None'},}
                        }
                        if (len(rng_startcell) > 0 and cell_ref not in rng_startcell) or (len(rng_startcell) == 0):
                            if cell_obj.value:
                                for key in formula_dict.keys():
                                    cell_obj_value = str(cell_obj.value)
                                    if key in cell_obj_value:
                                        cell_render_ref = formula_dict[key]
                                        break
                                    else:
                                        cell_render_ref = 'STATIC_TEXT'

                                self.db.transact('insert into {}(report_id,sheet_id ,cell_id ,cell_render_def ,cell_calc_ref)\
                                          values(%s,%s,%s,%s,%s)'.format(def_object,),
                                            (self.report_id, sheet.title, cell_ref, cell_render_ref, cell_obj_value.strip()))

                            if not self.check_if_cell_isin_range(cell_ref,rng_boundary):
                                self.db.transact('insert into {}(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref)\
                                                 values(%s,%s,%s,%s,%s)'.format(def_object,),(self.report_id, sheet.title, cell_ref, 'COMP_AGG_REF', agg_ref))
                                self.db.transact('insert into {}(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref)\
                                                 values(%s,%s,%s,%s,%s)'.format(def_object,),(self.report_id, sheet.title, cell_ref, 'CELL_STYLE', str(cell_style)))

            self.db.commit()
            return {"msg": "Report [" + self.report_id + "] template has been captured successfully using " + self.selected_file}, 200
        except Exception as e:
            app.logger.error(str(e))
            return {"msg": str(e)}, 500

    def check_if_cell_isin_range(self,cell_id,rng_boundary_list):

        def number_from_word(value):
            base='ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            result=0
            i=len(value)-1
            for j in range(0,len(value)):
              result+=math.pow(len(base),i)*(base.index(value[j])+1)
              i-=1
            return result

        def split_numbers_chars(string):
            sub=re.split('(\d+)', string)
            return sub

        def check_inclusion(cell,rng):
            start_cell, end_cell = rng.split(':')
            start_cell_sub=split_numbers_chars(start_cell)
            end_cell_sub=split_numbers_chars(end_cell)
            cell_sub=split_numbers_chars(cell)

            within_rng=False

            if number_from_word(cell_sub[0]) >=number_from_word(start_cell_sub[0]) \
            and number_from_word(cell_sub[0]) <= number_from_word(end_cell_sub[0]) \
            and  cell_sub[1] >= start_cell_sub[1] and cell_sub[1] <= end_cell_sub[1]:
                within_rng=True

            return within_rng

        for rng in rng_boundary_list:
            incl=check_inclusion(cell_id,rng)
            if incl:
                break
        return incl


    def report_template_suggesstion_list(self,report_id='ALL',country='ALL',report_type='ALL'):

        app.logger.info("Getting report template list")

        try:
            #db=DatabaseHelper()
            data_dict={}
            where_clause = ''

            sql = "select distinct country from report_def_catalog where 1 "
            country_suggestion = self.db.query(sql).fetchall()
            if country is not None and country !='ALL':
                 where_clause =  " and instr('" + country.upper() + "', upper(country)) > 0"
            if report_id is not None and report_id !='ALL':
                 where_clause +=  " and instr('" + report_id.upper() + "', upper(report_id)) > 0"
            if report_type is not None and report_type !='ALL':
                 where_clause +=  " and instr('" + report_type.upper() + "', upper(report_type)) > 0"

            app.logger.info("Getting list of countries")
            country = self.db.query(sql + where_clause).fetchall()

            # sql = "select distinct report_id from report_def_catalog"
            # report_suggestion = db.query(sql).fetchall()
            data_dict['country'] = country
            for i,c in enumerate(data_dict['country']):
                # sql = "select distinct report_id, report_description from report_def_catalog where country = '" + c['country'] + "'"
                sql = "select * from report_def_catalog where country = '" + c['country'] + "'"

                if report_id is not None and report_id !='ALL':
                     where_clause =  " and instr(upper('" + report_id + "'), upper(report_id)) > 0"
                if report_type is not None and report_type !='ALL':
                     where_clause =  " and instr(upper('" + report_type + "'), upper(report_type)) > 0"

                app.logger.info("Getting list of reports for country {}".format(data_dict["country"]))
                report = self.db.query(sql + where_clause).fetchall()
                #print(data_dict['country'][i])
                data_dict['country'][i]['report'] = report
                where_report = ''
                # for j,r in enumerate(data_dict['country'][i]['report']):
                #     sql = "select distinct report_id, valid_from, valid_to, last_updated_by,'{0}' as report_description from report_def where 1 "
                #     sql = sql.format(data_dict['country'][i]['report'][j]['report_description'])
                #     where_report =  " and report_id = '" + data_dict['country'][i]['report'][j]['report_id'] + "'"
                #     app.logger.info("Getting different version for report {}".format(data_dict['country'][i]['report'][j]))
                #     reportversions = self.db.query(sql + where_report).fetchone()
                #     #print(data_dict['country'][i]['report'][j])
                #     data_dict['country'][i]['report'][j] = reportversions
                #print(data_dict)
            #data_dict['report_suggestion'] = report_suggestion
            #data_dict['country_suggestion'] = country_suggestion

            if not data_dict:
                return {"msg":"No report matched found"},404
            else:
                return data_dict['country']
        except Exception as e:
            app.logger.error(e)
            return {"msg": e}, 500
