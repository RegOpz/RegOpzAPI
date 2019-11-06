from app import *
from flask_restful import Resource,abort
from Helpers.DatabaseHelper import DatabaseHelper
import openpyxl as xls
from openpyxl.utils import column_index_from_string,get_column_letter,coordinate_from_string,coordinate_to_tuple
import Helpers.utils as util
import json
from Helpers.utils import autheticateTenant
from Helpers.authenticate import *
import pandas as pd
from Constants.Status import *
import os
import uuid
from werkzeug.utils import secure_filename

UPLOAD_FOLDER='./uploads/templates'

class FFCaptureTemplateController(Resource):
    def __init__(self):
        self.domain_info = autheticateTenant()
        self.db_master=DatabaseHelper()
        if self.domain_info:
            tenant_info = json.loads(self.domain_info)
            self.tenant_info = json.loads(tenant_info['tenant_conn_details'])
            self.db=DatabaseHelper(self.tenant_info)

    def post(self, report_id=None):
        if request.endpoint=="capture_xls_template_ep":
            #app.logger.info(request)
            if 'file' not in request.files:
                return NO_FILE_SELECTED

            self.report_id = request.form.get('report_id')
            self.report_type = request.form.get('report_type').upper()
            self.country = request.form.get('country').upper()
            self.report_description = request.form.get('report_description')
            self.db_object_suffix = request.form.get('domain_type')

            if self.report_id == None or self.report_id == "":
                return REPORT_ID_EMPTY

            if self.country == None or self.country == "":
                return COUNTRY_EMPTY

            file = request.files['file']
            self.selected_file = file.filename
            filename = str(uuid.uuid4()) + "_" + secure_filename(file.filename)
            file.save(os.path.join(UPLOAD_FOLDER, filename))

            res = self.insert_report_def_catalog()
            if res:
                return self.load_report_template(filename)
            else:
                return res

        if request.endpoint=="capture_hot_template_ep":
            self.report_id = report_id
            self.db_object_suffix = request.args.get('domain_type')
            report_data = request.get_json(force=True)
            return self.save_hot_table_report_template(report_data)

    def insert_report_def_catalog(self):
        app.logger.info("Creating entry for report catalog")
        try:
            catalog_object = "report_def_catalog"
            if self.db_object_suffix:
                catalog_object += "_" + self.db_object_suffix
            app.logger.info("Checking if report {} for country {} already exists in catalog".format(self.report_id,self.country))
            count = self.db.query("select count(*) as count from {} where report_id=%s and country=%s".format(catalog_object,),\
                                (self.report_id,self.country,)).fetchone()
            app.logger.info("Creating catalog entry for country {} and report {}".format(self.country,self.report_id))
            if not count['count']:
                res = self.db.transact("insert into {}(report_id,report_type,country,report_description) values(%s,%s,%s,%s)".format(catalog_object,),\
                        (self.report_id,self.report_type,self.country,self.report_description,))
                # self.db.commit()
                return res
            return 1
        except Exception as e:
            app.logger.error(str(e))
            return {"msg":str(e)},500

    def extract_content_type(self,val):
        if not val:
            return '', 'BLANK'
        return val, 'STATIC_TEXT'

    def load_report_template(self,template_file_name):

        app.logger.info("Loading report template")
        try:
            def_object ="report_free_format_def"
            ref_object="report_free_format_ref"
            sec_object="report_free_format_section"

            if self.db_object_suffix:
                def_object += "_" + self.db_object_suffix
                ref_object += "_" + self.db_object_suffix
                sec_object += "_" + self.db_object_suffix

            ref_df = pd.DataFrame()
            def_df = pd.DataFrame()

            target_dir = UPLOAD_FOLDER + "/"
            app.logger.info("Loading {} file from {} directory".format(template_file_name, target_dir))
            wb = xls.load_workbook(target_dir + template_file_name)

            sheet_index = 0
            for sheet in wb.worksheets:

                app.logger.info("Processing definition entries for sheet {} ,report {}".format(sheet.title, self.report_id))

                row_heights = [sheet.row_dimensions[r + 1].height for r in range(sheet.max_row)]
                col_widths = [sheet.column_dimensions[get_column_letter(c + 1)].width for c in range(sheet.max_column)]

                app.logger.info("Creating entries for row height")
                for row, height in enumerate(row_heights):
                    def_df = def_df.append({'report_id': self.report_id, 'sheet_id': sheet.title, 'entity_type': 'ROW', \
                         'entity_ref': str(row+1), 'content_type': 'ROW_HEIGHT','content': str(height)}, ignore_index=True)

                app.logger.info("Creating entries for column width")
                for col, width in enumerate(col_widths):
                    def_df = def_df.append({'report_id': self.report_id, 'sheet_id': sheet.title, 'entity_type': 'COL',
                         'entity_ref': get_column_letter(col + 1), 'content_type': 'COLUMN_WIDTH', 'content': str(width)}, ignore_index=True)

                merged_cell = {}
                for cell_rng in sheet.merged_cell_ranges: #move to 2.5.14 sheet.merged_cells.ranges
                    # print rng
                    startcell, endcell = cell_rng.split(':')
                    # print sheet.cell(startcell).border
                    start_xy = coordinate_to_tuple(startcell)
                    end_xy=coordinate_to_tuple(endcell)
                    start_row,start_col = start_xy
                    end_row,end_col = end_xy
                    for r in range(start_row,end_row+1):
                        for c in range(start_col, end_col+1):
                            cell = get_column_letter(c) + str(r)
                            merged_cell[cell] = {'cell_rng': cell_rng, 'start_cell': startcell}

                cell_added=[]
                sheet_index += 1
                ref=0
                for row in sheet.rows:
                    for cell_idx in row:
                        cell = str(cell_idx.column) + str(cell_idx.row)
                        row = cell_idx.row
                        col = str(cell_idx.column)

                        is_merged_cell=False
                        if cell in merged_cell.keys():
                            cell_id = merged_cell[cell]['cell_rng']
                            start_cell=merged_cell[cell]['start_cell']
                            is_merged_cell = True
                        else:
                            cell_id = cell
                            start_cell=cell


                        try :
                            idx=cell_added.index(cell_id)
                        except ValueError:
                            ref+=1
                            cell_ref='S'+str(sheet_index)+'{:04d}'.format(ref)
                            content,content_type=self.extract_content_type(cell_idx.value)
                            ref_df=ref_df.append({'report_id':self.report_id,'sheet_id':sheet.title,'cell_id':cell_id,
                                           'col_id':col,'row_id':row,'cell_type':'MERGED' if is_merged_cell else 'SINGLE',
                                           'cell_ref':cell_ref,'section_id':None,'section_type':None,},ignore_index=True)
                            def_df=def_df.append({'report_id':self.report_id,'sheet_id':sheet.title,'entity_type':'CELL',
                                                  'entity_ref':cell_ref,'content_type':content_type,'content':content},ignore_index=True)
                            cell_style = util.get_css_style_from_openpyxl(cell_idx)
                            def_df=def_df.append({'report_id':self.report_id,'sheet_id':sheet.title,'entity_type':'CELL',\
                                                  'entity_ref':cell_ref,'content_type':'CELL_STYLE','content':json.dumps(cell_style)},ignore_index=True)
                            cell_added.append(cell_id)

            app.logger.info("Deleting definition entries for report {}".format(self.report_id, ))
            self.db.transact('delete from {} where report_id=%s'.format(def_object, ), (self.report_id,))
            self.db.transact('delete from {} where report_id=%s'.format(ref_object), (self.report_id,))
            self.db.transact('delete from {} where report_id=%s'.format(sec_object), (self.report_id,))

            def_df.fillna('',inplace=True)
            ref_df.fillna('',inplace=True)

            self.db.transactmany("insert into {0}({1}) values({2})".format(def_object, ",".join(def_df.columns),",".join(['%s'] * len(def_df.columns))),
                             list(def_df.itertuples(index=False, name=None)))

            self.db.transactmany("insert into {0}({1}) values({2})".format(ref_object, ",".join(ref_df.columns),",".join(['%s'] * len(ref_df.columns))),
                             list(ref_df.itertuples(index=False, name=None)))

            self.db.commit()

            return {"msg": "Report [" + self.report_id + "] template updates has been captured successfully "}, 200
        except Exception as e:
            app.logger.error(str(e))
            #return {"msg": str(e)}, 500
            raise e


    def save_hot_table_report_template(self,report_data):
        app.logger.info("Saving hot table report template")
        try:
            def cell_section(sec_df,cell_rng):
                cells=cell_rng.split(':')
                for cell in cells:
                    r,c=coordinate_to_tuple(cell)
                    for row in sec_df:
                        if row['start_row']<= r-1 and row['end_row'] >= r-1 and row['start_col'] <= c-1 and row['end_col'] >= c-1:
                            return row['section_id'],row['section_type']
                return None,None


            ref_object ="report_free_format_ref"
            def_object="report_free_format_def"
            sec_object="report_free_format_section"

            if self.db_object_suffix:
                def_object += "_" + self.db_object_suffix
                ref_object += "_" + self.db_object_suffix
                sec_object += "_" + self.db_object_suffix
                self.db = self.db_master

            ref_df=pd.DataFrame()
            def_df=pd.DataFrame()
            sec_df2 = pd.DataFrame()

            sheet_index =0
            for sheet in report_data:
                #{"ht_range": [0, 0, 0, 1], "range": "A1:B1", "section_id": "s1", "section_position": [], "section_type": "FIXEDFORMAT"}
                sec_df=[]
                for sec in sheet['sections']:
                    #print(sec)
                    sec_df.append({'section_id':sec['section_id'],'section_type':sec['section_type'],'start_row':sec['ht_range'][0],
                                'start_col':sec['ht_range'][1],'end_row':sec['ht_range'][2],'end_col':sec['ht_range'][3]})
                    sec_df2=sec_df2.append({'report_id':self.report_id ,'sheet_id':sheet['sheet'],'section_id':sec['section_id'] ,
                                    'section_range':sec['range'],'section_ht_range':",".join(map(str,sec['ht_range'])),
                                    'section_dependency':",".join(sec['section_position']),'section_type':sec['section_type']},ignore_index=True)

                merged_cell = {}
                for rng in sheet['mergedCells']:
                    # print rng
                    startcell = get_column_letter(rng['col'] + 1) + str(rng['row'] + 1)
                    endcell = get_column_letter(rng['col'] + rng['colspan']) + str(rng['row'] + rng['rowspan'])
                    cell_rng = startcell + ':' + endcell
                    start_row = rng['row'] + 1
                    start_col = get_column_letter(rng['col'] + 1)

                    for r in range(rng['row']+1,rng['row']+rng['rowspan']+1):
                      for c in range(rng['col']+1,rng['col']+rng['colspan']+1):
                          cell=get_column_letter(c)+str(r)
                          merged_cell[cell] = {'cell_rng': cell_rng,'start_cell':startcell}

                cell_added=[]
                sheet_index+=1
                ref=0
                for r,cols in enumerate(sheet['sheetData']):
                    for c,val in enumerate(cols):
                        is_merged_cell = False
                        row = r+1
                        col = get_column_letter(c+1)
                        cell = col + str(row)
                        if cell in merged_cell.keys():
                            cell_id = merged_cell[cell]['cell_rng']
                            start_cell=merged_cell[cell]['start_cell']
                            col,row=coordinate_from_string(start_cell)
                            is_merged_cell = True
                        else:
                            cell_id = cell
                            start_cell=cell
                        section,section_type=cell_section(sec_df,cell_id)
                        try :
                            idx=cell_added.index(cell_id)
                        except ValueError:
                            ref+=1
                            cell_ref='S'+'{:02d}'.format(sheet_index)+'{:04d}'.format(ref)
                            content,content_type=self.extract_content_type(val)
                            ref_df=ref_df.append({'report_id':self.report_id,'sheet_id':sheet['sheet'],'cell_id':cell_id,
                                           'col_id':col,'row_id':row,'cell_type':'MERGED' if is_merged_cell else 'SINGLE',
                                           'cell_ref':cell_ref,'section_id':section,'section_type':section_type,},ignore_index=True)
                            def_df=def_df.append({'report_id':self.report_id,'sheet_id':sheet['sheet'],'entity_type':'CELL',
                                                  'entity_ref':cell_ref,'content_type':content_type,'content':content},ignore_index=True)
                            if start_cell in sheet['sheetStyles'].keys():
                                cell_style = sheet['sheetStyles'][start_cell]
                                def_df=def_df.append({'report_id':self.report_id,'sheet_id':sheet['sheet'],'entity_type':'CELL',\
                                                  'entity_ref':cell_ref,'content_type':'CELL_STYLE','content':json.dumps(cell_style)},ignore_index=True)
                            cell_added.append(cell_id)



                app.logger.info("Creating entries for row height")
                for row in sheet['rowHeights'].keys():
                    def_df = def_df.append({'report_id': self.report_id, 'sheet_id': sheet['sheet'], 'entity_type': 'ROW',\
                'entity_ref':str(row),'content_type':'ROW_HEIGHT', 'content': str(sheet['rowHeights'][str(row)])}, ignore_index = True)

                app.logger.info("Creating entries for column width")
                for col in sheet['colWidths'].keys():
                    def_df = def_df.append(
                        {'report_id': self.report_id, 'sheet_id': sheet['sheet'], 'entity_type': 'COL','entity_ref': str(col), 'content_type': 'COLUMN_WIDTH',\
                         'content': str(sheet['colWidths'][col]/8)}, ignore_index=True)


            app.logger.info("Deleting definition entries for report {}".format(self.report_id, ))
            self.db.transact('delete from {} where report_id=%s'.format(def_object),(self.report_id,))
            self.db.transact('delete from {} where report_id=%s'.format(ref_object), (self.report_id,))
            self.db.transact('delete from {} where report_id=%s'.format(sec_object), (self.report_id,))

            self.db.transactmany("insert into {0}({1}) values({2})".format(def_object,",".join(def_df.columns),
                             ",".join(['%s'] * len(def_df.columns))),list(def_df.itertuples(index=False, name=None)))

            self.db.transactmany("insert into {0}({1}) values({2})".format(ref_object, ",".join(ref_df.columns),
                            ",".join(['%s'] * len(ref_df.columns))),list(ref_df.itertuples(index=False, name=None)))
            self.db.transactmany("insert into {0}({1}) values({2})".format(sec_object,",".join(sec_df2.columns),
                            ",".join(['%s'] * len(sec_df2.columns))),list(sec_df2.itertuples(index=False, name=None)))
            self.db.commit()

            return {"msg": "Report [" + self.report_id + "] template updates has been captured successfully "},200
        except Exception as e:
            app.logger.error(str(e))
            raise e
            # return {"msg": str(e)}, 500

    def update_content_type(self,data,db=None,content_type=None,check_before_update=True):
        app.logger.info("Updating content type for {}".format(data,))
        if not db:
            return True
        # Get the rule record to obtain report_id, sheet_id and cell_id for the rule in concern
        sql = "select * from {0} where id={1}".format(data['table_name'],data['id'])
        rule = db.query(sql).fetchone()
        if check_before_update:
            # Now check whether there is any existing valid rule for the cell_id
            sql = ("select cell_id from report_calc_def where report_id='{0}' and sheet_id='{1}' and cell_id='{2}' and in_use='Y' " + \
                " union " + \
                "select cell_id from report_comp_agg_def where report_id='{0}' and sheet_id='{1}' and cell_id='{2}' and in_use='Y' " + \
                "").format(rule['report_id'], rule['sheet_id'],rule['cell_id'])
            rule_in_use = db.query(sql).fetchall()
            # If no valid rule exits on the cell_id then mark it as BLANK cell
            if not rule_in_use:
                content_type = 'BLANK'

        # Now that content type is decided, set the value in def table
        app.logger.info("Updating content type for {0} {1} {2} {3}".format(content_type,rule['report_id'], rule['sheet_id'],rule['cell_id']))
        sql = ("update report_free_format_def set content_type='{0}', content='' " + \
            " where report_id='{1}' and sheet_id='{2}' and entity_ref='{3}'" + \
            " and content_type in ('BLANK','DATA','STATIC_TEXT')").format(content_type,rule['report_id'], rule['sheet_id'],rule['cell_id'])
        db.transact(sql)
