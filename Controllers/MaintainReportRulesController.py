from app import *
from flask import Flask, jsonify, request
from flask_restful import Resource
from Helpers.DatabaseHelper import DatabaseHelper
from Helpers.DatabaseOps import DatabaseOps
from Helpers.AuditHelper import AuditHelper
import csv
import time
from datetime import datetime
from Constants.Status import *
import openpyxl as xls
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Protection
from openpyxl.utils import get_column_letter
import json

class MaintainReportRulesController(Resource):
	def __init__(self):
		#print("Inside MaintainReportRulesController", request.headers.get('Tenant'))
		tenant_info = json.loads(request.headers.get('Tenant'))
		self.tenant_info = json.loads(tenant_info['tenant_conn_details'])
		self.dbOps=DatabaseOps('def_change_log',self.tenant_info)
		self.audit=AuditHelper('def_change_log',self.tenant_info)
		self.db=DatabaseHelper(self.tenant_info)

	def get(self,report_id=None):
		#print(request.endpoint)
		if request.endpoint == 'get_business_rules_suggestion_list_ep':
			source_id = request.args.get('source_id')
			return self.get_business_rules_suggestion_list(source_id=source_id)

		if request.endpoint == 'get_source_suggestion_list_ep':
			source_id = request.args.get('source_id')
			return self.get_source_suggestion_list(source_id=source_id)

		if request.endpoint == 'get_cell_calc_ref_suggestion_list_ep':
			report_id = request.args.get('report_id')
			return self.get_cell_calc_ref_suggestion_list(report_id=report_id)

		if request.endpoint == 'get_agg_function_column_suggestion_list_ep':
			table_name = request.args.get('table_name')
			return self.get_agg_function_column_suggestion_list(table_name=table_name)

		if request.endpoint == "report_rule_audit_ep":
			report_id = request.args.get('report_id')
			sheet_id = request.args.get('sheet_id')
			cell_id = request.args.get('cell_id')
			return self.get_report_audit_list(report_id=report_id, sheet_id=sheet_id, cell_id=cell_id)

		if request.endpoint == 'get_report_rule_export_to_excel_ep':
			self.report_id = request.args.get('report_id')
			return self.export_rules_to_excel()


	def post(self):
		data = request.get_json(force=True)
		res = self.dbOps.insert_data(data)
		return res

	def put(self, id=None, report=None):
		if id == None and report == None:
			return BUSINESS_RULE_EMPTY
		data = request.get_json(force=True)
		if request.endpoint == "report_rule_ep":
			res = self.dbOps.update_or_delete_data(data, id)
			return res
		if request.endpoint == "report_parameter_ep":
			return self.update_report_parameters(data, report)

	# def delete(self, id=None):
	# 	if id == None:
	# 		return BUSINESS_RULE_EMPTY
	# 	if request.endpoint == "report_rule_ep":
	# 		table_name = request.args.get("table_name")
	# 		data=request.get_json(force=True)
	# 		#comment=request.headers["comment"]
	# 		print("++++++++++++++")
	# 		print(data)
	# 		print("++++++++++++++")
    #
	# 		res = self.dbOps.delete_data(table_name,id)
	# 		return res

	def update_report_parameters(self,data,report):
		app.logger.info("Updating Report Parameters in Report Def Catalog {}".format(data,))
		sql = "update report_def_catalog set report_parameters='{0}' where report_id='{1}'"
		sql = sql.format(data['report_parameters'],report)
		try:
			self.db.transact(sql)
			self.db.commit()
			return {"msg": "Report parameters updated successfully for {0}".format(report,)}, 200
		except Exception as e:
			return {"msg": str(e)},500




	def get_source_suggestion_list(self,source_id='ALL'):

		app.logger.info("Getting source suggestion list")
		try:
			data_dict={}
			where_clause = ''

			sql = "select *  from data_source_information where 1 "
			if source_id is not None and source_id !='ALL':
				 where_clause =  " and source_id = " + source_id

			app.logger.info("Retrieving source suggestion")
			source = self.db.query(sql + where_clause).fetchall()
			data_dict['source_suggestion'] = source

			if not data_dict['source_suggestion']:
				return {"msg":"No matched sources found"},404
			else:
				return data_dict
		except Exception as e:
			app.logger.error(e)
			return {"msg": e}, 500

	def get_business_rules_suggestion_list(self,source_id='ALL'):

		app.logger.info("Getting business rule suggestion list")
		try:
			data_dict={}
			where_clause = ''

			sql = "select source_id, source_table_name from data_source_information where 1 "

			if source_id is not None and source_id !='ALL':
				 where_clause =  " and source_id = " + str(source_id)
			app.logger.info("Retrieving source list")
			source = self.db.query(sql + where_clause).fetchall()

			data_dict['source_suggestion'] = source
			for i,s in enumerate(data_dict['source_suggestion']):
				sql = "select * from business_rules where source_id = '{0}' ".format(str(s['source_id']))
				app.logger.info("Retrieving business rules for source {}".format(s['source_id']))
				rules_suggestion = self.db.query(sql).fetchall()
				data_dict['source_suggestion'][i]['rules_suggestion'] = rules_suggestion

			if not data_dict['source_suggestion']:
				return {"msg":"No sources found"},404
			else:
				return data_dict
		except Exception as e:
			app.logger.error(e)
			return {"msg": e}, 500

	def get_agg_function_column_suggestion_list(self,table_name):
		app.logger.info("Getting column suggestion list for agrregate function")
		try:

			app.logger.info("Retrieving column list for report_qualified_data_link")
			data_dict_report_link = self.db.query("describe report_qualified_data_link").fetchall()
			app.logger.info("Retrieving column list for table {}".format(table_name))
			data_dict = self.db.query("describe " + table_name).fetchall()

			#Now build the agg column list
			return data_dict + data_dict_report_link
		except Exception as e:
			app.logger.error(e)
			return {"msg": e}, 500

	def get_cell_calc_ref_suggestion_list(self,report_id):
		app.logger.info("Getting suggestion list for cell calc ref")
		try:
			data_dict={}
			sql = "select * from report_calc_def where report_id = '{}'".format(report_id)
			app.logger.info("Retrieving cell calc ref for report {}".format(report_id))
			cell_calc_ref_list = self.db.query(sql).fetchall()
			data_dict['cell_calc_ref'] = cell_calc_ref_list

			if not data_dict['cell_calc_ref']:
				return {"msg":"No report matched found"},404
			else:
				return data_dict
		except Exception as e:
			app.logger.error(e)
			return {"msg": e}, 500

	def get_report_audit_list(self, report_id=None, sheet_id=None, cell_id=None):
		app.logger.info("Getting report audit list")
		try:
			if report_id:
				calc_query = "SELECT id,'report_calc_def' FROM report_calc_def WHERE report_id=%s"
				comp_query = "SELECT id,'report_comp_agg_def' FROM report_comp_agg_def WHERE report_id=%s"
				queryParams = (report_id, report_id)
				if sheet_id:
					calc_query += " AND sheet_id=%s"
					comp_query += " AND sheet_id=%s"
					queryParams = (report_id, sheet_id, report_id, sheet_id,)
				if cell_id:
					calc_query += " AND cell_id=%s"
					comp_query += " AND cell_id=%s"
					queryParams = (report_id, sheet_id, cell_id, report_id, sheet_id, cell_id,)
				queryString = "SELECT DISTINCT id,table_name,change_type,change_reference,date_of_change,\
					maker,maker_comment,checker,checker_comment,status,date_of_checking FROM def_change_log\
					WHERE (id,table_name) IN (" + calc_query + " UNION " + comp_query + " )"
				return self.audit.get_audit_list(queryString, queryParams)
		except Exception as e:
			app.logger.error(e)
			return {"msg": e}, 500

	def export_rules_to_excel(self):
		app.logger.info("Exporting report rules to excel sheet")
		try:
			report_id = self.report_id
			target_file_name = report_id + '_report_rules' + str(time.time()) + '.xlsx'

			# Create report template
			wr = xls.Workbook()
			# target_dir='../output/'
			target_dir = './static/'
			#db = DatabaseHelper()

			# print sheets
			# The default sheet of the workbook
			al = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
			ws = wr.worksheets[0]
			for sheet in [{'table_name': 'report_calc_def'}, {'table_name': 'report_comp_agg_def'}]:
				# The first sheet title will be Sheet, so do not create any sheet, just rename the title
				if ws.title == 'Sheet':
					ws.title = sheet["table_name"]
				else:
					ws = wr.create_sheet(title=sheet["table_name"])
				app.logger.info("Retrieving data from {0} for report {1}".format(sheet["table_name"],report_id))
				cur = self.db.query('select * from ' + sheet["table_name"] + ' where report_id=%s ',
							   (report_id,))
				report_rules = cur.fetchall()

				cols = []
				for i, c in enumerate(cur.description):
					cols.append({'cell_id': get_column_letter(i + 1), 'cell_name': c[0]})
					print(cols, cols[i]['cell_id'] + '1')
					ws[cols[i]['cell_id'] + '1'].value = cols[i]['cell_name']
					ws[cols[i]['cell_id'] + '1'].fill = PatternFill("solid", fgColor="DDDDDD")
					ws[cols[i]['cell_id'] + '1'].font = Font(bold=True, size=9)
				for row, rule in enumerate(report_rules):
					for c in cols:
						ws[c['cell_id'] + str(row + 2)].value = rule[c['cell_name']]
						ws[cols[i]['cell_id'] + str(row + 2)].font = Font(bold=True, size=9)

			app.logger.info("Saving excel sheet to file {}".format(target_dir + target_file_name))
			wr.save(target_dir + target_file_name)
			# End create report template

			return {"file_name": target_file_name}
		except Exception as e:
			app.logger.error(e)
			return {"msg": e}, 500
