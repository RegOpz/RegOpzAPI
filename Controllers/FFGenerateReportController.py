from flask_restful import Resource
from app import *
import Helpers.utils as util
from collections import defaultdict
import pandas as pd
from Helpers.DatabaseHelper import DatabaseHelper
import json
from Helpers.utils import autheticateTenant
from Helpers.authenticate import *
from Parser.ExcelLib import *
from Controllers.OperationalLogController import OperationalLogController
from Controllers.FFRenderController import FFRenderController
from Controllers.FFCreateFixedFormatController import FFCreateFixedFormatController
import openpyxl as xls
from openpyxl.utils import column_index_from_string,get_column_letter,coordinate_from_string,coordinate_to_tuple

class FFGenerateReportController(Resource):
    def __init__(self):
        self.domain_info = autheticateTenant()
        if self.domain_info:
            tenant_info = json.loads(self.domain_info)
            self.tenant_info = json.loads(tenant_info['tenant_conn_details'])
            self.db=DatabaseHelper(self.tenant_info)
            self.user_id=Token().authenticate()

        self.opsLog = OperationalLogController()
        self.log_master_id = None

    @authenticate
    def get(self):
        if(request.endpoint=='get_report_list_ep'):
            country=request.args.get('country') if request.args.get('country') != None else 'ALL'
            return self.get_report_list(country)
        if (request.endpoint == 'get_country_list_ep'):
            return self.get_country_list()

    def post(self):
        report_parameters=request.get_json(force=True)
        return self.create_report(report_parameters)



    def  create_report(self, report_parameters):
        try:
            app.logger.info("Creating report")
            report_id = report_parameters["report_id"]
            reporting_date = report_parameters["reporting_date"]
            self.reporting_currency = report_parameters["reporting_currency"]
            report_version_no=self.create_report_catalog(report_parameters)

            self.update_report_catalog(status='RUNNING'
                                        , report_parameters=report_parameters
                                        , version=report_version_no)
            # report_snapshot=self.create_report_snapshot(report_version_no,**report_parameters)
            report_snapshot=self.create_report_snapshot(**report_parameters)
            self.update_report_catalog(status='SNAPSHOTCREATED'
                                        , report_parameters=report_parameters
                                        , report_snapshot=report_snapshot
                                        , version=report_version_no)
            self.create_report_detail(report_version_no,report_snapshot,**report_parameters)

            self.update_report_catalog(status='SUCCESS'
                                        , report_parameters=report_parameters
                                        , report_snapshot=report_snapshot
                                        , version=report_version_no)
            if self.log_master_id:
                self.opsLog.write_log_detail(master_id=self.log_master_id
                    , operation_sub_type='End of Create Report'
                    , operation_status='Complete'
                    , operation_narration="Report generated SUCCESSFULLY for [{0}] Reporting date [{1}].".format(str(report_id),str(reporting_date))
                    )
                self.opsLog.update_master_status(id=self.log_master_id,operation_status="SUCCESS")

            return {"msg": "Report generated SUCCESSFULLY for [{0}] Reporting date [{1}].".format(str(report_id),str(reporting_date))}, 200
        except Exception as e:
            app.logger.error(str(e))
            if self.log_master_id:
                self.opsLog.write_log_detail(master_id=self.log_master_id
                    , operation_sub_type='Error occured while creating report'
                    , operation_status='Failed'
                    , operation_narration="Report not generated  for [{0}] Reporting date [{1}].".format(str(report_id),str(reporting_date))
                    )
                self.opsLog.update_master_status(id=self.log_master_id,operation_status="ERROR")
            return {'msg': str(e)}, 400
            # raise e

    def create_report_detail(self,report_version_no,report_snapshot,**report_parameters):
        try:
            report_id=report_parameters['report_id']
            reporting_date=report_parameters['business_date_from']+report_parameters['business_date_to']
            render_controller=FFRenderController()
            fixed_controller=FFCreateFixedFormatController()
            wb_data=xls.Workbook()
            # report_id=report_parameters['report_id']
            sec_object = "report_free_format_section"
            ff_summary_object="report_free_format_summary"

            sec_df = self.db.query("select sheet_id,section_id,section_type,section_range,section_ht_range,section_dependency\
                     from {} where report_id=%s".format(sec_object),(report_id,)).fetchall()

            fixed_created=False
            for sec in sec_df:
                if sec['section_type']=="FIXEDFORMAT" and not fixed_created:
                    fixed_controller.create_fixed_format_sections(report_version_no,report_snapshot,self.log_master_id,**report_parameters)
                    fixed_created=True

            wb_data =render_controller.render_report_intermediate(wb_data,report_version_no, **report_parameters)
            # Delete the first default Sheet
            wb_data.remove(wb_data.get_sheet_by_name('Sheet'))

            ff_summary_df=pd.DataFrame()
            for sheet in wb_data.sheetnames:
                if sheet[sheet.find('-template'):]=="-template":
                    continue

                ws=wb_data[sheet]
                ws_template=wb_data[sheet+'-template']

                # for row in ws_template.rows:
                #     for col in row:
                #         tcell = str(col.column) + str(col.row)
                #         app.logger.info("cell_template {} {}".format(tcell, col.value))
                app.logger.info("cell_template sheet row column {} {} {} {}".format(sheet, ws.max_row,ws.max_column,ws_template['G8'].value))

                merged_cell = {}
                for cell_rng in ws.merged_cell_ranges:  # move to 2.5.14 sheet.merged_cells.ranges
                    # print rng
                    startcell, endcell = cell_rng.split(':')
                    # print sheet.cell(startcell).border
                    start_xy = coordinate_to_tuple(startcell)
                    end_xy = coordinate_to_tuple(endcell)
                    start_row, start_col = start_xy
                    end_row, end_col = end_xy
                    for r in range(start_row, end_row + 1):
                        for c in range(start_col, end_col + 1):
                            cell = get_column_letter(c) + str(r)
                            merged_cell[cell] = {'cell_rng': cell_rng, 'start_cell': startcell}

                cell_added = []

                source_range_start, source_range_end = ws_template.dimensions.split(':')
                ws_start_xy = coordinate_to_tuple(source_range_start)
                ws_end_xy = coordinate_to_tuple(source_range_end)
                end_of_sheet = False
                for row in range(1,ws_end_xy[0]+1):
                    if end_of_sheet:
                        break
                    for col in range(1,ws_end_xy[1]+1):
                        cell_value = ws.cell(row=row,column=col).value
                        template_cell_value = ws_template.cell(row=row,column=col).value
                        column = get_column_letter(col)
                        app.logger.info("Processing cell [{}] [{}]".format((sheet,row,column),cell_value))
                        if template_cell_value and template_cell_value == "END-OF-TECHNICAL-SHEET-DATA-FOR-INDIVIDUALSHEET":
                            end_of_sheet = True
                            break
                        cell = str(column) + str(row)
                        is_merged_cell=False
                        if cell in merged_cell.keys():
                            cell_id = merged_cell[cell]['cell_rng']
                            start_cell=merged_cell[cell]['start_cell']
                            is_merged_cell = True
                        else:
                            cell_id = cell
                            start_cell=cell


                        try :
                            idx=cell_added.index(cell_id)
                        except ValueError:
                            content=cell_value
                            template_content = ws_template[start_cell].value
                            # # app.logger.info("start cell being processed [{}] [{}]".format(start_cell,content))
                            if not template_content and not content:
                                cell_style='{}'
                                cell_ref=json.dumps({'cell_ref': start_cell, 'section': None})
                            else:
                                cell_template=json.loads(ws_template[start_cell].value)
                                # app.logger.info("cell_template {}".format(cell_template))
                                cell_style=json.dumps(cell_template.get('cell_style'))
                                cell_style=cell_style if cell_style!='null' else None
                                cell_ref=cell_template.get('cell_ref')

                            ff_summary_df=ff_summary_df.append({'version':report_version_no,'report_id':report_id,
                                          'reporting_date':reporting_date,'sheet_id':sheet,
                                          'entity_type':'CELL','entity_id':cell_id,'col_id':str(column),'row_id':int(row),
                                          'cell_type':'MERGED' if is_merged_cell else 'SINGLE','cell_ref':cell_ref,
                                          'cell_style':cell_style,'content':content,},ignore_index=True)
                            cell_added.append(cell_id)


                end_of_sheet = False
                for row in range(1,ws_end_xy[0]+1):
                    row_height=0
                    # if end_of_sheet:
                    #     break
                    for col in range(1,ws_end_xy[1]+1):
                        app.logger.info("Row height processing....{}".format(row,col))
                        col_value = ws_template.cell(row=row,column=col).value
                        if col_value and col_value == "END-OF-TECHNICAL-SHEET-DATA-FOR-INDIVIDUALSHEET":
                            end_of_sheet = True
                            break
                        cell = get_column_letter(col) + str(row)
                        content=ws_template[cell].value
                        if content:
                            rh=float(json.loads(content)['row_height'])
                            if rh > row_height:
                                row_height=rh

                    if end_of_sheet:
                        break

                    row_height=row_height if row_height else 25
                    ff_summary_df = ff_summary_df.append({'version': report_version_no, 'report_id': report_id, 'reporting_date':reporting_date,'sheet_id': sheet,
                         'entity_type': 'ROW', 'entity_id': str(row), 'col_id': None,'row_id': None,
                         'cell_type': 'ROW_HEIGHT','cell_ref':None,'cell_style':None, 'content': row_height, }, ignore_index=True)

                for col in range(1,ws_end_xy[1]+1):
                    col_width=0
                    for row in range(1,ws_end_xy[0]+1):
                        cell = get_column_letter(col) + str(row)
                        content=ws_template[cell].value
                        if content and content == "END-OF-TECHNICAL-SHEET-DATA-FOR-INDIVIDUALSHEET":
                            break
                        if content:
                            cw=float(json.loads(content)['col_width'])
                            if cw > col_width:
                                col_width=cw

                    col_width=col_width if col_width else 90/8
                    ff_summary_df = ff_summary_df.append({'version': report_version_no, 'report_id': report_id, 'sheet_id': sheet,
                         'reporting_date':reporting_date,'entity_type': 'COL', 'entity_id': str(col), 'col_id': None, 'row_id': None,
                         'cell_type': 'COLUMN_WIDTH', 'cell_ref': None, 'cell_style': None, 'content': col_width, },ignore_index=True)

            self.db.transactmany("insert into {0}({1}) values({2})".format(ff_summary_object, ",".join(ff_summary_df.columns),
                                ",".join(['%s'] * len(ff_summary_df.columns))),list(ff_summary_df.itertuples(index=False, name=None)))

            self.db.commit


        except Exception as e:
            raise (e)

    def create_report_snapshot(self,**report_parameters):
        try:
            app.logger.info("Creating report snapshot for {}".format(report_parameters['report_id']))
            report_snapshot={}
            data_snapshot=self.create_data_snapshot(**report_parameters)
            fixed_format_snapshot=self.create_fixed_format_snapshot(**report_parameters)
            report_snapshot.update(data_snapshot)
            report_snapshot.update(fixed_format_snapshot)
            return json.dumps(report_snapshot)
        except Exception as e:
            raise(e)

    def create_data_snapshot(self,**report_parameters):
        try:
            report_id = report_parameters["report_id"]
            business_date_from = report_parameters["business_date_from"]
            business_date_to = report_parameters["business_date_to"]

            # Union with these sources from transaction report
            all_sources = self.db.query('select distinct source_id from report_calc_def where report_id=%s and in_use=\'Y\' \
                                        and source_id !=0',(report_id,)).fetchall()
            # Union with these sources from transaction report

            qualified_data_version = defaultdict(dict)

            for source in all_sources:
                qdvers = self.db.query("select a.business_date,a.source_id,a.version from qualified_data_vers a,\
                                  (select business_date,source_id,max(version) version from qualified_data_vers where \
                                   business_date between %s and %s and source_id=%s group by business_date,source_id) b \
                                   where a.business_date=b.business_date and a.source_id=b.source_id and a.version=b.version",
                                   (business_date_from, business_date_to, source['source_id'])).fetchall()
                for vers in qdvers:
                    qualified_data_version[str(source['source_id'])][vers['business_date']] = vers['version']


            return {'qualified_data':qualified_data_version}

        except Exception as e:
            app.logger.error(str(e))
            raise(e)

    def create_fixed_format_snapshot(self,**report_paramteres):
        try:
            report_id = report_paramteres["report_id"]

            cardf = pd.DataFrame(self.db.query("select id,report_id,sheet_id,cell_id,comp_agg_ref,comp_agg_rule,reporting_scale,rounding_option\
                                     from report_comp_agg_def WHERE report_id=%s AND in_use='Y'",
                                               (report_id,)).fetchall())
            car_version = self.db.query("select report_id,version,id_list from report_comp_agg_def_vers where report_id=%s\
                                  and version=(select max(version) from report_comp_agg_def_vers where report_id=%s)",
                                        (report_id, report_id)).fetchone()
            if len(cardf.index) == 0:
                car_id_list = []
            else:
                cardf['id'] = cardf['id'].astype(dtype='int64', errors='ignore')
                car_id_list = list(map(int, cardf['id'].tolist()))
            car_id_list.sort()
            car_id_list_str = ",".join(map(str, car_id_list))

            if not car_version:
                car_version_no = 1
            else:
                old_id_list = list(map(int, car_version['id_list'].split(',')))
                car_version_no = car_version['version'] + 1 if set(car_id_list) != set(old_id_list) else car_version[
                    'version']

            if not car_version or car_version_no != car_version['version']:
                self.db.transact("insert into report_comp_agg_def_vers(report_id,version,id_list) values(%s,%s,%s)",
                                 (report_id, car_version_no, car_id_list_str))
                self.db.commit()

            all_sources = self.db.query('select distinct source_id from report_calc_def where report_id=%s and in_use=\'Y\' \
                                        and source_id !=0',(report_id,)).fetchall()
            report_rule_version={}
            for source in all_sources:
                all_business_rules_df = pd.DataFrame(self.db.query("select id,report_id,sheet_id,cell_id,cell_calc_ref,\
                                        cell_business_rules from report_calc_def where report_id=%s and source_id=%s and in_use='Y'",
                                        (report_id, source['source_id'],)).fetchall())

                rr_version = self.db.query("select version,id_list from report_calc_def_vers where source_id=%s and report_id=%s \
                                          and version=(select max(version) version from report_calc_def_vers where source_id=%s \
                                          and report_id=%s)",(source['source_id'], report_id, source['source_id'], report_id)).fetchone()
                all_business_rules_df['id'] = all_business_rules_df['id'].astype(dtype='int64', errors='ignore')
                # print(all_business_rules.dtypes)
                rr_id_list = list(map(int, all_business_rules_df['id'].tolist()))
                rr_id_list.sort()
                rr_id_list_str = ",".join(map(str, rr_id_list))

                if not rr_version:
                    rr_version_no = 1
                else:
                    old_id_list = list(map(int, rr_version['id_list'].split(',')))
                    rr_version_no = rr_version['version'] + 1 if set(rr_id_list) != set(old_id_list) else rr_version[
                        'version']

                if not rr_version or rr_version_no != rr_version['version']:
                    self.db.transact(
                        "insert into report_calc_def_vers(report_id,source_id,version,id_list) values(%s,%s,%s,%s)",
                        (report_id, source['source_id'], rr_version_no, rr_id_list_str))
                    self.db.commit()
                report_rule_version[str(source['source_id'])] = rr_version_no

            return {"report_calc_def":report_rule_version,"report_comp_agg_def":car_version_no}

        except Exception as e:
            app.logger.error(str(e))
            raise(e)


    def get_report_list(self,country='ALL'):
        try:
            report_list=self.db.query("select distinct report_id from report_def_catalog where country='"+country+"'").fetchall()
            return report_list
        except Exception as e:
            app.logger.error(str(e))
            return {'msg': str(e)}, 500

    def get_country_list(self):
        try:
            country_list=self.db.query("select distinct country from report_def_catalog").fetchall()
            return country_list
        except Exception as e:
            app.logger.error(str(e))
            return {'msg': str(e)}, 500

    def create_report_catalog(self,report_parameters):
        try:
            report_id = report_parameters["report_id"]
            reporting_date = report_parameters["reporting_date"]
            report_parameters_str=json.dumps(report_parameters)
            report_create_status = 'CREATE'
            report_create_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            as_of_reporting_date = report_parameters["as_of_reporting_date"]

            report_version=self.db.query("select max(version) version from report_catalog where report_id=%s and reporting_date=%s",
                                         (report_id,reporting_date)).fetchone()
            report_version_no=1 if not report_version['version'] else  report_version['version']+1

            sql="insert into report_catalog(report_id,reporting_date,report_create_date," + \
                "report_parameters,report_create_status,as_of_reporting_date,version,report_created_by)" + \
                " values(%s,%s,%s,%s,%s,%s,%s,%s)"
            catalog_id = self.db.transact(sql,(report_id,reporting_date,report_create_date,report_parameters_str,report_create_status,
            as_of_reporting_date,report_version_no,self.user_id))
            self.log_master_id = self.opsLog.write_log_master(operation_type='Create Report'
                , operation_status = 'RUNNING'
                , operation_narration = 'Create report {0} for {1} as on {2}'.format(report_id,reporting_date,as_of_reporting_date)
                , entity_type = 'Report'
                , entity_name = report_id
                , entity_table_name = 'report_catalog'
                , entity_id = catalog_id
                )
            self.db.commit()
            if self.log_master_id:
                self.opsLog.write_log_detail(master_id=self.log_master_id
                    , operation_sub_type='Create catalog entry'
                    , operation_status='Complete'
                    , operation_narration='Report creation started with the parameters : {0}'.format(report_parameters_str,))
            return report_version_no
        except Exception as e:
            app.logger.error(str(e))
            if self.log_master_id:
                self.opsLog.write_log_detail(master_id=self.log_master_id
                    , operation_sub_type='Error occured while creating catalog'
                    , operation_status='Failed'
                    , operation_narration='Report creation Failed with error : {0}'.format(str(e),))
            raise(e)

    def update_report_catalog(self,status,report_parameters,report_snapshot=None,version=0):
        try:
            report_id = report_parameters["report_id"]
            reporting_date = report_parameters["reporting_date"]

            update_clause = "report_create_status='{0}'".format(status,)
            update_clause += ", report_create_date='{0}'".format(datetime.now().strftime('%Y-%m-%d %H:%M:%S'),)
            if report_parameters != None:
                report_parameters_str=json.dumps(report_parameters)
                update_clause += ", report_parameters='{0}'".format(report_parameters_str,)
            if report_snapshot !=None:
                update_clause +=", report_snapshot='{0}'".format(report_snapshot)
            sql = "update report_catalog set {0} where report_id='{1}' and reporting_date='{2}' and version={3}".format(update_clause,report_id,reporting_date,version)
            self.db.transact(sql)
            self.db.commit()
            if self.log_master_id:
                self.opsLog.write_log_detail(master_id=self.log_master_id
                    , operation_sub_type='Updated report catalog'
                    , operation_status='Complete'
                    , operation_narration='Report catalog updated with : {0}'.format(update_clause,))
        except Exception as e:
            app.logger.error(str(e))
            if self.log_master_id:
                self.opsLog.write_log_detail(master_id=self.log_master_id
                    , operation_sub_type='Error occured while updating report catalog'
                    , operation_status='Failed'
                    , operation_narration='Report creation Failed with error : {0}'.format(str(e),))
            raise(e)
