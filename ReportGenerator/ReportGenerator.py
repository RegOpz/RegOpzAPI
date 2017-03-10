#!/usr/bin/python3
# -*- coding: utf-8 -*-
import openpyxl as xls
import utils as util
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Protection
from openpyxl.utils import get_column_letter
import json
from Helpers.DatabaseHelper import *

class ReportGenerator(object):

    def __init__(self,report_id,reporting_date,business_date_from,business_date_to):
        self.report_id=report_id
        self.reporting_date=reporting_date
        self.bsuniess_date_from=business_date_from
        self.business_date_to=business_date_to

    def load_report_template(self,template_file_name):
        formula_dict = {'SUM': 'CALCULATE_FORMULA',
                        '=SUM': 'CALCULATE_FORMULA',
                        }
        cell_render_ref = None
        target_dir = './'
        wb = xls.load_workbook(target_dir + template_file_name)

        db = DatabaseHelper()

        for sheet in wb.worksheets:

            db.transact('delete from report_def where report_id=%s and sheet_id=%s', (self.report_id, sheet.title,))

            # First capture the dimensions of the cells of the sheet
            rowHeights = [sheet.row_dimensions[r + 1].height for r in range(sheet.max_row)]
            colWidths = [sheet.column_dimensions[get_column_letter(c + 1)].width for c in range(sheet.max_column)]

            for row, height in enumerate(rowHeights):
                db.transact('insert into report_def(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref)\
                             values(%s,%s,%s,%s,%s)', (self.report_id, sheet.title, str(row + 1), 'ROW_HEIGHT', str(height)))

            for col, width in enumerate(colWidths):
                db.transact('insert into report_def(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref)\
                            values(%s,%s,%s,%s,%s)',
                            (self.report_id, sheet.title, get_column_letter(col + 1), 'COLUMN_WIDTH', str(width)))

            rng_startcell = []
            for rng in sheet.merged_cell_ranges:
                # print rng
                startcell, endcell = rng.split(':')
                # print sheet.cell(startcell).border
                rng_startcell.append(startcell)

                db.transact('insert into report_def(report_id,sheet_id,cell_id,cell_render_def,cell_calc_ref)\
                            values(%s,%s,%s,%s,%s)',
                            (self.report_id, sheet.title, rng, 'MERGED_CELL', sheet[startcell].value))

            for all_obj in sheet['A1':util.cell_index(sheet.max_column, sheet.max_row)]:
                for cell_obj in all_obj:
                    cell_ref = str(cell_obj.column) + str(cell_obj.row)
                    if len(rng_startcell) > 0 and cell_ref not in rng_startcell:
                        if cell_obj.value:
                            for key in formula_dict.keys():
                                cell_obj_value = str(cell_obj.value)
                                if key in cell_obj_value:
                                    cell_render_ref = formula_dict[key]
                                    break
                                else:
                                    cell_render_ref = 'STATIC_TEXT'

                            db.transact('insert into report_def(report_id,sheet_id ,cell_id ,cell_render_def ,cell_calc_ref)\
                                      values(%s,%s,%s,%s,%s)',
                                        (self.report_id, sheet.title, cell_ref, cell_render_ref, cell_obj_value.strip()))
        db.commit()
        return 0

    def render_report_template(self):

        target_file_name = self.report_id + '_' + self.reporting_date + '.xlsx'
        wr = xls.Workbook()
        # target_dir='../output/'
        target_dir = './'

        db = DatabaseHelper()

        cur = db.query('select distinct sheet_id from report_def where report_id=%s', (self.report_id,))
        sheets = cur.fetchall()

        # The default sheet of the workbook
        al = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
        ws = wr.worksheets[0]
        for sheet in sheets:
            # The first sheet title will be Sheet, so do not create any sheet, just rename the title
            if ws.title == 'Sheet':
                ws.title = sheet['sheet_id']
            else:
                ws = wr.create_sheet(title=sheet["sheet_id"])

            cur = db.query(
                'select cell_id,cell_render_def,cell_calc_ref from report_def where report_id=%s and sheet_id=%s',
                (self.report_id, sheet["sheet_id"]))
            report_template = cur.fetchall()

            for row in report_template:
                if row['cell_render_def'] == 'MERGED_CELL':
                    ws.merge_cells(row["cell_id"])
                    startcell, endcell = row["cell_id"].split(':')
                    ws[startcell].value = row["cell_calc_ref"]
                    ws[startcell].value = row["cell_calc_ref"]
                    ws[startcell].fill = PatternFill("solid", fgColor="DDDDDD")
                    ws[startcell].font = Font(size=9)
                    ws[startcell].alignment = al
                elif row["cell_render_def"] == 'STATIC_TEXT':
                    ws[row["cell_id"]].value = row["cell_calc_ref"]
                    ws[row["cell_id"]].fill = PatternFill("solid", fgColor="DDDDDD")
                    ws[row["cell_id"]].font = Font(size=9)
                    if row["cell_calc_ref"][:1] != '=' and row["cell_calc_ref"][
                                                           len(row["cell_calc_ref"]) - 1:1] != '%' and 'SUM' not in row[
                        "cell_calc_ref"]:
                        ws[row["cell_id"]].alignment = al
                elif row["cell_render_def"] == 'CALCULATE_FORMULA':
                    ws[row["cell_id"]].fill = PatternFill("solid", fgColor="DDDDDD")
                    ws[row["cell_id"]].font = Font(bold=True, size=9)
                    if '=' in row["cell_calc_ref"]:
                        ws[row["cell_id"]].value = row["cell_calc_ref"]
                    else:
                        ws[row["cell_id"]].value = '=' + row["cell_calc_ref"]
                if row["cell_render_def"] == 'ROW_HEIGHT' and row["cell_calc_ref"] != 'None':
                    # print('row ' + row["cell_id"])
                    ws.row_dimensions[int(row["cell_id"])].height = float(row["cell_calc_ref"])
                elif row["cell_render_def"] == 'COLUMN_WIDTH' and row["cell_calc_ref"] != 'None':
                    # print(row["cell_id"])
                    ws.column_dimensions[row["cell_id"]].width = float(row["cell_calc_ref"])
                else:
                    pass

        wr.save(target_dir + target_file_name)
        return 0

    def render_report_template_json(self):

        db = DatabaseHelper()

        cur = db.query("select distinct sheet_id from report_def where report_id=%s ", (self.report_id,))
        sheets = cur.fetchall()
        print(sheets)

        sheet_d_list=[]
        for sheet in sheets:
            matrix_list = []
            cur = db.query(
                'select cell_id,cell_render_def,cell_calc_ref from report_def where report_id=%s and sheet_id=%s',
                (self.report_id, sheet["sheet_id"]))
            report_template = cur.fetchall()

            for row in report_template:
                cell_d = {}
                if row["cell_render_def"] == 'STATIC_TEXT':
                    cell_d['cell'] =row['cell_id']
                    cell_d['value']=row['cell_calc_ref']
                    matrix_list.append(cell_d)


                elif row['cell_render_def'] == 'MERGED_CELL':
                    start_cell, end_cell = row['cell_id'].split(':')
                    cell_d['cell'] = start_cell
                    cell_d['value'] = row['cell_calc_ref']
                    cell_d['merged'] = end_cell
                    matrix_list.append(cell_d)

            sheet_d={}
            sheet_d['sheet']=sheet['sheet_id']
            #print(sheet_d['sheet'])
            sheet_d['matrix']=matrix_list
            sheet_d_list.append(sheet_d)


        json_dump = json.dumps(sheet_d_list)
        #print(json_dump)

        return json_dump


if __name__=="__main__":

    report=ReportGenerator('MAS1003','20160930','20160901','20160930')

    #report.load_report_template('Functional_Specification_MAS10032.xlsx')

    #report.render_report_template()

    report.render_report_template_json()
